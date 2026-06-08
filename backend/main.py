import os
import db_hardener
import sqlite3
import json
import socket
import logging
from typing import List, Optional, Dict, Tuple, Any
import pandas as pd
from fastapi import FastAPI, UploadFile, File, HTTPException, Response, BackgroundTasks, Form
from fastapi.responses import FileResponse, RedirectResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from pathlib import Path
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Explicitly find root directory .env even if uvicorn is spawned inside /backend
root_dir = Path(__file__).resolve().parent.parent
root_env = root_dir / ".env"
if root_env.exists():
    load_dotenv(dotenv_path=root_env)
else:
    load_dotenv() # Fallback to local

# Setup unified logging
from logger import logger

# Import our modular processor, AI, and crypto modules
from crypto_vault import encrypt_field, decrypt_field, encrypt_raw_data, decrypt_raw_data
from credit_guard import init_credit_db, check_credit_limit
from dsm_tracker import init_dsm_db, save_dsm_shift, delete_dsm_shifts_by_date, calculate_dsm_expected_sales
from lube_sales import init_lube_db
from density_logger import init_density_db, save_density_record, get_density_records

app = FastAPI(
    title="Petrol Pump Daily Register OCR & Auditing Backend",
    version="1.0.0",
    description="Processes handwritten daily ledgers, runs image shadow cleaning, extracts data with Gemini, and validates arithmetic with Claude 3.5 Sonnet."
)

# Enable CORS for frontend integration
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Register Local Network Presentation Gateway router
from local_gateway import router as local_gateway_router
app.include_router(local_gateway_router)

# Register Customer Ledger WhatsApp Alerts router
from customer_alerts import router as customer_alerts_router
app.include_router(customer_alerts_router)

# Register Underground Tank Tilt Profile Adjuster router
from dip_profiler import router as dip_profiler_router
app.include_router(dip_profiler_router)

# Register Database Hardener & Optimization router
from db_hardener import router as db_hardener_router
app.include_router(db_hardener_router)


from fastapi.staticfiles import StaticFiles
upload_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "uploaded_raw_photos")
os.makedirs(upload_dir, exist_ok=True)
app.mount("/uploaded_raw_photos", StaticFiles(directory=upload_dir), name="uploaded_raw_photos")

# Mount charts directory
charts_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "pump_exports", "charts")
os.makedirs(charts_dir, exist_ok=True)
app.mount("/pump_exports/charts", StaticFiles(directory=charts_dir), name="charts")

# Mount processed split/straightened images directory
processed_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "processed_images")
os.makedirs(processed_dir, exist_ok=True)
app.mount("/processed_images", StaticFiles(directory=processed_dir), name="processed_images")

# Mount raw_backlog directory (Prompt 100)
raw_backlog_dir = os.path.abspath(os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "raw_backlog"))
os.makedirs(raw_backlog_dir, exist_ok=True)
app.mount("/raw_backlog", StaticFiles(directory=raw_backlog_dir), name="raw_backlog")

# Constants
DB_PATH = "ledger.db"

# Dynamically resolve default master Excel path from environment
_export_path = os.getenv("EXPORT_EXCEL_PATH")
if not _export_path:
    EXCEL_PATH = os.path.join(str(root_dir), "pump_exports", "Pump_Accounts.xlsx")
else:
    if not os.path.isabs(_export_path):
        EXCEL_PATH = os.path.abspath(os.path.join(str(root_dir), _export_path))
    else:
        EXCEL_PATH = _export_path

# --- Workspace Profile Management Setup ---
try:
    from workspace_manager import get_active_profile, rebind_all_modules
    active_profile = get_active_profile()
    rebind_all_modules(active_profile)
    
    # Re-route FastAPI static serving directories for the active workspace
    for route in app.routes:
        if getattr(route, "name", None) == "processed_images":
            route.app.directory = processed_dir
            route.app.all_directories = [processed_dir]
        elif getattr(route, "name", None) == "charts":
            route.app.directory = charts_dir
            route.app.all_directories = [charts_dir]
except Exception as ws_err:
    logger.error(f"Failed to initialize workspace profile: {ws_err}")

# Initialize SQLite database
def init_db():
    logger.info("Initializing SQLite database...")
    try:
        from self_healer import auto_heal_if_corrupted
        auto_heal_if_corrupted(DB_PATH)
    except Exception as heal_err:
        logger.error(f"Failed to execute self-healer check: {str(heal_err)}")

    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS daily_ledger (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT UNIQUE,
            total_sales_liters REAL,
            total_amount_inr REAL,
            cash_tender REAL,
            upi_tender REAL,
            paytm_transfers REAL,
            card_tender REAL,
            udhaar_sales REAL,
            expenses_amount REAL,
            validation_status TEXT,
            raw_data TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """)
        conn.commit()
        conn.close()
        logger.info("Database initialized successfully.")
        
        # Initialize fuel_rates table
        try:
            from price_registry import init_rates_db
            init_rates_db()
        except Exception as rates_err:
            logger.warning(f"Failed to auto-initialize fuel_rates table: {str(rates_err)}")
            
        # Initialize stock_recon table
        try:
            from reconciliation import init_recon_db
            init_recon_db(DB_PATH)
        except Exception as recon_err:
            logger.warning(f"Failed to auto-initialize stock_recon table: {str(recon_err)}")
            
        # Initialize credit_limits table
        try:
            init_credit_db(DB_PATH)
        except Exception as credit_err:
            logger.warning(f"Failed to auto-initialize credit_limits table: {str(credit_err)}")
            
        # Initialize dsm_shifts table
        try:
            init_dsm_db(DB_PATH)
        except Exception as dsm_err:
            logger.warning(f"Failed to auto-initialize dsm_shifts table: {str(dsm_err)}")
            
        # Initialize inventory_sales table
        try:
            init_lube_db(DB_PATH)
        except Exception as lube_err:
            logger.warning(f"Failed to auto-initialize inventory_sales table: {str(lube_err)}")
            
        # Initialize card_settlements table
        try:
            from card_settlements import init_card_settlements_db
            init_card_settlements_db(DB_PATH)
        except Exception as card_err:
            logger.warning(f"Failed to auto-initialize card_settlements table: {str(card_err)}")

        # Initialize density_register table
        try:
            init_density_db(DB_PATH)
        except Exception as density_err:
            logger.warning(f"Failed to auto-initialize density_register table: {str(density_err)}")

        # Initialize bank_matcher tables (bank_statement_credits + digital_settlement_status)
        try:
            from bank_matcher import init_bank_matcher_db
            init_bank_matcher_db(DB_PATH)
        except Exception as bm_err:
            logger.warning(f"Failed to auto-initialize bank_matcher tables: {str(bm_err)}")

        # Initialize cash_denominations table
        try:
            from cash_denominations import init_cash_denominations_db
            init_cash_denominations_db(DB_PATH)
        except Exception as cash_denom_err:
            logger.warning(f"Failed to auto-initialize cash_denominations table: {str(cash_denom_err)}")

        # Initialize internal_consumption table
        try:
            from internal_consumption import init_internal_consumption_db
            init_internal_consumption_db(DB_PATH)
        except Exception as ic_err:
            logger.warning(f"Failed to auto-initialize internal_consumption table: {str(ic_err)}")

        # Initialize nozzle_testing_logs table
        try:
            from testing_rts import init_testing_rts_db
            init_testing_rts_db(DB_PATH)
        except Exception as rts_err:
            logger.warning(f"Failed to auto-initialize nozzle_testing_logs table: {str(rts_err)}")

        # Initialize staff_advances table
        try:
            from staff_ledger import init_staff_ledger_db
            init_staff_ledger_db(DB_PATH)
        except Exception as staff_err:
            logger.warning(f"Failed to auto-initialize staff_advances table: {str(staff_err)}")

        # Initialize fleet_card tables
        try:
            from fleet_cards import init_fleet_cards_db
            init_fleet_cards_db(DB_PATH)
        except Exception as fleet_err:
            logger.warning(f"Failed to auto-initialize fleet_cards table: {str(fleet_err)}")

        # Initialize decanting tanker_receipts table
        try:
            from decanting_auditor import init_decanting_db
            init_decanting_db(DB_PATH)
        except Exception as decant_err:
            logger.warning(f"Failed to auto-initialize tanker_receipts table: {str(decant_err)}")

        # Run database schema migrations to keep database future-proof
        try:
            from migrations import apply_schema_updates
            apply_schema_updates(DB_PATH)
        except Exception as migration_err:
            logger.error(f"Failed to apply database migrations at startup: {str(migration_err)}")
            raise migration_err
    except Exception as e:
        logger.error(f"Failed to initialize database: {str(e)}")
        raise e

init_db()

# Initialise API cost accounting table
try:
    from cost_tracker import init_cost_db
    init_cost_db(DB_PATH)
except Exception as cost_err:
    logger.warning(f"Failed to auto-initialize cost tracker: {str(cost_err)}")

# Pre-render analytics trend charts on startup
try:
    from local_analytics import generate_all_charts
    generate_all_charts(DB_PATH)
except Exception as charts_init_err:
    logger.warning(f"Failed to pre-render trend charts on startup: {str(charts_init_err)}")


@app.on_event("startup")
def startup_backup_trigger():
    logger.info("Triggering automatic startup backup sequence...")
    try:
        from backup import execute_local_backup
        execute_local_backup()
    except Exception as e:
        logger.warning(f"FastAPI Startup backup failed: {str(e)}")
        
    logger.info("Retrying any pending automated backups on startup...")
    try:
        import threading
        from backup_dispatcher import retry_pending_backups
        # Run in a background thread to prevent startup block
        threading.Thread(target=retry_pending_backups, args=(DB_PATH,), daemon=True).start()
    except Exception as e:
        logger.warning(f"FastAPI Startup pending backups retry trigger failed: {str(e)}")
        
    logger.info("Triggering automatic startup database optimization...")
    try:
        from optimize import optimize_database
        optimize_database(DB_PATH)
    except Exception as e:
        logger.warning(f"FastAPI Startup database optimization failed: {str(e)}")

    # --- Crash recovery: reset any stuck PROCESSING jobs back to PENDING ---
    try:
        from state_tracker import reset_stuck_jobs, init_state_db
        init_state_db(DB_PATH)
        count = reset_stuck_jobs(DB_PATH)
        if count > 0:
            logger.warning(
                f"Startup recovery: reset {count} stuck batch job(s) "
                "from PROCESSING → PENDING."
            )
    except Exception as e:
        logger.warning(f"Startup crash-recovery failed: {str(e)}")

    # Start USB synchronization background mirroring service
    try:
        import asyncio
        from usb_sync import set_main_loop, start_usb_sync_service
        loop = asyncio.get_running_loop()
        set_main_loop(loop)
        start_usb_sync_service(interval_sec=5)
    except Exception as usb_init_err:
        logger.warning(f"Failed to start USB sync background mirroring service: {str(usb_init_err)}")

    # --- LAN IP Detection + QR Code Banner ---
    _print_network_qr()


def _get_lan_ip() -> str:
    """
    Detects the machine's local network IP address by opening a UDP socket
    towards a public address (no actual packet is sent).  Falls back to
    127.0.0.1 if no network interface is active.
    """
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.settimeout(0)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return "127.0.0.1"


def _print_network_qr(port: int = 8000) -> None:
    """
    Prints a clean ASCII QR code for the local Wi-Fi network URL so any
    phone on the same network can instantly open the mobile capture screen.
    """
    lan_ip = _get_lan_ip()
    mobile_url = f"http://{lan_ip}:{port}/mobile"

    border = "═" * 54
    print("")
    print(f"  ╔{border}╗")
    print(f"  ║{'  📡  LOCAL WI-FI MOBILE ACCESS':^54}║")
    print(f"  ╠{border}╣")
    print(f"  ║  Backend API : http://{lan_ip}:{port}{'':>{54 - 28 - len(lan_ip)}}║")
    print(f"  ║  Mobile URL  : {mobile_url:<38}║")
    print(f"  ╠{border}╣")
    print(f"  ║{'  Scan with your phone camera to open capture UI':^54}║")
    print(f"  ╚{border}╝")
    print("")

    try:
        import qrcode
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_M,
            box_size=1,
            border=1,
        )
        qr.add_data(mobile_url)
        qr.make(fit=True)
        # Print using unicode half-blocks for a compact terminal QR code
        matrix = qr.get_matrix()
        for row_idx in range(0, len(matrix), 2):
            line = "  "
            for col in range(len(matrix[row_idx])):
                top = matrix[row_idx][col]
                bot = matrix[row_idx + 1][col] if (row_idx + 1) < len(matrix) else False
                if top and bot:
                    line += "█"
                elif top:
                    line += "▀"
                elif bot:
                    line += "▄"
                else:
                    line += " "
            print(line)
        print("")
        logger.info(f"Mobile capture URL: {mobile_url}")
    except ImportError:
        logger.warning("qrcode library not installed — skipping QR render. Run: pip install qrcode")


@app.on_event("shutdown")
def shutdown_backup_trigger():
    logger.info("Triggering automatic shutdown backup sequence...")
    try:
        from backup import execute_local_backup
        execute_local_backup()
    except Exception as e:
        logger.warning(f"FastAPI Shutdown backup failed: {str(e)}")

    logger.info("Triggering database WAL checkpoint on shutdown...")
    try:
        from db_hardener import execute_wal_checkpoint
        execute_wal_checkpoint()
    except Exception as e:
        logger.warning(f"Database WAL checkpoint on shutdown failed: {str(e)}")

    logger.info("Triggering database storage optimization on shutdown...")
    try:
        from db_vacuum import execute_db_vacuum_background
        execute_db_vacuum_background(DB_PATH)
    except Exception as e:
        logger.warning(f"Database storage optimization on shutdown failed: {str(e)}")

# Pydantic schemas for verification
class SaveEntryRequest(BaseModel):
    date: str
    nozzles: list
    hinglish_notes: list
    total_sales_liters: float
    total_amount_inr: float
    validation_status: str
    audit_explanation: str
    cash_tender: float = 0.0
    upi_tender: float = 0.0
    paytm_transfers: float = 0.0
    card_tender: float = 0.0
    udhaar_sales: float = 0.0
    expenses_amount: float = 0.0

# Helper to style the Excel workbook
def apply_excel_styling(path: str):
    logger.info(f"Applying styling and number formatting to {path}...")
    try:
        wb = load_workbook(path)
        
        # Color & Typography Tokens
        header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Elegant Navy Blue
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        regular_font = Font(name="Segoe UI", size=11)
        
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Explicitly enable gridlines (often disabled by default in Python-generated sheets)
            ws.views.sheetView[0].showGridLines = True
            
            # 1. Format Headers
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border
            
            # Set header row height to 28pt for comfortable padding
            ws.row_dimensions[1].height = 28
            
            # 2. Format Data Rows
            for row in range(2, ws.max_row + 1):
                ws.row_dimensions[row].height = 20  # Row height 20pt for premium spacing
                
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.font = regular_font
                    cell.border = thin_border
                    
                    col_name = str(ws.cell(row=1, column=col).value)
                    
                    # Align and format based on values and column headers
                    if isinstance(cell.value, (int, float)):
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                        
                        # INR Currency formatting
                        if "Amount" in col_name or "INR" in col_name or "Rate" in col_name:
                            cell.number_format = "[$₹-4009] #,##0.00"  # Format as ₹ with decimals
                        # Liters/Decimals formatting
                        else:
                            cell.number_format = "#,##0.00"
                    else:
                        # Center status and dates
                        if "Date" in col_name or "Status" in col_name or "Type" in col_name:
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        else:
                            cell.alignment = Alignment(horizontal="left", vertical="center")
            
            # 3. Auto-fit columns with safety margin
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    val_str = str(cell.value or '')
                    if len(val_str) > max_len:
                        max_len = len(val_str)
                # Cap minimum column width at 14 to avoid title compression
                ws.column_dimensions[col_letter].width = max(max_len + 4, 14)
                
        wb.save(path)
        logger.info("Excel styling applied successfully.")
    except Exception as e:
        logger.warning(f"Failed to style Excel sheet: {str(e)}")

# Endpoint to test if backend is live
@app.get("/")
def read_root():
    return {
        "status": "online",
        "service": "Petrol Pump OCR Backend",
        "primary_model": "gemini-1.5-flash",
        "validation_model": "claude-3-5-sonnet-20241022"
    }


# Endpoint 2: Save finalized JSON ledger
@app.post("/api/save-entry")
def save_entry(entry: SaveEntryRequest):
    """
    Saves a verified daily entry directly to local SQLite and appends rows to a clean Excel ledger.
    Using 'INSERT OR REPLACE' on SQLite date to allow safe updates.
    """
    logger.info(f"Saving ledger entry for date: {entry.date}")
    
    # 1. Save to SQLite
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # Store raw payload as JSON string in database for historical records
        raw_json_str = json.dumps(entry.dict(), ensure_ascii=False)
        
        cursor.execute("""
            INSERT OR REPLACE INTO daily_ledger 
            (date, total_sales_liters, total_amount_inr, cash_tender, upi_tender, paytm_transfers, card_tender, udhaar_sales, expenses_amount, validation_status, raw_data)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            entry.date, 
            entry.total_sales_liters, 
            entry.total_amount_inr, 
            entry.cash_tender,
            entry.upi_tender,
            entry.paytm_transfers,
            entry.card_tender,
            entry.udhaar_sales,
            entry.expenses_amount,
            entry.validation_status, 
            raw_json_str
        ))
        conn.commit()
        conn.close()
        logger.info("SQLite insert successful.")
    except Exception as e:
        logger.error(f"SQLite save failed: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Database Save Error: {str(e)}")
        
    # 2. Append to clean Excel workbook
    try:
        # Generate new DataFrames for appending
        summary_row = {
            "Date": [entry.date],
            "Total Sales (Liters)": [entry.total_sales_liters],
            "Total Amount (INR)": [entry.total_amount_inr],
            "Cash Tender (INR)": [entry.cash_tender],
            "UPI Tender (INR)": [entry.upi_tender],
            "Paytm Transfers (INR)": [entry.paytm_transfers],
            "Card Tender (INR)": [entry.card_tender],
            "Udhaar Sales (INR)": [entry.udhaar_sales],
            "Expenses (INR)": [entry.expenses_amount],
            "Validation Status": [entry.validation_status],
            "Audit Notes Summary": [entry.audit_explanation],
            "Hinglish Notes": [", ".join(entry.hinglish_notes)]
        }
        df_summary_new = pd.DataFrame(summary_row)
        
        nozzle_rows = []
        for nozzle in entry.nozzles:
            nozzle_rows.append({
                "Date": entry.date,
                "Nozzle Name": nozzle.get("nozzle_name"),
                "Fuel Type": nozzle.get("fuel_type"),
                "Opening Meter": nozzle.get("opening"),
                "Closing Meter": nozzle.get("closing"),
                "Sales (Liters)": nozzle.get("sales_liters_calculated"),
                "Rate (Price/Ltr)": nozzle.get("rate"),
                "Amount Calculated (INR)": nozzle.get("amount_calculated"),
                "Math Validated": "Yes" if nozzle.get("arithmetic_valid") else "No",
                "Discrepancy Notes": nozzle.get("discrepancy_details") or "None"
            })
        df_nozzles_new = pd.DataFrame(nozzle_rows)
        
        # Read or create Excel structure
        if not os.path.exists(EXCEL_PATH):
            # Create fresh workbook
            with pd.ExcelWriter(EXCEL_PATH, engine='openpyxl') as writer:
                df_summary_new.to_excel(writer, sheet_name="Daily Summary", index=False)
                df_nozzles_new.to_excel(writer, sheet_name="Nozzle Readings", index=False)
        else:
            # Overwrite or append while removing previous records for that date (avoids duplicates)
            with pd.ExcelWriter(EXCEL_PATH, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                # Append Summary
                try:
                    df_summary_existing = pd.read_excel(EXCEL_PATH, sheet_name="Daily Summary")
                    df_summary_existing = df_summary_existing[df_summary_existing["Date"] != entry.date]
                    df_summary_combined = pd.concat([df_summary_existing, df_summary_new], ignore_index=True)
                except Exception:
                    df_summary_combined = df_summary_new
                    
                # Append Nozzles
                try:
                    df_nozzles_existing = pd.read_excel(EXCEL_PATH, sheet_name="Nozzle Readings")
                    df_nozzles_existing = df_nozzles_existing[df_nozzles_existing["Date"] != entry.date]
                    df_nozzles_combined = pd.concat([df_nozzles_existing, df_nozzles_new], ignore_index=True)
                except Exception:
                    df_nozzles_combined = df_nozzles_new
                
                # Overwrite sheets safely
                df_summary_combined.to_excel(writer, sheet_name="Daily Summary", index=False)
                df_nozzles_combined.to_excel(writer, sheet_name="Nozzle Readings", index=False)
                
        # Apply visual aesthetics to excel output
        apply_excel_styling(EXCEL_PATH)
        logger.info("Excel write successful.")
        
    except Exception as e:
        logger.error(f"Excel generation failed: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Excel Export Error: {str(e)}")
        
    return {
        "status": "success",
        "message": f"Ledger entry for {entry.date} saved successfully.",
        "sqlite_db": DB_PATH,
        "excel_file": EXCEL_PATH
    }

# Endpoint: High-speed FTS Search Route (Prompt 70)
@app.get("/api/search/archive")
def search_archive(q: str):
    """
    Accepts query parameter 'q' to search unstructured OCR text using SQLite FTS5 MATCH.
    """
    from text_search import search_ledger_fts
    try:
        matching_dates = search_ledger_fts(DB_PATH, q)
        return {"dates": matching_dates}
    except Exception as e:
        logger.error(f"FTS search_archive endpoint failed: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Search execution failed: {str(e)}")

# Endpoint 3: Fetch historical log (limit to 400 entries)
@app.get("/api/historical-log")
def get_historical_log():
    """
    Fetches the last 400 daily entries from SQLite database, joining daily_summary with daily_ledger.
    Sorted newest first.
    """
    logger.info("Fetching joined daily_summary & daily_ledger historical log entries...")
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # Pull directly from both daily_summary and daily_ledger to support HistoryView columns
        cursor.execute("""
            SELECT 
                s.date, 
                s.total_hsd_liters, 
                s.total_ms_liters, 
                s.total_cash_calculated, 
                s.is_verified, 
                l.raw_data 
            FROM daily_summary s
            LEFT JOIN daily_ledger l ON s.date = l.date
            ORDER BY s.date DESC
            LIMIT 400
        """)
        rows = cursor.fetchall()
        conn.close()
        
        log_entries = []
        for row in rows:
            date_str = row[0]
            total_hsd = float(row[1] or 0.0)
            total_ms = float(row[2] or 0.0)
            total_cash = float(row[3] or 0.0)
            is_verified = int(row[4] or 0)
            raw_data = json.loads(row[5]) if row[5] else {}
            raw_data = decrypt_raw_data(raw_data)
            
            # Formulate backward-compatible UI dictionary
            log_entries.append({
                "date": date_str,
                "total_hsd_liters": total_hsd,
                "total_ms_liters": total_ms,
                "total_cash_calculated": total_cash,
                "is_verified": is_verified,
                "total_sales_liters": total_hsd + total_ms,
                "total_amount_inr": total_cash,
                "cash_tender": raw_data.get("cash_tender") or max(0.0, total_cash - raw_data.get("udhaar_sales", 0.0) - raw_data.get("expenses_amount", 0.0)),
                "upi_tender": raw_data.get("upi_tender") or 0.0,
                "paytm_transfers": raw_data.get("paytm_transfers") or 0.0,
                "card_tender": raw_data.get("card_tender") or 0.0,
                "udhaar_sales": raw_data.get("udhaar_sales") or 0.0,
                "expenses_amount": raw_data.get("expenses_amount") or 0.0,
                "validation_status": "valid" if is_verified == 1 else "needs_review",
                "raw_data": raw_data
            })
        return log_entries
    except Exception as e:
        logger.error(f"Failed to fetch historical log: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Database Read Error: {str(e)}")

# Chitti Slip Matcher Endpoint: Upload and parse paper credit slips
@app.post("/api/slips/upload")
async def upload_credit_slips(
    date: str,
    files: List[UploadFile] = File(...)
):
    """
    Accepts multiple paper credit slip images, extracts their text via Gemini Vision,
    saves them in the database, and executes double-entry cross-examination matching.
    """
    logger.info(f"Received {len(files)} credit slip uploads for date: {date}")
    
    extracted_slips = []
    
    for upload_file in files:
        try:
            image_bytes = await upload_file.read()
            
            # Execute Gemini vision parsing
            from slip_matcher import extract_credit_slips_from_image
            slips = extract_credit_slips_from_image(image_bytes)
            if isinstance(slips, list):
                extracted_slips.extend(slips)
            elif isinstance(slips, dict):
                extracted_slips.append(slips)
        except Exception as err:
            logger.error(f"Failed to process credit slip image '{upload_file.filename}': {err}")
            # Do not fail entire batch, process others
            
    if not extracted_slips:
        raise HTTPException(status_code=400, detail="No valid slips could be extracted from the uploaded files.")
        
    try:
        # Save slips to DB
        from slip_matcher import save_extracted_slips, cross_reference_slips_to_ledger
        save_extracted_slips(extracted_slips, date, DB_PATH)
        
        # Execute cross-examination matching
        report = cross_reference_slips_to_ledger(date, DB_PATH)
        return report
    except Exception as e:
        logger.error(f"Failed to process and cross-reference uploaded slips: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

# Chitti Slip Matcher Endpoint: Retrieve matching report for date
@app.get("/api/slips/cross-reference")
def get_slips_cross_reference(date: str):
    """
    Returns the slip-to-ledger matching status array and discrepancy report for target date.
    """
    logger.info(f"Retrieving slips cross-reference report for date: {date}")
    try:
        from slip_matcher import cross_reference_slips_to_ledger
        report = cross_reference_slips_to_ledger(date, DB_PATH)
        return report
    except Exception as e:
        logger.error(f"Failed to fetch slips report: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

class CashDenominationRequest(BaseModel):
    date: str
    notes_500: int = 0
    notes_200: int = 0
    notes_100: int = 0
    notes_50: int = 0
    notes_20: int = 0
    notes_10: int = 0
    coins_total: float = 0.0

@app.post("/api/cash-denominations")
def post_cash_denominations(payload: CashDenominationRequest):
    """
    Verifies physical currency counts, matches against expected book sales,
    and commits the discrepancy delta.
    """
    logger.info(f"Received cash denominations submission for date: {payload.date}")
    try:
        from cash_denominations import verify_cash_vault_balance
        report = verify_cash_vault_balance(
            payload.date,
            {
                "notes_500": payload.notes_500,
                "notes_200": payload.notes_200,
                "notes_100": payload.notes_100,
                "notes_50": payload.notes_50,
                "notes_20": payload.notes_20,
                "notes_10": payload.notes_10,
                "coins_total": payload.coins_total
            },
            DB_PATH
        )
        return report
    except Exception as err:
        logger.error(f"Failed to verify and save cash denominations: {err}")
        raise HTTPException(status_code=500, detail=str(err))

@app.get("/api/cash-denominations")
def get_cash_denominations(date: str):
    """
    Retrieves the saved cash denomination breakdown and mismatch details for the target date.
    """
    logger.info(f"Retrieving cash denominations report for date: {date}")
    try:
        from cash_denominations import get_cash_denomination
        report = get_cash_denomination(date, DB_PATH)
        return report
    except Exception as err:
        logger.error(f"Failed to fetch cash denominations report: {err}")
        raise HTTPException(status_code=500, detail=str(err))

class InternalConsumptionRequest(BaseModel):
    date: str
    product_type: str # 'HSD' or 'MS'
    liters_drawn: float
    purpose_head: str
    authorized_by: str

@app.post("/api/internal-consumption")
def post_internal_consumption(payload: InternalConsumptionRequest):
    """
    Records an internal operational fuel draw, logs the financial cost
    as an Operational Station Expense, and returns the result.
    """
    logger.info(f"Recording internal fuel consumption draw: {payload.liters_drawn}L of {payload.product_type} for {payload.purpose_head}")
    try:
        from internal_consumption import record_internal_consumption
        result = record_internal_consumption(
            payload.date,
            payload.product_type,
            payload.liters_drawn,
            payload.purpose_head,
            payload.authorized_by,
            DB_PATH
        )
        return result
    except ValueError as val_err:
        raise HTTPException(status_code=400, detail=str(val_err))
    except Exception as err:
        logger.error(f"Failed to record internal consumption draw: {err}")
        raise HTTPException(status_code=500, detail=str(err))

@app.get("/api/internal-consumption")
def get_internal_consumption_logs(date: str):
    """
    Retrieves detailed internal draw logs recorded on the specified date.
    """
    logger.info(f"Retrieving internal consumption logs for date: {date}")
    try:
        from internal_consumption import get_daily_internal_consumption
        logs = get_daily_internal_consumption(date, DB_PATH)
        return logs
    except Exception as err:
        logger.error(f"Failed to fetch daily internal consumption logs: {err}")
        raise HTTPException(status_code=500, detail=str(err))

@app.get("/api/internal-consumption/summary")
def get_internal_consumption_summary():
    """
    Retrieves the monthly cumulative internal consumption summaries.
    """
    logger.info("Retrieving monthly cumulative internal consumption summaries")
    try:
        from internal_consumption import get_monthly_cumulative_consumption
        summary = get_monthly_cumulative_consumption(DB_PATH)
        return summary
    except Exception as err:
        logger.error(f"Failed to fetch internal consumption summaries: {err}")
        raise HTTPException(status_code=500, detail=str(err))

class NozzleTestingRequest(BaseModel):
    date: str
    nozzle_id: str
    product_type: str # 'HSD' or 'MS'
    testing_volume_liters: float = 5.0
    rts_verified: bool = True

@app.post("/api/nozzle-testing")
def post_nozzle_testing(payload: NozzleTestingRequest):
    """
    Records a nozzle testing calibration event and marks Return-to-Stock verification status.
    """
    logger.info(f"Recording nozzle testing draw: {payload.testing_volume_liters}L of {payload.product_type} on nozzle {payload.nozzle_id} (RTS: {payload.rts_verified})")
    try:
        from testing_rts import record_nozzle_testing
        result = record_nozzle_testing(
            payload.date,
            payload.nozzle_id,
            payload.product_type,
            payload.testing_volume_liters,
            payload.rts_verified,
            DB_PATH
        )
        return result
    except ValueError as val_err:
        raise HTTPException(status_code=400, detail=str(val_err))
    except Exception as err:
        logger.error(f"Failed to record nozzle testing draw: {err}")
        raise HTTPException(status_code=500, detail=str(err))

@app.get("/api/nozzle-testing")
def get_nozzle_testing_logs(date: str):
    """
    Retrieves detailed nozzle testing logs recorded on the specified date.
    """
    logger.info(f"Retrieving nozzle testing logs for date: {date}")
    try:
        from testing_rts import get_daily_testing_logs
        logs = get_daily_testing_logs(date, DB_PATH)
        return {"status": "success", "date": date, "logs": logs}
    except Exception as err:
        logger.error(f"Failed to fetch daily nozzle testing logs: {err}")
        raise HTTPException(status_code=500, detail=str(err))

# ---------------------------------------------------------------------------
# Party Name Auto-Complete Suggestion Endpoint
# ---------------------------------------------------------------------------

# Simple in-process suggestion cache to avoid re-decrypting the full corpus
# on every keystroke. TTL = 90 seconds. Invalidated by any ledger write.
_suggest_cache: dict = {"ts": 0.0, "rows": []}   # rows = list of (party_name, count)
_SUGGEST_CACHE_TTL = 90  # seconds

def _invalidate_suggest_cache():
    """Call after any ledger write to ensure fresh suggestions on next fetch."""
    global _suggest_cache
    _suggest_cache = {"ts": 0.0, "rows": []}


def _load_suggest_corpus(db_path: str) -> list[tuple[str, int]]:
    """
    Returns a frequency-ranked list of (plaintext_party_name, occurrence_count)
    for all rows of type 'udhaar' in ledger_entries.

    Result is cached in-process for _SUGGEST_CACHE_TTL seconds.
    """
    import time
    global _suggest_cache

    now = time.time()
    if (now - _suggest_cache["ts"]) < _SUGGEST_CACHE_TTL and _suggest_cache["rows"]:
        return _suggest_cache["rows"]

    try:
        conn = sqlite3.connect(db_path)
        # Fetch only the party_name column for udhaar entries (credit sales)
        # Also accept 'expense' type to surface common expense parties
        raw_rows = conn.execute(
            "SELECT party_name FROM ledger_entries WHERE type IN ('udhaar', 'expense')"
        ).fetchall()
        conn.close()
    except Exception as e:
        logger.error(f"suggest-parties: corpus load failed — {e}")
        return []

    # Decrypt and count
    from collections import Counter
    counter: Counter = Counter()
    for (enc_name,) in raw_rows:
        try:
            plain = decrypt_field(enc_name, return_type=str)
            if plain and plain.strip():
                counter[plain.strip()] += 1
        except Exception:
            pass  # skip rows that fail decryption silently

    # Sort by frequency descending, then alphabetically for ties
    ranked = sorted(counter.items(), key=lambda x: (-x[1], x[0].lower()))

    _suggest_cache = {"ts": now, "rows": ranked}
    return ranked


@app.get("/api/ledger/suggest-parties")
def api_suggest_parties(q: str = "", limit: int = 15):
    """
    High-speed party-name auto-complete endpoint.

    Query params:
        q:     Partial text string to match against known party names.
        limit: Maximum suggestions to return (default 15).

    Returns:
        { "suggestions": ["Party A", "Party B", …] }
        Results are ordered by historical frequency of occurrence so the
        most active customers always appear first in the dropdown.
    """
    try:
        corpus = _load_suggest_corpus(DB_PATH)
        q_lower = q.strip().lower()

        if not q_lower:
            # No query — return the top N most frequent parties overall
            suggestions = [name for name, _ in corpus[:limit]]
        else:
            # Priority 1: names that START with the query string
            starts = [(n, c) for n, c in corpus if n.lower().startswith(q_lower)]
            # Priority 2: names that CONTAIN the query string anywhere (but don't start with it)
            contains = [(n, c) for n, c in corpus if q_lower in n.lower() and not n.lower().startswith(q_lower)]
            # Merge and deduplicate while preserving frequency order
            merged = starts + contains
            suggestions = [name for name, _ in merged[:limit]]

        return {"suggestions": suggestions, "count": len(suggestions)}

    except Exception as e:
        logger.error(f"suggest-parties failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# =====================================================================
# PROMPT 88: Customer Contracts API Endpoints
# =====================================================================
class ContractUpsertRequest(BaseModel):
    party_name: str
    discount_type: str
    discount_value: float

class ContractDeleteRequest(BaseModel):
    party_name: str

@app.get("/api/contracts")
def api_get_contracts():
    try:
        from discount_manager import get_all_contracts
        return get_all_contracts(DB_PATH)
    except Exception as e:
        logger.error(f"Failed to fetch contracts: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/contracts/upsert")
def api_upsert_contract(req: ContractUpsertRequest):
    if req.discount_type not in ('FIXED_PER_LITER', 'PERCENTAGE_REDUCTION'):
        raise HTTPException(status_code=400, detail="Invalid discount type. Must be 'FIXED_PER_LITER' or 'PERCENTAGE_REDUCTION'.")
    if req.discount_value < 0:
        raise HTTPException(status_code=400, detail="Discount value must be non-negative.")
    try:
        from discount_manager import upsert_contract
        upsert_contract(req.party_name, req.discount_type, req.discount_value, DB_PATH)
        _invalidate_suggest_cache() # Clear suggest cache
        return {"status": "success", "message": f"Contract for '{req.party_name}' saved."}
    except Exception as e:
        logger.error(f"Failed to upsert contract: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/contracts/delete")
def api_delete_contract(req: ContractDeleteRequest):
    try:
        from discount_manager import delete_contract
        delete_contract(req.party_name, DB_PATH)
        _invalidate_suggest_cache() # Clear suggest cache
        return {"status": "success", "message": f"Contract for '{req.party_name}' deleted."}
    except Exception as e:
        logger.error(f"Failed to delete contract: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# =====================================================================
# Fuel Purchase Cost Price (CP) Registry Endpoints
# =====================================================================
class PurchaseRateUpsertRequest(BaseModel):
    effective_date: str
    product_type: str
    purchase_rate_per_liter: float
    invoice_reference: Optional[str] = None

class PurchaseRateDeleteRequest(BaseModel):
    effective_date: str
    product_type: str

@app.get("/api/purchase-rates")
def api_get_purchase_rates():
    try:
        from purchase_registry import get_all_purchase_rates
        return get_all_purchase_rates(DB_PATH)
    except Exception as e:
        logger.error(f"Failed to fetch purchase rates: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/purchase-rates/upsert")
def api_upsert_purchase_rate(req: PurchaseRateUpsertRequest):
    if req.product_type not in ('HSD', 'MS'):
        raise HTTPException(status_code=400, detail="Invalid product type. Must be 'HSD' or 'MS'.")
    if req.purchase_rate_per_liter <= 0:
        raise HTTPException(status_code=400, detail="Purchase rate must be a positive number.")
    try:
        from datetime import datetime
        datetime.strptime(req.effective_date, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail="effective_date must be in YYYY-MM-DD format.")

    try:
        from purchase_registry import upsert_purchase_rate
        upsert_purchase_rate(
            req.effective_date,
            req.product_type,
            req.purchase_rate_per_liter,
            req.invoice_reference or "",
            DB_PATH
        )
        return {"status": "success", "message": f"Purchase rate for {req.product_type} on {req.effective_date} saved."}
    except Exception as e:
        logger.error(f"Failed to upsert purchase rate: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/purchase-rates/delete")
def api_delete_purchase_rate(req: PurchaseRateDeleteRequest):
    try:
        from purchase_registry import delete_purchase_rate
        delete_purchase_rate(req.effective_date, req.product_type, DB_PATH)
        return {"status": "success", "message": f"Purchase rate for {req.product_type} on {req.effective_date} deleted."}
    except Exception as e:
        logger.error(f"Failed to delete purchase rate: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# Bulk Editor Endpoint: Fetch detailed ledger entries for date range
@app.get("/api/ledger/bulk-fetch")

def bulk_fetch_ledger(start_date: Optional[str] = None, end_date: Optional[str] = None):
    """
    Fetches all ledger entries within the specified date range.
    Decrypts party_name and amount for editing in the grid.
    """
    logger.info(f"Bulk fetching ledger entries from {start_date} to {end_date}...")
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        query = """
            SELECT entry_id, date, party_name, vehicle_wheel_no, amount, type, remarks, 
                   payment_status, amount_remaining, linked_payment_id,
                   base_amount, discount_applied, base_rate
            FROM ledger_entries
        """
        params = []
        conditions = []
        
        if start_date:
            conditions.append("date >= ?")
            params.append(start_date)
        if end_date:
            conditions.append("date <= ?")
            params.append(end_date)
            
        if conditions:
            query += " WHERE " + " AND ".join(conditions)
            
        query += " ORDER BY date ASC, entry_id ASC"
        
        cursor.execute(query, tuple(params))
        rows = cursor.fetchall()
        conn.close()
        
        entries = []
        for row in rows:
            entry_id = row[0]
            date_str = row[1]
            party_enc = row[2]
            vehicle = row[3] or ""
            amount_enc = row[4]
            etype = row[5]
            remarks = row[6] or ""
            payment_status = row[7] or "UNPAID"
            amount_remaining = row[8]
            linked_payment_id = row[9] or ""
            base_amount_enc = row[10] if len(row) > 10 else None
            discount_applied_enc = row[11] if len(row) > 11 else None
            base_rate_enc = row[12] if len(row) > 12 else None
            
            try:
                party_name = decrypt_field(party_enc, return_type=str)
                amount = decrypt_field(amount_enc, return_type=float)
            except Exception:
                party_name = str(party_enc or "")
                amount = float(amount_enc or 0.0)
                
            try:
                base_amount = decrypt_field(base_amount_enc, return_type=float) if base_amount_enc is not None else amount
            except Exception:
                base_amount = float(base_amount_enc) if base_amount_enc is not None else amount
                
            try:
                discount_applied = decrypt_field(discount_applied_enc, return_type=float) if discount_applied_enc is not None else 0.0
            except Exception:
                discount_applied = float(discount_applied_enc) if discount_applied_enc is not None else 0.0
                
            try:
                base_rate = decrypt_field(base_rate_enc, return_type=float) if base_rate_enc is not None else None
            except Exception:
                base_rate = float(base_rate_enc) if base_rate_enc is not None else None
                
            entries.append({
                "entry_id": entry_id,
                "date": date_str,
                "party_name": party_name,
                "vehicle_wheel_no": vehicle,
                "amount": amount,
                "type": etype,
                "remarks": remarks,
                "payment_status": payment_status,
                "amount_remaining": amount_remaining,
                "linked_payment_id": linked_payment_id,
                "base_amount": base_amount,
                "discount_applied": discount_applied,
                "base_rate": base_rate
            })
            
        return entries
    except Exception as e:
        logger.error(f"Bulk fetch failed: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Bulk Fetch Error: {str(e)}")

# Bulk Editor Endpoint: Apply batch updates to ledger entries inside a strict transaction
@app.post("/api/ledger/bulk-update")
def bulk_update_ledger(patches: List[Dict[str, Any]]):
    """
    Applies bulk updates (patches) to ledger entries inside a safe transaction block.
    """
    logger.info(f"Applying bulk updates to {len(patches)} ledger entries...")
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    try:
        conn.execute("BEGIN TRANSACTION")
        for patch in patches:
            entry_id = patch.get("entry_id")
            if not entry_id:
                raise ValueError("Each patch must include 'entry_id'.")
            
            updates = []
            params = []
            
            if "date" in patch:
                updates.append("date = ?")
                params.append(patch["date"])
            
            if "party_name" in patch:
                updates.append("party_name = ?")
                params.append(encrypt_field(patch["party_name"].strip()))
                
            if "vehicle_wheel_no" in patch:
                updates.append("vehicle_wheel_no = ?")
                params.append(patch["vehicle_wheel_no"])
                
            if "amount" in patch:
                updates.append("amount = ?")
                params.append(encrypt_field(float(patch["amount"])))
                
            if "type" in patch:
                updates.append("type = ?")
                params.append(patch["type"])
                
            if "remarks" in patch:
                updates.append("remarks = ?")
                params.append(patch["remarks"])
                
            if updates:
                sql = f"UPDATE ledger_entries SET {', '.join(updates)} WHERE entry_id = ?"
                params.append(entry_id)
                cursor.execute(sql, tuple(params))
                
        conn.commit()
        
        # Proactively trigger Excel ledger synchronization
        try:
            from exporter import export_db_to_excel
            export_db_to_excel(EXCEL_PATH)
            logger.info("✓ SQLite changes synchronized to master Excel ledger successfully.")
        except Exception as sync_err:
            logger.warning(f"Excel synchronization failed after bulk update: {sync_err}")
            
        return {"status": "success", "message": f"Successfully updated {len(patches)} records."}
        
    except Exception as e:
        conn.rollback()
        logger.error(f"Bulk update failed, rolled back cleanly. Error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Transaction rolled back: {str(e)}")
    finally:
        conn.close()

# Endpoint 4: Export complete styled Excel database
@app.get("/api/export-excel")
def export_excel():
    """
    Downloads the professionally styled Excel ledger directly from the server.
    Refreshes the Excel database from SQLite first.
    """
    logger.info("Exporting and refreshing complete Excel ledger...")
    try:
        from exporter import export_db_to_excel
        export_db_to_excel(EXCEL_PATH)
    except Exception as e:
        logger.error(f"Failed to refresh Excel database: {str(e)}")
        
    if not os.path.exists(EXCEL_PATH):
        raise HTTPException(status_code=404, detail="No ledger file found. Please process and save some entries first.")
    return FileResponse(
        path=EXCEL_PATH, 
        filename=os.path.basename(EXCEL_PATH), 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# Endpoint 4b: Export PetroByte sync CSV
@app.get("/api/export-petrobyte")
def export_petrobyte(date: str = "all"):
    """
    Generates and returns the double-entry PetroByte sync CSV file directly as a download response.
    """
    logger.info(f"Generating PetroByte CSV for date: {date}")
    try:
        from exporter import generate_accounting_export
        excel_p, csv_p = generate_accounting_export(date)
        if os.path.exists(csv_p):
            return FileResponse(
                path=csv_p,
                filename=os.path.basename(csv_p),
                media_type="text/csv"
            )
        else:
            raise HTTPException(status_code=500, detail="CSV file was not created by the exporter.")
    except Exception as e:
        logger.error(f"PetroByte CSV generation failed: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to generate PetroByte CSV: {str(e)}")

# Endpoint 4c: Database recovery and spreadsheet reconstruction
@app.post("/api/rebuild-master-excel")
def rebuild_master_excel():
    """
    Triggers the complete database recovery and spreadsheet reconstruction utility.
    Checks and isolates open/missing/corrupted spreadsheets and compiles a fresh Pump_Accounts.xlsx.
    """
    logger.info("Triggering database recovery and spreadsheet rebuild...")
    try:
        from repair_kit import rebuild_master_spreadsheet
        result_path = rebuild_master_spreadsheet()
        return {
            "status": "success",
            "message": "Master Excel spreadsheet reconstructed successfully from database!",
            "file_path": result_path
        }
    except Exception as e:
        logger.error(f"Reconstruction failed: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Failed to rebuild master spreadsheet: {str(e)}"
        )

# =====================================================================
# PROMPT 6: New Core Operational Routes
# =====================================================================

# Pydantic schema for save-ledger-day request payload
class SaveLedgerDayRequest(BaseModel):
    date: str
    total_calculated_liters_hsd: float = 0.0
    total_calculated_liters_ms: float = 0.0
    total_cash_calculated: float = 0.0
    total_credit_sales: float = 0.0
    total_testing_deductions: float = 0.0
    total_regular_hsd_liters: float = 0.0
    total_premium_hsd_liters: float = 0.0
    total_regular_ms_liters: float = 0.0
    total_premium_ms_liters: float = 0.0
    validation_status: str = "balanced"
    mathematical_warnings: list = []
    nozzles: list = []
    credit_sales: list = []
    cash_expenses: list = []
    dsm_shifts: list = []
    lube_sales: list = []
    card_settlements: list = []
    credit_realizations: list = []
    staff_advances: list = []
    meter_replaced: bool = False
    replacement_offset_liters: float = 0.0

@app.post("/api/upload")
async def upload_sheet(
    files: List[UploadFile] = File(None),
    image: UploadFile = File(None)
):
    """
    Accepts a list of multipart form files (or a single file fallback).
    Decodes multi-page PDFs or direct images, generates tracking hashes,
    saves clean pages, executes shadow division preprocessing and AI audits,
    and returns structured audited accounting JSON schema(s).
    """
    import cv2
    import uuid
    import time
    from intake import convert_upload_to_cv2_matrices, generate_matrix_hash
    from processor import optimize_register_image
    from ai_engine import analyze_register_sheet
    
    # 1. Resolve files list
    upload_list = []
    if files:
        upload_list = files
    elif image:
        upload_list = [image]
    else:
        raise HTTPException(status_code=400, detail="No files or image provided for upload.")
        
    upload_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "uploaded_raw_photos")
    os.makedirs(upload_dir, exist_ok=True)
    
    results = []
    
    try:
        for upload_file in upload_list:
            logger.info(f"Processing upload in bulk intake pipeline: {upload_file.filename}")
            
            # Convert multi-format pages to CV2 matrices
            sheets = convert_upload_to_cv2_matrices(upload_file)
            
            for page_idx, cv2_matrix, page_filename in sheets:
                # A. Generate stable content hash
                page_hash = generate_matrix_hash(cv2_matrix)
                
                # B. Save temporary raw file locally
                file_ext = os.path.splitext(page_filename)[1] or ".png"
                local_filename = f"raw_{page_hash}{file_ext}"
                local_path = os.path.join(upload_dir, local_filename)
                
                cv2.imwrite(local_path, cv2_matrix)
                logger.info(f"Saved intake page to raw path: {local_path}")

                # C. Detect double-page spreads (landscape scans) and split if needed.
                #    Portrait images are returned as-is ([local_path]);
                #    Landscape spreads are split into [left_path, right_path].
                from stitcher import detect_and_split_double_pages
                split_paths = detect_and_split_double_pages(local_path)
                logger.info(
                    f"Stitcher produced {len(split_paths)} page(s) from "
                    f"'{os.path.basename(local_path)}': "
                    f"{[os.path.basename(p) for p in split_paths]}"
                )

                for split_idx, split_path in enumerate(split_paths):
                    # Clarity Check Guardrail (PROMPT 46)
                    from image_guard import validate_image_clarity
                    from state_tracker import calculate_file_hash, upsert_job, mark_preprocessing_failed
                    from archiver import archive_processed_file
                    
                    clarity_res = validate_image_clarity(split_path)
                    if not clarity_res["success"]:
                        split_hash = calculate_file_hash(split_path)
                        split_filename = os.path.basename(split_path)
                        
                        # Upsert job to database state machine
                        upsert_job(split_hash, split_filename, db_path=DB_PATH)
                        # Mark job as failed preprocessing with the reason
                        reason_msg = f"Clarity check failed: {clarity_res['status']} (Focus score: {clarity_res['focus_score']:.2f}, Contrast: {clarity_res['contrast_score']:.2f})"
                        mark_preprocessing_failed(split_hash, reason_msg, db_path=DB_PATH)
                        
                        # Slide file straight to requires_human_review folder
                        archived_path = archive_processed_file(split_path, status="FAILED_PREPROCESSING")
                        
                        # Return failed audited JSON bypass
                        failed_json = {
                            "date": "N/A",
                            "validation_status": "FAILED_PREPROCESSING",
                            "mathematical_warnings": [f"FAILED_PREPROCESSING: Clarity check failed: {clarity_res['status']}"],
                            "image_url": f"http://localhost:8000/uploaded_raw_photos/{local_filename}",
                            "split_image_url": f"http://localhost:8000/uploaded_raw_photos/{os.path.basename(archived_path)}",
                            "page_hash": split_hash,
                            "page_index": page_idx,
                            "split_index": split_idx,
                            "split_total": len(split_paths),
                            "original_filename": upload_file.filename,
                            "error": f"Image clarity check failed: {clarity_res['status']}"
                        }
                        results.append(failed_json)
                        continue
                        
                    # D. Run OpenCV optimizer shadow divisions & CLAHE
                    optimized_path = optimize_register_image(split_path)
                    
                    # E. Run Gemini + Claude Accounting vision auditing pipeline
                    audited_json = analyze_register_sheet(optimized_path)
                    
                    # FTS Indexing Hook
                    try:
                        raw_text = audited_json.get("raw_transcription_text")
                        date_val = audited_json.get("date")
                        if date_val and date_val != "N/A" and raw_text:
                            from text_search import index_ledger_text
                            index_ledger_text(DB_PATH, date_val, raw_text)
                    except Exception as fts_index_err:
                        logger.warning(f"Failed to auto-index FTS text: {str(fts_index_err)}")
                    
                    # Run chronological continuity check
                    try:
                        from interlock_checker import verify_chronological_continuity
                        date_extracted = audited_json.get("date")
                        nozzles_extracted = audited_json.get("nozzles") or []
                        if date_extracted:
                            continuity = verify_chronological_continuity(
                                target_date_string=date_extracted,
                                current_nozzles=nozzles_extracted,
                                db_path=DB_PATH
                            )
                            audited_json["continuity_check"] = continuity
                    except Exception as continuity_err:
                        logger.warning(f"Failed to run chronological interlock check on upload: {str(continuity_err)}")
                    
                    # F. Inject static retrieval elements
                    split_filename = os.path.basename(split_path)
                    audited_json["image_url"] = f"http://localhost:8000/uploaded_raw_photos/{local_filename}"
                    audited_json["split_image_url"] = f"http://localhost:8000/processed_images/{split_filename}"
                    audited_json["page_hash"] = page_hash
                    audited_json["page_index"] = page_idx
                    audited_json["split_index"] = split_idx
                    audited_json["split_total"] = len(split_paths)
                    audited_json["original_filename"] = upload_file.filename
                    
                    results.append(audited_json)
                
        if len(results) == 0:
            raise HTTPException(status_code=400, detail="No valid sheets could be parsed from the files.")
        
        # Trigger database WAL checkpoint after the bulk historical execution finishes its loop
        try:
            from db_hardener import execute_wal_checkpoint
            execute_wal_checkpoint()
        except Exception as checkpoint_err:
            logger.warning(f"Failed to execute database WAL checkpoint after bulk upload: {str(checkpoint_err)}")

        if len(results) == 1:
            return results[0]
        else:
            return results
            
    except ValueError as ve:
        logger.error(f"Validation error in intake pipeline: {str(ve)}")
        raise HTTPException(status_code=400, detail=str(ve))
    except Exception as e:
        logger.error(f"Failed in multi-format /api/upload: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Multi-format upload and vision pipeline failed: {str(e)}")


@app.post("/api/save-ledger-day")
def save_ledger_day(payload: SaveLedgerDayRequest):
    """
    Accepts the verified JSON payload containing corrected values.
    Inserts a summary block into daily_summary, unpacks individual credit sales
    and expenses line-by-line into ledger_entries, and automatically triggers
    an update to the master spreadsheet.
    """
    logger.info(f"Saving ledger day for date: {payload.date}")
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # 1. Insert/Replace summary block in daily_summary table
        date_str = payload.date
        
        # Apply customer special contract pricing and discounts (Prompt 88)
        try:
            from discount_manager import apply_contract_discounts
            updated_credit_sales = apply_contract_discounts(payload.credit_sales, date_str, DB_PATH)
        except Exception as discount_err:
            logger.error(f"Failed to apply customer contract discounts on save (non-fatal): {str(discount_err)}")
            updated_credit_sales = payload.credit_sales
            
        # 1. Insert/Replace summary block in daily_summary table
        date_str = payload.date
        
        # Apply customer special contract pricing and discounts (Prompt 88)
        try:
            from discount_manager import apply_contract_discounts
            updated_credit_sales = apply_contract_discounts(payload.credit_sales, date_str, DB_PATH)
        except Exception as discount_err:
            logger.error(f"Failed to apply customer contract discounts on save (non-fatal): {str(discount_err)}")
            updated_credit_sales = payload.credit_sales
            
        total_credit = sum(float(c.get("amount") or 0.0) for c in updated_credit_sales)
        
        # Recalculate all nozzle flow rates dynamically using calculate_net_nozzle_volume
        from meter_rollover import calculate_net_nozzle_volume
        
        # Calculate original totals from nozzles payload
        orig_nozzles_hsd = 0.0
        orig_nozzles_ms = 0.0
        orig_nozzles_cash = 0.0
        orig_nozzles_testing = 0.0
        
        for n in payload.nozzles:
            ft = str(n.get("fuel_type") or "").strip().upper()
            sales = float(n.get("net_sales_liters") or n.get("sales_liters_calculated") or 0.0)
            amt = float(n.get("amount_calculated") or 0.0)
            testing = float(n.get("testing_liters") or 0.0)
            
            if ft in ("REGULAR_HSD", "PREMIUM_HSD", "HSD"):
                orig_nozzles_hsd += sales
            elif ft in ("REGULAR_MS", "PREMIUM_MS", "MS"):
                orig_nozzles_ms += sales
                
            orig_nozzles_cash += amt
            orig_nozzles_testing += testing
            
        recalculated_nozzles = []
        new_nozzles_hsd = 0.0
        new_nozzles_ms = 0.0
        new_nozzles_cash = 0.0
        new_nozzles_testing = 0.0
        
        for n in payload.nozzles:
            opening = float(n.get("opening") or 0.0)
            closing = float(n.get("closing") or 0.0)
            testing = float(n.get("testing_liters") or 0.0)
            rate = float(n.get("rate") or 0.0)
            ft = str(n.get("fuel_type") or "").strip().upper()
            
            # Determine replacement override
            is_replaced = bool(n.get("meter_replaced") or False)
            offset = float(n.get("replacement_offset_liters") or 0.0)
            
            if payload.meter_replaced and (is_replaced or closing < opening or len(payload.nozzles) == 1):
                is_replaced = True
                if offset == 0.0:
                    offset = float(payload.replacement_offset_liters or 0.0)
            
            calc_flow = calculate_net_nozzle_volume(
                opening=opening,
                closing=closing,
                meter_replaced=is_replaced,
                replacement_offset_liters=offset
            )
            
            net_sales = round(calc_flow - testing, 2)
            amt = round(net_sales * rate, 2)
            
            n_copy = dict(n)
            n_copy["calculated_flow"] = calc_flow
            n_copy["net_sales_liters"] = net_sales
            n_copy["sales_liters_calculated"] = net_sales
            n_copy["amount_calculated"] = amt
            n_copy["meter_replaced"] = is_replaced
            n_copy["replacement_offset_liters"] = offset
            
            recalculated_nozzles.append(n_copy)
            
            # Aggregate new totals
            if ft in ("REGULAR_HSD", "PREMIUM_HSD", "HSD"):
                new_nozzles_hsd += net_sales
            elif ft in ("REGULAR_MS", "PREMIUM_MS", "MS"):
                new_nozzles_ms += net_sales
                
            new_nozzles_cash += amt
            new_nozzles_testing += testing
            
        # Override payload nozzles list with recalculated ones so later processing is consistent
        payload.nozzles = recalculated_nozzles
        
        # Calculate final totals by adjusting the original payload values
        total_hsd = float(payload.total_calculated_liters_hsd)
        total_ms = float(payload.total_calculated_liters_ms)
        total_cash = float(payload.total_cash_calculated)
        total_testing = float(payload.total_testing_deductions)
        
        if payload.nozzles:
            total_hsd = round(total_hsd - orig_nozzles_hsd + new_nozzles_hsd, 2)
            total_ms = round(total_ms - orig_nozzles_ms + new_nozzles_ms, 2)
            total_cash = round(total_cash - orig_nozzles_cash + new_nozzles_cash, 2)
            total_testing = round(total_testing - orig_nozzles_testing + new_nozzles_testing, 2)
            
        # If validation status is balanced and there are no warnings, is_verified is set to 1
        is_verified = 1 if (payload.validation_status == "balanced" and len(payload.mathematical_warnings) == 0) else 0
        
        cursor.execute("PRAGMA table_info(daily_summary)")
        summary_cols = [c[1] for c in cursor.fetchall()]
        
        summary_data = {
            "date": date_str,
            "total_hsd_liters": total_hsd,
            "total_ms_liters": total_ms,
            "total_cash_calculated": total_cash,
            "total_credit_sales": total_credit,
            "total_testing_deductions": total_testing,
            "is_verified": is_verified,
        }
        
        # Fill in extended splits if they are in the database schema
        if "total_regular_hsd_liters" in summary_cols:
            total_reg_hsd = float(payload.total_regular_hsd_liters or 0.0)
            total_prem_hsd = float(payload.total_premium_hsd_liters or 0.0)
            total_reg_ms = float(payload.total_regular_ms_liters or 0.0)
            total_prem_ms = float(payload.total_premium_ms_liters or 0.0)
            
            # Re-aggregate splits if they are zero
            if total_reg_hsd == 0.0 and total_prem_hsd == 0.0 and total_reg_ms == 0.0 and total_prem_ms == 0.0:
                for n in payload.nozzles:
                    ft = str(n.get("fuel_type") or "").strip().upper()
                    sales = float(n.get("net_sales_liters") or 0.0)
                    if ft == "REGULAR_HSD":
                        total_reg_hsd += sales
                    elif ft == "PREMIUM_HSD":
                        total_prem_hsd += sales
                    elif ft == "REGULAR_MS":
                        total_reg_ms += sales
                    elif ft == "PREMIUM_MS":
                        total_prem_ms += sales
                    elif ft == "HSD":
                        total_reg_hsd += sales
                    elif ft == "MS":
                        total_reg_ms += sales
            
            summary_data["total_regular_hsd_liters"] = round(total_reg_hsd, 2)
            summary_data["total_premium_hsd_liters"] = round(total_prem_hsd, 2)
            summary_data["total_regular_ms_liters"] = round(total_reg_ms, 2)
            summary_data["total_premium_ms_liters"] = round(total_prem_ms, 2)
            
        if "meter_replaced" in summary_cols:
            summary_data["meter_replaced"] = 1 if payload.meter_replaced else 0
            summary_data["replacement_offset_liters"] = float(payload.replacement_offset_liters or 0.0)

        fields = list(summary_data.keys())
        placeholders = ", ".join(["?"] * len(fields))
        columns_str = ", ".join(fields)
        query = f"INSERT OR REPLACE INTO daily_summary ({columns_str}) VALUES ({placeholders})"
        cursor.execute(query, tuple(summary_data[f] for f in fields))
        
        # Clean existing entries for this date in ledger_entries to support safe manual overrides/updates
        
        # Clean existing entries for this date in ledger_entries to support safe manual overrides/updates
        cursor.execute("DELETE FROM ledger_entries WHERE date = ?", (date_str,))
        
        # 2. Unpack individual credit sales ('udhaar') and insert line-by-line
        for credit in updated_credit_sales:
            base_amount = credit.get("base_amount")
            discount_applied = credit.get("discount_applied")
            base_rate = credit.get("base_rate")
            
            cursor.execute("""
                INSERT INTO ledger_entries (
                    date, party_name, vehicle_wheel_no, amount, type, remarks,
                    base_amount, discount_applied, base_rate
                )
                VALUES (?, ?, ?, ?, 'udhaar', ?, ?, ?, ?)
            """, (
                date_str,
                encrypt_field(credit.get("party_name") or "Unknown Party"),
                credit.get("vehicle_no") or "N/A",
                encrypt_field(float(credit.get("amount") or 0.0)),
                credit.get("remarks") or "HSD credit sale",
                encrypt_field(base_amount) if base_amount is not None else None,
                encrypt_field(discount_applied) if discount_applied is not None else None,
                encrypt_field(base_rate) if base_rate is not None else None
            ))
            
        # 3. Unpack individual cash expenses ('expense') and insert line-by-line
        for expense in payload.cash_expenses:
            # Resolve accounting_head: prefer pre-assigned (from AI engine categorizer),
            # otherwise run classifier live as a safety net.
            acct_head = expense.get("accounting_head")
            if not acct_head:
                try:
                    from expense_mapper import get_accounting_head
                    combined = f"{expense.get('party_name') or ''} {expense.get('remarks') or ''}".strip()
                    acct_head = get_accounting_head(combined)
                except Exception:
                    acct_head = "Unclassified Operational Expenses"

            cursor.execute("""
                INSERT INTO ledger_entries (date, party_name, vehicle_wheel_no, amount, type, remarks, accounting_head)
                VALUES (?, ?, ?, ?, 'expense', ?, ?)
            """, (
                date_str,
                encrypt_field(expense.get("party_name") or "Office Expense"),
                "N/A",
                encrypt_field(float(expense.get("amount") or 0.0)),
                expense.get("remarks") or "Cash expense",
                acct_head
            ))

            
        # 3.5. Unpack and save individual staff advances
        try:
            from staff_ledger import delete_daily_staff_advances, record_staff_advance_with_cursor
            delete_daily_staff_advances(date_str, conn)
            for adv in getattr(payload, "staff_advances", []) or []:
                record_staff_advance_with_cursor(
                    cursor=cursor,
                    date=date_str,
                    employee_name=adv.get("employee_name"),
                    amount_drawn=float(adv.get("amount_drawn") or 0.0),
                    atype=adv.get("type") or "CASH_ADVANCE",
                    remarks=adv.get("remarks") or ""
                )
        except Exception as staff_save_err:
            logger.warning(f"Failed to save staff advances during save-ledger-day: {str(staff_save_err)}")
            
        # 4. Insert into daily_ledger as well for historical logs UI view parity
        total_sales_liters = total_hsd + total_ms
        val_status_str = "valid" if is_verified == 1 else "needs_review"
        
        nozzle_list = []
        for n in payload.nozzles:
            nozzle_list.append({
                "nozzle_name": n.get("nozzle_name"),
                "fuel_type": n.get("fuel_type"),
                "opening": float(n.get("opening") or 0.0),
                "closing": float(n.get("closing") or 0.0),
                "sales_liters_calculated": float(n.get("net_sales_liters") or n.get("sales_liters_calculated") or 0.0),
                "rate": float(n.get("rate") or 0.0),
                "amount_calculated": float(n.get("amount_calculated") or 0.0),
                "arithmetic_valid": n.get("is_valid" or n.get("arithmetic_valid") or True),
                "discrepancy_details": n.get("math_warning" or n.get("discrepancy_details")),
                "meter_replaced": bool(n.get("meter_replaced") or False),
                "replacement_offset_liters": float(n.get("replacement_offset_liters") or 0.0)
            })
            
        raw_payload = {
            "date": date_str,
            "nozzles": nozzle_list,
            "hinglish_notes": payload.mathematical_warnings,
            "total_sales_liters": total_sales_liters,
            "total_amount_inr": total_cash,
            "validation_status": val_status_str,
            "audit_explanation": ", ".join(payload.mathematical_warnings) or "Manual verified entry",
            "cash_tender": max(0.0, total_cash - total_credit - sum(float(e.get("amount") or 0.0) for e in payload.cash_expenses)),
            "upi_tender": 0.0,
            "paytm_transfers": 0.0,
            "card_tender": 0.0,
            "udhaar_sales": total_credit,
            "expenses_amount": sum(float(e.get("amount") or 0.0) for e in payload.cash_expenses),
            "meter_replaced": payload.meter_replaced,
            "replacement_offset_liters": payload.replacement_offset_liters
        }
        
        raw_payload["dsm_shifts"] = getattr(payload, "dsm_shifts", [])
        raw_payload["lube_sales"] = getattr(payload, "lube_sales", [])
        raw_payload["card_settlements"] = getattr(payload, "card_settlements", [])
        raw_payload["credit_realizations"] = getattr(payload, "credit_realizations", [])
        raw_payload["staff_advances"] = getattr(payload, "staff_advances", [])
        encrypted_raw_payload = encrypt_raw_data(raw_payload)
        raw_json_str = json.dumps(encrypted_raw_payload, ensure_ascii=False)
        
        cursor.execute("""
            INSERT OR REPLACE INTO daily_ledger 
            (date, total_sales_liters, total_amount_inr, cash_tender, upi_tender, paytm_transfers, card_tender, udhaar_sales, expenses_amount, validation_status, raw_data)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            date_str,
            total_sales_liters,
            total_cash,
            raw_payload["cash_tender"],
            0.0,
            0.0,
            0.0,
            total_credit,
            raw_payload["expenses_amount"],
            val_status_str,
            raw_json_str
        ))
        
        # Save credit payment realizations and run FIFO allocation inside the same transaction block
        try:
            from credit_realization import save_credit_realization
            from fifo_settler import allocate_realization_fifo
            
            cursor.execute("DELETE FROM credit_realizations WHERE date = ?", (date_str,))
            
            credit_realizations_list = getattr(payload, "credit_realizations", []) or []
            for real in credit_realizations_list:
                party = real.get("party_name")
                if party:
                    amt = float(real.get("amount_received") or 0.0)
                    mode = real.get("payment_mode") or "CASH"
                    remarks = real.get("bank_utr_or_remarks") or ""
                    invoice = real.get("linked_invoice_no") or ""
                    save_credit_realization(
                        date_str=date_str,
                        party_name=party,
                        amount_received=amt,
                        payment_mode=mode,
                        bank_utr_or_remarks=remarks,
                        linked_invoice_no=invoice,
                        db_path=DB_PATH,
                        conn=conn
                    )
                    # Settle customer's outstanding invoices via FIFO allocation
                    allocate_realization_fifo(
                        party_name=party,
                        payment_amount=amt,
                        payment_date=date_str,
                        db_path=DB_PATH,
                        conn=conn
                    )
        except Exception as real_save_err:
            logger.error(f"Failed to save credit realizations during save: {str(real_save_err)}")
            raise real_save_err
            
        # Calculate daily gross profit and store it inside the daily_summary table (Prompt 97)
        try:
            from profit_engine import calculate_and_store_daily_profit
            calculate_and_store_daily_profit(date_str, db_path=DB_PATH, conn=conn)
        except Exception as profit_err:
            logger.warning(f"Failed to calculate and store daily profit during save (non-fatal): {str(profit_err)}")
            
        conn.commit()
        conn.close()
        logger.info("Local SQLite transaction committed successfully.")
        _invalidate_suggest_cache()   # Ensure autocomplete sees new party names immediately
        
        # Save individual DSM shifts
        try:
            delete_dsm_shifts_by_date(date_str, db_path=DB_PATH)
            for dsm in getattr(payload, "dsm_shifts", []):
                dsm_name = dsm.get("dsm_name")
                if dsm_name:
                    shift_type = dsm.get("shift_type") or "Day"
                    assigned_nozzles = dsm.get("assigned_nozzles")
                    if isinstance(assigned_nozzles, list):
                        assigned_nozzles_str = ", ".join(assigned_nozzles)
                    else:
                        assigned_nozzles_str = str(assigned_nozzles or "")
                        
                    cash_handed = float(dsm.get("cash_handed_over") or 0.0)
                    digital_value = float(dsm.get("digital_slips_value") or 0.0)
                    short_excess = dsm.get("calculated_shortage_or_excess")
                    
                    if short_excess is None or short_excess == 0.0:
                        nozzles_list = [n.strip() for n in assigned_nozzles_str.split(",") if n.strip()]
                        expected_sales = calculate_dsm_expected_sales(date_str, nozzles_list, db_path=DB_PATH)
                        if expected_sales > 0.0:
                            short_excess = (cash_handed + digital_value) - expected_sales
                        else:
                            short_excess = float(short_excess or 0.0)
                    else:
                        short_excess = float(short_excess)
                        
                    save_dsm_shift(
                        date_str=date_str,
                        shift_type=shift_type,
                        dsm_name=dsm_name,
                        assigned_nozzles=assigned_nozzles_str,
                        cash_handed_over=cash_handed,
                        digital_slips_value=digital_value,
                        calculated_shortage_or_excess=short_excess,
                        db_path=DB_PATH
                    )
        except Exception as dsm_save_err:
            logger.warning(f"Failed to save DSM shifts during save: {str(dsm_save_err)}")
            
        # Save lube/ancillary sales and run validation guard to update daily summary cash expected
        try:
            from lube_sales import delete_lube_sales_by_date, save_lube_sale, verify_inventory_totals
            delete_lube_sales_by_date(date_str, db_path=DB_PATH)
            for lube in getattr(payload, "lube_sales", []):
                item = lube.get("item_name")
                if item:
                    qty = float(lube.get("quantity_sold") or 0.0)
                    price = float(lube.get("unit_price") or 0.0)
                    rev = float(lube.get("total_item_revenue") or 0.0)
                    if rev == 0.0:
                        rev = qty * price
                    save_lube_sale(
                        date_str=date_str,
                        item_name=item,
                        quantity_sold=qty,
                        unit_price=price,
                        total_item_revenue=rev,
                        db_path=DB_PATH
                    )
            verify_inventory_totals(date_str, db_path=DB_PATH)
        except Exception as lube_save_err:
            logger.warning(f"Failed to save lube sales during save: {str(lube_save_err)}")
            
        # Save card swipe settlements
        try:
            from card_settlements import save_card_settlements
            card_settlements_list = getattr(payload, "card_settlements", [])
            save_card_settlements(date_str, card_settlements_list, db_path=DB_PATH)
        except Exception as card_save_err:
            logger.warning(f"Failed to save card settlements during save: {str(card_save_err)}")
            
        # Credit payment realizations and FIFO allocation have been moved inside the main transaction block before commit
        pass
        
        # Check if there is a flagged image for this date in requires_human_review, and archive it
        try:
            backend_dir = os.path.dirname(os.path.abspath(__file__))
            workspace_dir = os.path.dirname(backend_dir)
            review_dir = os.path.join(workspace_dir, "ledger_photos", "requires_human_review")
            if os.path.exists(review_dir):
                import glob
                matching_files = glob.glob(os.path.join(review_dir, f"{date_str}_review_*"))
                if matching_files:
                    from archiver import archive_processed_file
                    for mf in matching_files:
                        logger.info(f"Moving active review file out of requires_human_review: {mf}")
                        # Move it out to successfully_imported with status balanced
                        archive_processed_file(mf, "balanced", date_str)
        except Exception as archive_err:
            logger.warning(f"Failed to auto-archive review image: {str(archive_err)}")
        
        # 5. Automatically trigger an update to the master spreadsheet and pump exports
        try:
            from exporter import export_db_to_excel, generate_accounting_export
            export_db_to_excel(EXCEL_PATH)
            generate_accounting_export(payload.date)
            logger.info("Master spreadsheet and pump exports generated successfully.")
            
            # Save evaporation allowance and trigger logs
            try:
                from evaporation_handler import calculate_evaporation_allowances
                calculate_evaporation_allowances(payload.date, db_path=DB_PATH)
            except Exception as evap_log_err:
                logger.warning(f"Failed to trigger evaporation calculation logging: {str(evap_log_err)}")
                
            # Regenerate trend charts on new entry commit
            try:
                from local_analytics import generate_all_charts
                generate_all_charts(db_path=DB_PATH)
            except Exception as chart_gen_err:
                logger.warning(f"Failed to regenerate trend charts on commit: {str(chart_gen_err)}")
        except Exception as sheet_err:
            logger.error(f"Failed to auto-update Excel and pump exports: {str(sheet_err)}")
            
        # Check credit limits for each credit entry after commit
        credit_alert_message = None
        try:
            for credit in payload.credit_sales:
                p_name = credit.get("party_name")
                if p_name:
                    alert = check_credit_limit(p_name, 0.0, db_path=DB_PATH)
                    if alert and "credit cap" in alert:
                        credit_alert_message = alert
                        break
        except Exception as guard_err:
            logger.warning(f"Failed to check credit limits during save: {str(guard_err)}")
            
        # Trigger database shadow mirror clone on verified commit
        try:
            from self_healer import save_shadow_mirror
            save_shadow_mirror(DB_PATH)
        except Exception as shadow_err:
            logger.error(f"Failed to create shadow mirror clone: {str(shadow_err)}")
            
        # Trigger daily ledger backup dispatcher in background
        try:
            from backup_dispatcher import dispatch_daily_ledger_backup_background
            dispatch_daily_ledger_backup_background(payload.date, DB_PATH)
        except Exception as dispatch_err:
            logger.warning(f"Failed to trigger backup dispatch: {str(dispatch_err)}")
            
        res_dict = {
            "status": "success",
            "message": f"Ledger day for {payload.date} logged successfully and Excel updated.",
            "excel_file": EXCEL_PATH
        }
        if credit_alert_message:
            res_dict["credit_alert"] = credit_alert_message
            
        return res_dict
        
    except Exception as e:
        logger.error(f"Failed to save ledger day: {str(e)}")
        if 'conn' in locals() and conn:
            try:
                conn.rollback()
                conn.close()
            except Exception as rollback_err:
                logger.warning(f"Failed to rollback connection: {str(rollback_err)}")
        raise HTTPException(status_code=500, detail=f"Database ledger save failed: {str(e)}")

# =====================================================================
# PROMPT 15: Review Queue endpoint
# =====================================================================

@app.get("/api/review-queue")
def get_review_queue():
    """
    Queries SQLite database for all entries where status is explicitly set to 'needs_review' or 'math_discrepancy',
    and scans the requires_human_review directory for files.
    """
    logger.info("Fetching review queue entries...")
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # Select from daily_ledger where validation_status is not valid
        cursor.execute("""
            SELECT date, total_sales_liters, total_amount_inr, cash_tender, upi_tender, paytm_transfers, card_tender, udhaar_sales, expenses_amount, validation_status, raw_data 
            FROM daily_ledger
            WHERE validation_status != 'valid'
            ORDER BY date ASC
        """)
        rows = cursor.fetchall()
        conn.close()
        
        queue = []
        db_dates = set()
        
        for row in rows:
            date_str = row[0]
            db_dates.add(date_str)
            raw_data = json.loads(row[10]) if row[10] else {}
            raw_data = decrypt_raw_data(raw_data)
            
            queue.append({
                "date": date_str,
                "total_sales_liters": row[1],
                "total_amount_inr": row[2],
                "cash_tender": row[3],
                "upi_tender": row[4],
                "paytm_transfers": row[5],
                "card_tender": row[6],
                "udhaar_sales": row[7],
                "expenses_amount": row[8],
                "validation_status": row[9],
                "raw_data": raw_data
            })
            
        # Also check files in requires_human_review folder that don't have matching needs_review database records, or just check what files exist
        backend_dir = os.path.dirname(os.path.abspath(__file__))
        workspace_dir = os.path.dirname(backend_dir)
        review_dir = os.path.join(workspace_dir, "ledger_photos", "requires_human_review")
        
        if os.path.exists(review_dir):
            for filename in os.listdir(review_dir):
                if filename.endswith(".png") or filename.endswith(".jpg") or filename.endswith(".jpeg"):
                    # Parse date from filename, e.g. YYYY-MM-DD_review_filename
                    parts = filename.split("_review_", 1)
                    if len(parts) == 2:
                        date_str = parts[0]
                        if date_str not in db_dates:
                            # Create a dummy Needs Review entry for files that have no DB row yet
                            queue.append({
                                "date": date_str,
                                "total_sales_liters": 0.0,
                                "total_amount_inr": 0.0,
                                "cash_tender": 0.0,
                                "upi_tender": 0.0,
                                "paytm_transfers": 0.0,
                                "card_tender": 0.0,
                                "udhaar_sales": 0.0,
                                "expenses_amount": 0.0,
                                "validation_status": "needs_review",
                                "raw_data": {
                                    "date": date_str,
                                    "nozzles": [],
                                    "credit_sales": [],
                                    "cash_expenses": [],
                                    "offline_mode": True,
                                    "image_url": f"http://localhost:8000/uploaded_raw_photos/{filename}"
                                }
                            })
                            db_dates.add(date_str)
                            
        # Sort queue by date ascending so oldest gets reviewed first
        queue.sort(key=lambda x: x["date"])
        return queue
        
    except Exception as e:
        logger.error(f"Failed to fetch review queue: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

# =====================================================================
# PROMPT 7: Continuous Learning Mapped Corrections Route
# =====================================================================

class RecordCorrectionRequest(BaseModel):
    original_raw_text: str
    field_type: str
    corrected_text: str

@app.post("/api/record-correction")
def record_correction(req: RecordCorrectionRequest):
    """
    Exposes an API route to record manual handwriting corrections.
    Updates the corrections_cache.json recurrence counts.
    """
    logger.info(f"Received manual correction: '{req.original_raw_text}' -> '{req.corrected_text}' ({req.field_type})")
    try:
        from learning_cache import HandwritingMemory
        HandwritingMemory.record_correction(req.original_raw_text, req.field_type, req.corrected_text)
        return {
            "status": "success",
            "message": "Correction recorded successfully.",
            "original": req.original_raw_text,
            "corrected": req.corrected_text
        }
    except Exception as e:
        logger.error(f"Failed to record correction: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to record correction: {str(e)}")

# =====================================================================
# PROMPT 19: Operational Reconciliation Endpoints
# =====================================================================

class ReconciliationSaveRequest(BaseModel):
    date: str
    hsd_opening_dip_liters: float = 0.0
    hsd_receipt_liters: float = 0.0
    hsd_closing_dip_liters: float = 0.0
    ms_opening_dip_liters: float = 0.0
    ms_receipt_liters: float = 0.0
    ms_closing_dip_liters: float = 0.0
    actual_cash_deposited: float = 0.0
    digital_wallet_settlements: float = 0.0
    logged_udhaar_entries: float = 0.0

@app.get("/api/reconciliation")
def get_reconciliation_calculations(date: str):
    """
    Computes real-time variance calculations and cash reconciliation results for a date.
    """
    try:
        import reconciliation
        result = reconciliation.calculate_daily_variance(date, db_path=DB_PATH)
        return result
    except Exception as e:
        logger.error(f"Failed to calculate reconciliation for date {date}: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/reconciliation")
def save_reconciliation_calculations(payload: ReconciliationSaveRequest):
    """
    Saves evening dip numbers and cash collection values, and returns real-time calculations.
    """
    try:
        import reconciliation
        reconciliation.save_reconciliation(
            date_str=payload.date,
            hsd_opening=payload.hsd_opening_dip_liters,
            hsd_receipt=payload.hsd_receipt_liters,
            hsd_closing=payload.hsd_closing_dip_liters,
            ms_opening=payload.ms_opening_dip_liters,
            ms_receipt=payload.ms_receipt_liters,
            ms_closing=payload.ms_closing_dip_liters,
            actual_cash=payload.actual_cash_deposited,
            digital_settlements=payload.digital_wallet_settlements,
            udhaar_entries=payload.logged_udhaar_entries,
            db_path=DB_PATH
        )
        result = reconciliation.calculate_daily_variance(payload.date, db_path=DB_PATH)
        return result
    except Exception as e:
        logger.error(f"Failed to save reconciliation for date {payload.date}: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ---------------------------------------------------------------------------
# Mobile Redirect Route
# ---------------------------------------------------------------------------

@app.get("/mobile", include_in_schema=False)
def mobile_page():
    """
    Serves the dedicated mobile capture page directly from the backend
    so any phone on the local Wi-Fi can access it without frontend port bindings.
    """
    mobile_html_path = os.path.abspath(os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "pages", "mobile_capture.html"))
    if not os.path.exists(mobile_html_path):
        mobile_html_path = os.path.abspath(os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "pages", "mobile-capture.html"))
    return FileResponse(mobile_html_path)


# ---------------------------------------------------------------------------
# API Cost Accounting Endpoints
# ---------------------------------------------------------------------------

@app.get("/api/cost-summary")
def api_cost_summary(days: int = 7):
    """
    Returns a breakdown of API token usage and spend for the last `days` days,
    grouped by date and provider.
    """
    try:
        from cost_tracker import get_usage_summary
        rows = get_usage_summary(days=days, db_path=DB_PATH)
        return {"status": "ok", "days": days, "rows": rows}
    except Exception as e:
        logger.error(f"Cost summary query failed: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/budget-status")
def api_budget_status():
    """
    Returns today's cumulative API spend and the configured daily budget cap.
    The frontend uses this to render a spend indicator widget.
    """
    try:
        from cost_tracker import get_today_spend, DAILY_BUDGET_CAP_USD
        spent = get_today_spend(db_path=DB_PATH)
        return {
            "status":          "ok",
            "today_spend_usd": round(spent, 6),
            "daily_cap_usd":   DAILY_BUDGET_CAP_USD,
            "remaining_usd":   round(max(0.0, DAILY_BUDGET_CAP_USD - spent), 6),
            "budget_exceeded": spent >= DAILY_BUDGET_CAP_USD,
        }
    except Exception as e:
        logger.error(f"Budget status query failed: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/analytics/sales-targets")
def api_sales_targets():
    """
    Returns overhead configurations, calculated daily breakeven targets,
    current daily volume progress, and monthly historical profit accumulations.
    """
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # 1. Fetch fixed overhead config
        cursor.execute("SELECT value FROM fixed_overhead_config WHERE key = 'monthly_overhead'")
        row = cursor.fetchone()
        monthly_overhead = row[0] if row else 1500000.0
        
        # 2. Fetch trailing 30 daily summaries to compute average margin per liter
        cursor.execute("PRAGMA table_info(daily_summary)")
        cols = [c[1] for c in cursor.fetchall()]
        
        has_usd = "gross_spread_usd" in cols
        
        # Fetch up to 30 days
        cursor.execute("SELECT date, total_hsd_liters, total_ms_liters, gross_spread_usd FROM daily_summary ORDER BY date DESC LIMIT 30")
        rows = cursor.fetchall()
        
        total_liters = 0.0
        total_spread_inr = 0.0
        
        for r in rows:
            hsd = float(r[1] or 0.0)
            ms = float(r[2] or 0.0)
            usd_spread = float(r[3] or 0.0) if has_usd else 0.0
            
            total_liters += (hsd + ms)
            total_spread_inr += (usd_spread * 83.0)
            
        if total_liters > 0.0:
            average_margin_per_liter = round(total_spread_inr / total_liters, 4)
        else:
            average_margin_per_liter = 4.0 # default fallback
            
        # Target daily volume math
        target_daily_sales_volume = round((monthly_overhead / 30.0) / average_margin_per_liter, 2)
        
        # 3. Fetch latest day's total sales volume
        cursor.execute("SELECT total_hsd_liters, total_ms_liters, date FROM daily_summary ORDER BY date DESC LIMIT 1")
        latest_row = cursor.fetchone()
        
        if latest_row:
            latest_hsd = float(latest_row[0] or 0.0)
            latest_ms = float(latest_row[1] or 0.0)
            latest_date = latest_row[2]
            current_daily_volume = latest_hsd + latest_ms
        else:
            current_daily_volume = 0.0
            latest_date = None
            
        progress_percentage = round((current_daily_volume / target_daily_sales_volume) * 100.0, 2) if target_daily_sales_volume > 0.0 else 0.0
        volume_gap = max(0.0, target_daily_sales_volume - current_daily_volume)
        
        # 4. Live micro-label metric explicitly reading:
        # "Volume Gap: You need to sell 1,240 more Liters of HSD today to hit net operational profitability."
        micro_label = f"Volume Gap: You need to sell {int(round(volume_gap)):,} more Liters of HSD today to hit net operational profitability."
        
        # 5. Monthly profit accumulation over trailing 400 recorded days
        cursor.execute("""
            SELECT strftime('%Y-%m', date) AS month, SUM(realized_profit) AS monthly_profit
            FROM (
                SELECT date, realized_profit FROM daily_summary
                ORDER BY date DESC LIMIT 400
            )
            GROUP BY month
            ORDER BY month ASC
        """)
        history_rows = cursor.fetchall()
        conn.close()
        
        monthly_history = []
        for hr in history_rows:
            monthly_history.append({
                "month": hr[0] or "Unknown",
                "monthly_profit": round(float(hr[1] or 0.0), 2)
            })
            
        return {
            "status": "ok",
            "monthly_overhead": monthly_overhead,
            "average_margin_per_liter": average_margin_per_liter,
            "target_daily_sales_volume": target_daily_sales_volume,
            "current_daily_volume": current_daily_volume,
            "progress_percentage": progress_percentage,
            "volume_gap": round(volume_gap, 2),
            "micro_label": micro_label,
            "monthly_history": monthly_history,
            "latest_date": latest_date
        }
        
    except Exception as e:
        logger.error(f"Sales target analytics query failed: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/analytics/sales-targets/config")
def api_save_sales_targets_config(payload: dict):
    """
    Saves/updates the monthly fixed operational overhead configuration.
    """
    try:
        monthly_overhead = payload.get("monthly_overhead")
        if monthly_overhead is None or float(monthly_overhead) <= 0.0:
            raise HTTPException(status_code=400, detail="Invalid monthly overhead value")
            
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute("""
            INSERT OR REPLACE INTO fixed_overhead_config (key, value)
            VALUES ('monthly_overhead', ?)
        """, (float(monthly_overhead),))
        conn.commit()
        conn.close()
        return {"status": "ok", "message": "Fixed overhead cost updated successfully"}
    except HTTPException as he:
        raise he
    except Exception as e:
        logger.error(f"Save sales targets config failed: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/credit/verify-safety")
def api_verify_credit_safety(party_name: str, amount: float):
    """
    Exposes verify_transaction_credit_safety for frontend check verification.
    """
    try:
        from credit_guard import verify_transaction_credit_safety
        res = verify_transaction_credit_safety(party_name, amount, db_path=DB_PATH)
        return res
    except Exception as e:
        logger.error(f"API verify credit safety failed: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ---------------------------------------------------------------------------
# Stitcher Preview Endpoint
# ---------------------------------------------------------------------------

@app.post("/api/split-preview")
async def api_split_preview(image: UploadFile = File(...)):
    """
    Accepts a single image upload, runs the stitcher spine detection, and
    returns metadata about the detected split — without triggering the AI
    pipeline.  Useful for testing the double-page detection on new scans.

    Returns:
        {
          "split_detected": bool,
          "pages": int,
          "files": ["/processed_images/name_left.jpg", ...]
        }
    """
    try:
        import tempfile
        from stitcher import detect_and_split_double_pages

        # Write upload to a temp file so stitcher can read it by path
        suffix = os.path.splitext(image.filename)[1] or ".jpg"
        with tempfile.NamedTemporaryFile(
            suffix=suffix,
            dir=os.path.join(BACKEND_DIR, "uploaded_raw_photos"),
            delete=False
        ) as tmp:
            tmp.write(await image.read())
            tmp_path = tmp.name

        split_paths = detect_and_split_double_pages(tmp_path)
        split_detected = len(split_paths) > 1

        return {
            "status":         "ok",
            "original":       image.filename,
            "split_detected": split_detected,
            "pages":          len(split_paths),
            "files":          [os.path.basename(p) for p in split_paths],
        }
    except FileNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        logger.error(f"Split preview failed: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ---------------------------------------------------------------------------
# Batch State Machine Endpoints
# ---------------------------------------------------------------------------

@app.get("/api/batch-stats")
def api_batch_stats():
    """
    Returns a count of batch jobs in each status category.
    Used by the frontend progress dashboard.
    """
    try:
        from state_tracker import get_batch_stats
        return {"status": "ok", **get_batch_stats(db_path=DB_PATH)}
    except Exception as e:
        logger.error(f"batch-stats failed: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/batch-jobs")
def api_batch_jobs(status: str = None, limit: int = 200):
    """
    Returns individual job rows from the batch_status table.

    Query params:
        status:  Filter by status string (e.g. 'FAILED_MATH').  Omit for all.
        limit:   Maximum rows to return (default 200).
    """
    try:
        from state_tracker import get_all_jobs, JobStatus
        if status and status.upper() not in JobStatus.all_values():
            raise HTTPException(
                status_code=400,
                detail=f"Invalid status '{status}'. Valid: {JobStatus.all_values()}"
            )
        rows = get_all_jobs(status_filter=status.upper() if status else None,
                            db_path=DB_PATH)
        return {"status": "ok", "count": len(rows), "jobs": rows[:limit]}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"batch-jobs failed: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ---------------------------------------------------------------------------
# Workspace Profile Management Routing
# ---------------------------------------------------------------------------

class SwitchWorkspaceRequest(BaseModel):
    profile_name: str

@app.get("/api/system/active-workspace")
def api_active_workspace():
    try:
        from workspace_manager import get_active_profile
        return {
            "status": "ok",
            "active_profile": get_active_profile()
        }
    except Exception as e:
        logger.error(f"Failed to get active workspace: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/system/switch-workspace")
def api_switch_workspace(req: SwitchWorkspaceRequest):
    try:
        from workspace_manager import switch_active_workspace
        switch_active_workspace(req.profile_name)
        return {
            "status": "ok",
            "active_profile": req.profile_name,
            "database": DB_PATH,
            "excel_path": EXCEL_PATH
        }
    except Exception as e:
        logger.error(f"Failed to switch workspace to {req.profile_name}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


class RecalculateRevenueRequest(BaseModel):
    target_date: str
    product_type: str
    new_correct_rate: float

@app.post("/api/system/recalculate-revenue")
def api_recalculate_revenue(req: RecalculateRevenueRequest):
    try:
        from price_gap_filler import recalculate_ledger_revenue_by_rate
        success = recalculate_ledger_revenue_by_rate(req.target_date, req.product_type, req.new_correct_rate)
        if success:
            return {
                "status": "ok",
                "message": f"Successfully recalculated revenue for {req.product_type} on {req.target_date} with rate {req.new_correct_rate}"
            }
        else:
            raise HTTPException(status_code=500, detail="Recalculation logic failed.")
    except Exception as e:
        logger.error(f"Failed to recalculate revenue: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ---------------------------------------------------------------------------
# Expense Categorization API Endpoints
# ---------------------------------------------------------------------------

@app.get("/api/system/expense-categories")
def api_get_expense_categories():
    """
    Returns the full list of formal accounting heads configured in
    expense_categories.json, plus the fallback 'Unclassified Operational Expenses'.
    Used by the frontend dropdown to stay in sync with backend categories.
    """
    try:
        from expense_mapper import get_all_categories, reload_categories
        reload_categories()   # pick up any live edits to the JSON
        categories = get_all_categories()
        return {"status": "ok", "categories": categories}
    except Exception as e:
        logger.error(f"Failed to fetch expense categories: {e}")
        raise HTTPException(status_code=500, detail=str(e))


class ReclassifyExpenseRequest(BaseModel):
    entry_id: int
    accounting_head: str


@app.post("/api/system/reclassify-expense")
def api_reclassify_expense(req: ReclassifyExpenseRequest):
    """
    Allows an operator to manually overwrite the accounting_head on an existing
    ledger_entries row (type='expense') identified by entry_id.
    Validates the head against the configured category list.
    """
    try:
        from expense_mapper import get_all_categories
        valid_categories = get_all_categories()
        if req.accounting_head not in valid_categories:
            raise HTTPException(
                status_code=400,
                detail=f"Unknown accounting head '{req.accounting_head}'. "
                       f"Valid heads: {valid_categories}"
            )

        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()

        # Verify entry exists and is an expense
        cursor.execute(
            "SELECT entry_id, type FROM ledger_entries WHERE entry_id = ?",
            (req.entry_id,)
        )
        row = cursor.fetchone()
        if not row:
            conn.close()
            raise HTTPException(status_code=404, detail=f"Entry ID {req.entry_id} not found.")
        if row[1] != "expense":
            conn.close()
            raise HTTPException(
                status_code=400,
                detail=f"Entry {req.entry_id} is type '{row[1]}', not 'expense'."
            )

        cursor.execute(
            "UPDATE ledger_entries SET accounting_head = ? WHERE entry_id = ?",
            (req.accounting_head, req.entry_id)
        )
        conn.commit()
        conn.close()

        logger.info(
            f"Expense entry {req.entry_id} reclassified to '{req.accounting_head}'"
        )
        return {
            "status": "ok",
            "entry_id": req.entry_id,
            "accounting_head": req.accounting_head
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to reclassify expense: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ---------------------------------------------------------------------------
# Database Snapshot & Rollback API Endpoints
# ---------------------------------------------------------------------------

@app.get("/api/system/snapshots")
def api_list_snapshots():
    """
    Returns metadata for all available .bak snapshot files in /backend/snapshots/,
    newest first.  Includes filename, size, creation time, and label.
    """
    try:
        from db_rollback import list_available_snapshots
        snapshots = list_available_snapshots()
        return {
            "status": "ok",
            "count":  len(snapshots),
            "snapshots": snapshots
        }
    except Exception as e:
        logger.error(f"Failed to list snapshots: {e}")
        raise HTTPException(status_code=500, detail=str(e))


class CreateSnapshotRequest(BaseModel):
    label: str = "manual"


@app.post("/api/system/create-snapshot")
def api_create_snapshot(req: CreateSnapshotRequest = CreateSnapshotRequest()):
    """
    On-demand manual snapshot.  Copies the active production DB to /backend/snapshots/.
    Useful before any risky manual edit, bulk import, or migration.
    """
    try:
        from db_rollback import create_pre_batch_snapshot
        snap_path = create_pre_batch_snapshot(db_path=DB_PATH, label=req.label)
        snap_name = os.path.basename(snap_path)
        snap_size_kb = os.path.getsize(snap_path) // 1024
        logger.info(f"Manual snapshot created via API: '{snap_name}'")
        return {
            "status":      "ok",
            "message":     f"Snapshot '{snap_name}' created successfully.",
            "snapshot":    snap_name,
            "size_kb":     snap_size_kb,
        }
    except Exception as e:
        logger.error(f"Snapshot creation failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/system/rollback-last-session")
def api_rollback_last_session():
    """
    Automated Reversion Control.

    Safely restores the active production database to the most recently
    created .bak snapshot:
      1. Flushes WAL + closes open SQLite connection handles.
      2. Locates the most-recent .bak file in /backend/snapshots/.
      3. Evicts the current DB (renamed to .rollback_evicted for safety).
      4. Copies the snapshot over and reinitializes connection pools.

    This endpoint is IRREVERSIBLE unless a newer snapshot exists.
    The frontend must show a confirmation dialog before calling this.
    """
    try:
        from db_rollback import rollback_to_last_snapshot
        result = rollback_to_last_snapshot(
            db_path=DB_PATH,
            reinitialize_connections=True,
        )
        logger.info(
            f"[API] Rollback complete. Restored from: '{result['snapshot_used']}'"
        )
        return result
    except RuntimeError as e:
        # Expected errors (no snapshot, file locked, etc.)
        logger.error(f"[API] Rollback failed: {e}")
        raise HTTPException(status_code=409, detail=str(e))
    except Exception as e:
        logger.error(f"[API] Rollback unexpected error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ---------------------------------------------------------------------------
# Error Resolution View — Failed Jobs List + Resolve Endpoints
# ---------------------------------------------------------------------------

FAILURE_STATUSES = ("FAILED_PREPROCESSING", "FAILED_MATH", "FAILED_OCR", "BLURRY")



@app.get("/api/failed-jobs")
def api_failed_jobs(limit: int = 500):
    """
    Returns all batch_status rows whose status is one of the recognised
    failure tags: FAILED_PREPROCESSING, FAILED_MATH, FAILED_OCR, BLURRY.

    Each row is enriched with:
      - image_url: best-effort static URL for the raw image file.
      - ledger_payload: previously saved daily_ledger row (raw_data JSON)
        matched by date prefix extracted from the file name.
    """
    try:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row

        placeholders = ",".join("?" for _ in FAILURE_STATUSES)
        rows = conn.execute(
            f"""
            SELECT image_hash, file_name, status, attempts, last_error,
                   created_at, updated_at
            FROM batch_status
            WHERE status IN ({placeholders})
            ORDER BY updated_at DESC
            LIMIT ?
            """,
            (*FAILURE_STATUSES, limit),
        ).fetchall()

        result = []
        for row in rows:
            d = dict(row)
            fname = d.get("file_name", "")

            # --- Reconstruct image URL from the static mount -----------------
            upload_dir = os.path.join(BACKEND_DIR, "uploaded_raw_photos")
            candidate = os.path.join(upload_dir, fname)
            if os.path.exists(candidate):
                d["image_url"] = f"http://localhost:8000/uploaded_raw_photos/{fname}"
            else:
                # Try without any leading path components
                base = os.path.basename(fname)
                candidate2 = os.path.join(upload_dir, base)
                d["image_url"] = (
                    f"http://localhost:8000/uploaded_raw_photos/{base}"
                    if os.path.exists(candidate2)
                    else None
                )

            # --- Pull the best matching previously-saved ledger payload ------
            # File names are often formatted as YYYY-MM-DD_..., try to extract date
            date_guess = None
            import re as _re
            m = _re.search(r"(\d{4}-\d{2}-\d{2})", fname)
            if m:
                date_guess = m.group(1)

            ledger_payload = None
            if date_guess:
                try:
                    ledger_row = conn.execute(
                        "SELECT raw_data FROM daily_ledger WHERE date = ? LIMIT 1",
                        (date_guess,),
                    ).fetchone()
                    if ledger_row and ledger_row["raw_data"]:
                        raw = ledger_row["raw_data"]
                        try:
                            ledger_payload = json.loads(raw) if isinstance(raw, str) else raw
                            # Decrypt if needed
                            if isinstance(ledger_payload, dict) and ledger_payload.get("_encrypted"):
                                try:
                                    ledger_payload = decrypt_raw_data(ledger_payload)
                                except Exception:
                                    pass
                        except Exception:
                            ledger_payload = None
                except Exception:
                    pass

            d["date_guess"] = date_guess
            d["ledger_payload"] = ledger_payload
            result.append(d)

        conn.close()
        return {"status": "ok", "count": len(result), "jobs": result}

    except Exception as e:
        logger.error(f"failed-jobs endpoint error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


class ResolveJobRequest(BaseModel):
    image_hash: str
    payload: dict   # SaveLedgerDayRequest-shaped dict from the frontend form


@app.post("/api/resolve-job")
def api_resolve_job(req: ResolveJobRequest):
    """
    Accepts a manual-override payload from the ErrorResolutionView form.

    Steps:
      1. Coerce the free-form dict into a SaveLedgerDayRequest and call the
         existing save_ledger_day helper logic (shared code path).
      2. Mark the job as COMPLETED in batch_status.
      3. Return the next still-failing job hash so the UI can auto-advance.
    """
    try:
        from state_tracker import mark_completed

        # --- Build a proper SaveLedgerDayRequest from the incoming dict ------
        try:
            typed_payload = SaveLedgerDayRequest(**req.payload)
        except Exception as val_err:
            raise HTTPException(status_code=422, detail=f"Payload validation error: {val_err}")

        # --- Re-use the existing save_ledger_day handler ---------------------
        # Call it directly (it's a normal synchronous function)
        save_result = save_ledger_day(typed_payload)
        logger.info(f"resolve-job: ledger saved for {typed_payload.date} via override")

        # --- Mark the batch_status row as COMPLETED --------------------------
        mark_completed(req.image_hash, db_path=DB_PATH)
        logger.info(f"resolve-job: marked COMPLETED for hash {req.image_hash[:12]}…")

        # --- Find the next pending failed job --------------------------------
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        placeholders = ",".join("?" for _ in FAILURE_STATUSES)
        next_row = conn.execute(
            f"""
            SELECT image_hash, file_name, status
            FROM batch_status
            WHERE status IN ({placeholders})
              AND image_hash != ?
            ORDER BY updated_at DESC
            LIMIT 1
            """,
            (*FAILURE_STATUSES, req.image_hash),
        ).fetchone()
        conn.close()

        next_hash = dict(next_row)["image_hash"] if next_row else None

        return {
            "status": "ok",
            "resolved_hash": req.image_hash,
            "next_hash": next_hash,
            "save_result": save_result,
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"resolve-job failed for hash {req.image_hash[:12]}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ---------------------------------------------------------------------------
# Daily Summary Text Compiler Endpoint
# ---------------------------------------------------------------------------

@app.get("/api/daily-summary-text")
def api_daily_summary_text(date: str):
    """
    Exposes daily summary text compiler for WhatsApp / SMS clipboard sharing.
    """
    try:
        from summary_bot import compile_whatsapp_sms_draft
        digest = compile_whatsapp_sms_draft(date, db_path=DB_PATH)
        return {"status": "ok", "date": date, "digest": digest}
    except Exception as e:
        logger.error(f"daily-summary-text failed for date {date}: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ---------------------------------------------------------------------------
# Chronological Continuity Interlock Checker Endpoint
# ---------------------------------------------------------------------------

@app.get("/api/interlock-check")
def api_interlock_check(date: str):
    """
    Exposes continuity interlock checker for totalizer meter lock validation.
    """
    try:
        from interlock_checker import verify_chronological_continuity
        result = verify_chronological_continuity(date, db_path=DB_PATH)
        return result
    except Exception as e:
        logger.error(f"interlock-check failed for date {date}: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ---------------------------------------------------------------------------
# Density Register & Compliance Endpoints
# ---------------------------------------------------------------------------

class DensityLogRequest(BaseModel):
    date: str
    product_type: str           # 'HSD' or 'MS'
    observed_temperature_celsius: float
    observed_density_raw: float
    invoice_density_reference: float
    method: str = 'astm'        # 'astm' (default) or 'linear'


@app.post("/api/density-log")
def api_save_density_log(req: DensityLogRequest):
    """
    Logs a daily fuel density reading, runs ASTM conversion,
    and asserts ±3 kg/m³ compliance threshold.
    """
    try:
        result = save_density_record(
            date_str=req.date,
            product_type=req.product_type,
            temp=req.observed_temperature_celsius,
            raw_density=req.observed_density_raw,
            invoice_ref=req.invoice_density_reference,
            db_path=DB_PATH,
            method=req.method
        )

        response = {"status": "ok", **result}

        # Attach high-priority alert tag if compliance failed
        if not result["permissible_variation_passed"]:
            response["density_alert"] = (
                f"⚠️ REGULATORY ALERT: {result['product_type']} density on {result['date']} "
                f"deviates by {result['variation']:+.2f} kg/m³ — exceeds ±3.0 kg/m³ permissible limit!"
            )

        return response
    except ValueError as ve:
        raise HTTPException(status_code=400, detail=str(ve))
    except Exception as e:
        logger.error(f"density-log save failed: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/density-log")
def api_get_density_logs():
    """
    Returns all density compliance records sorted chronologically.
    """
    try:
        records = get_density_records(db_path=DB_PATH)
        return {"status": "ok", "count": len(records), "records": records}
    except Exception as e:
        logger.error(f"density-log fetch failed: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/density-register-pdf")
def api_export_density_register_pdf():
    """
    Generates and streams the Statutory Density Register as a downloadable PDF.
    Matches the standard tabular format verified during IOC/BPCL/HPCL inspections.
    """
    try:
        from exporter import export_density_register_pdf
        pdf_path = export_density_register_pdf()
        return FileResponse(
            path=pdf_path,
            media_type="application/pdf",
            filename="Statutory_Density_Register.pdf"
        )
    except Exception as e:
        logger.error(f"density-register-pdf export failed: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ===========================================================================
# Bank Statement Cross-Referencing Engine Endpoints
# ===========================================================================

import tempfile
import shutil


@app.post("/api/bank-statement/upload")
async def api_upload_bank_statement(
    file: UploadFile = File(...),
    bank_name: str = "generic",
):
    """
    Accepts a bank statement PDF upload, parses it with PyMuPDF,
    and stores all extracted credit/debit rows in SQLite.

    Query params:
      bank_name — hint for parsing pattern: 'SBI', 'HDFC', 'Axis', or 'generic'
    """
    from bank_matcher import (
        parse_bank_statement_pdf,
        save_bank_statement_credits,
    )

    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Only PDF files are accepted.")

    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
        tmp_path = tmp.name
        shutil.copyfileobj(file.file, tmp)

    try:
        transactions = parse_bank_statement_pdf(tmp_path, bank_name=bank_name)
        inserted = save_bank_statement_credits(transactions, db_path=DB_PATH)
        return {
            "status": "ok",
            "bank_name": bank_name,
            "rows_parsed": len(transactions),
            "rows_inserted": inserted,
            "message": (
                f"Successfully extracted {len(transactions)} transaction rows "
                f"from {file.filename}."
            ),
        }
    except Exception as e:
        logger.error(f"Bank statement upload failed: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        try:
            os.unlink(tmp_path)
        except Exception:
            pass


@app.post("/api/bank-statement/reconcile")
def api_reconcile_bank(date: str):
    """
    Runs the cross-matching engine for a specific diary date.
    Stamps each digital entry as SETTLED_IN_BANK or UNSETTLED_MISSING_CASH.
    """
    from bank_matcher import reconcile_diary_against_bank
    try:
        result = reconcile_diary_against_bank(target_date=date, db_path=DB_PATH)
        return {"status": "ok", **result}
    except Exception as e:
        logger.error(f"Bank reconciliation failed for {date}: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/bank-statement/unsettled")
def api_unsettled_digital_entries(limit: int = 200):
    """
    Returns all UNSETTLED_MISSING_CASH diary entries across all dates.
    Feeds the HistoryView discrepancy alert panel.
    """
    from bank_matcher import get_unsettled_digital_entries
    try:
        entries = get_unsettled_digital_entries(db_path=DB_PATH, limit=limit)
        return {"status": "ok", "count": len(entries), "entries": entries}
    except Exception as e:
        logger.error(f"Unsettled entries query failed: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/bank-statement/summary")
def api_bank_settlement_summary():
    """
    Returns aggregate KPI counts and amounts for the bank settlement dashboard.
    """
    from bank_matcher import get_settlement_summary
    try:
        summary = get_settlement_summary(db_path=DB_PATH)
        return {"status": "ok", **summary}
    except Exception as e:
        logger.error(f"Settlement summary query failed: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/bank-statement/credits")
def api_list_bank_credits(
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    limit: int = 500,
):
    """
    Lists raw bank statement credits stored in SQLite, optionally filtered
    by date range. Used for audit cross-reference verification.
    """
    import sqlite3 as _sq
    try:
        conn = _sq.connect(DB_PATH)
        conn.row_factory = _sq.Row
        cursor = conn.cursor()

        query = """
            SELECT id, bank_name, transaction_date, description, utr_string,
                   credit_amount, debit_amount, uploaded_at
            FROM bank_statement_credits
        """
        params: list = []
        conditions: list = []

        if from_date:
            conditions.append("transaction_date >= ?")
            params.append(from_date)
        if to_date:
            conditions.append("transaction_date <= ?")
            params.append(to_date)
        if conditions:
            query += " WHERE " + " AND ".join(conditions)
        query += " ORDER BY transaction_date DESC LIMIT ?"
        params.append(limit)

        cursor.execute(query, params)
        rows = [dict(r) for r in cursor.fetchall()]
        conn.close()

        return {"status": "ok", "count": len(rows), "credits": rows}
    except Exception as e:
        logger.error(f"Bank credits list query failed: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


def background_save_and_reconcile_bank(transactions: List[Dict[str, Any]], db_path: str):
    """
    Background worker that saves the extracted transaction list to the SQLite database
    and runs date-wise reconciliation for each unique date present in the statement.
    """
    from bank_matcher import save_bank_statement_credits, reconcile_diary_against_bank
    try:
        # Save parsed credits to database
        save_bank_statement_credits(transactions, db_path=db_path)
        
        # Identify all unique transaction dates to trigger reconciliation loops
        unique_dates = sorted(list(set(t["transaction_date"] for t in transactions if t.get("transaction_date"))))
        for t_date in unique_dates:
            reconcile_diary_against_bank(target_date=t_date, db_path=db_path)
        logger.info(f"Background bank reconciliation finished successfully for {len(unique_dates)} dates.")
    except Exception as e:
        logger.error(f"Background bank reconciliation failed: {str(e)}")


@app.post("/api/reconcile-bank-statement")
async def api_reconcile_bank_statement(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    bank_name: str = "generic"
):
    """
    Accepts an uploaded official bank statement PDF (SBI/HDFC/generic),
    triggers text extraction synchronously using PyMuPDF, calculates matched
    and missing drops in-memory, schedules database write & reconciliation
    as a background task, and returns the match results immediately.
    """
    from bank_matcher import parse_bank_statement_pdf, calculate_matched_and_missing_in_memory

    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Only PDF files are accepted.")

    # Save to a temporary file
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
        tmp_path = tmp.name
        shutil.copyfileobj(file.file, tmp)

    try:
        # 1. Parse statement synchronously
        transactions = parse_bank_statement_pdf(tmp_path, bank_name=bank_name)
        
        # 2. Calculate matched hashes and missing drops in-memory
        matched_hashes, missing_drops = calculate_matched_and_missing_in_memory(transactions, db_path=DB_PATH)
        
        # 3. Schedule the database updates and reconciliations in the background
        background_tasks.add_task(background_save_and_reconcile_bank, transactions, DB_PATH)
        
        return {
            "status": "ok",
            "matched_hashes": matched_hashes,
            "missing_drops": missing_drops
        }
    except Exception as e:
        logger.error(f"Bank statement cross-referencing endpoint failed: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        # Clean up temporary file
        try:
            os.unlink(tmp_path)
        except Exception:
            pass


# ---------------------------------------------------------------------------
# Credit Account Invoicing & Statement Compiler Endpoint
# ---------------------------------------------------------------------------

@app.get("/api/customer-invoice/generate")
def api_generate_customer_invoice(
    party_name: str,
    start_date: str,
    end_date: str,
):
    """
    Triggers generation of a custom PDF invoice for the selected credit party.
    """
    from invoice_generator import generate_customer_invoice
    try:
        pdf_path = generate_customer_invoice(party_name, start_date, end_date, db_path=DB_PATH)
        filename = os.path.basename(pdf_path)
        return FileResponse(
            pdf_path,
            media_type="application/pdf",
            filename=filename
        )
    except Exception as e:
        logger.error(f"Invoicing failed for customer {party_name}: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ---------------------------------------------------------------------------
# Minimalist Local Trend Visualization Endpoint
# ---------------------------------------------------------------------------

@app.post("/api/analytics/regenerate")
def api_regenerate_analytics():
    """
    Triggers manual background regeneration of all Matplotlib operational charts.
    """
    from local_analytics import generate_all_charts
    try:
        paths = generate_all_charts(db_path=DB_PATH)
        return {
            "status": "success",
            "message": "Matplotlib static charts generated successfully.",
            "charts": {k: os.path.basename(v) for k, v in paths.items()}
        }
    except Exception as e:
        logger.error(f"Failed to regenerate static trend charts: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ---------------------------------------------------------------------------
# Credit Account Realization and Settlement Endpoints
# ---------------------------------------------------------------------------

class CreditRealizationRequest(BaseModel):
    date: str
    party_name: str
    amount_received: float
    payment_mode: str
    bank_utr_or_remarks: str = None
    linked_invoice_no: str = None

@app.post("/api/credit-realizations")
def api_save_credit_realization(payload: CreditRealizationRequest):
    """
    Manually logs a credit payment realization and settles customer balance.
    """
    from credit_realization import save_credit_realization
    try:
        rid = save_credit_realization(
            date_str=payload.date,
            party_name=payload.party_name,
            amount_received=payload.amount_received,
            payment_mode=payload.payment_mode,
            bank_utr_or_remarks=payload.bank_utr_or_remarks,
            linked_invoice_no=payload.linked_invoice_no,
            db_path=DB_PATH
        )
        return {"status": "success", "realization_id": rid, "message": "Credit realization recorded and ledger adjusted."}
    except Exception as e:
        logger.error(f"Failed to save credit realization: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/credit-realizations")
def api_get_credit_realizations():
    """
    Returns a running list of all credit payment realizations.
    """
    from credit_realization import get_all_realizations
    try:
        realizations = get_all_realizations(db_path=DB_PATH)
        return {"status": "success", "realizations": realizations}
    except Exception as e:
        logger.error(f"Failed to get credit realizations: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/logs/stream")
def api_stream_logs():
    """
    Streams the last 100 lines of the local pipeline log file.
    """
    from logger import LOG_FILE
    try:
        if not os.path.exists(LOG_FILE):
            return {"status": "success", "logs": []}
        with open(LOG_FILE, "r", encoding="utf-8") as f:
            lines = f.readlines()
        # Take the last 100 lines, strip newlines
        last_100 = [line.rstrip("\n") for line in lines[-100:]]
        return {"status": "success", "logs": last_100}
    except Exception as e:
        logger.error(f"Failed to stream diagnostic logs: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/logs/clear")
def api_clear_logs():
    """
    Clears the local pipeline log file, truncating it cleanly.
    """
    from logger import LOG_FILE
    try:
        if os.path.exists(LOG_FILE):
            with open(LOG_FILE, "w", encoding="utf-8") as f:
                f.truncate(0)
        return {"status": "success", "message": "Diagnostic history cleared successfully."}
    except Exception as e:
        logger.error(f"Failed to clear diagnostic logs: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/usb-sync/events")
async def api_usb_sync_events():
    """
    Server-Sent Events endpoint streaming USB sync state updates.
    Yields the current status immediately on connection.
    """
    from fastapi.responses import StreamingResponse
    from usb_sync import register_subscriber, unregister_subscriber, get_last_sync_status
    import asyncio
    import json

    async def event_generator():
        # 1. Yield current completed state immediately upon client connection
        current_status = get_last_sync_status()
        if current_status:
            yield f"data: {json.dumps(current_status)}\n\n"

        # 2. Listen to subsequent updates from background thread
        queue = asyncio.Queue()
        register_subscriber(queue)
        try:
            while True:
                data = await queue.get()
                yield f"data: {json.dumps(data)}\n\n"
        except asyncio.CancelledError:
            pass
        finally:
            unregister_subscriber(queue)

    return StreamingResponse(event_generator(), media_type="text/event-stream")


# ─────────────────────────────────────────────────────────────
# Thermal Receipt Printer Integration
# ─────────────────────────────────────────────────────────────

class PrintSlipRequest(BaseModel):
    party_name: str
    dry_run: Optional[bool] = False

@app.post("/api/print-slip")
def api_print_slip(req: PrintSlipRequest):
    """
    Fetches all credit ledger transactions for a customer,
    calculates the net outstanding balance, and dispatches
    a formatted receipt slip to the ESC/POS thermal printer.
    Falls back to saving an ASCII dry-run file on disk.
    """
    from thermal_printer import print_credit_ledger_slip
    
    party_name = req.party_name.strip()
    if not party_name:
        raise HTTPException(status_code=400, detail="party_name is required.")
    
    logger.info(f"[ThermalPrint] Compiling receipt slip for customer: {party_name}")
    
    BACKEND_DIR_LOCAL = os.path.dirname(os.path.abspath(__file__))
    db_path = os.path.join(BACKEND_DIR_LOCAL, "ledger.db")
    
    try:
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        cursor.execute("""
            SELECT date, party_name, vehicle_wheel_no, amount, type, remarks
            FROM ledger_entries
            ORDER BY date ASC, entry_id ASC
        """)
        rows = cursor.fetchall()
        conn.close()
    except Exception as db_err:
        logger.error(f"[ThermalPrint] Database read failed: {str(db_err)}")
        raise HTTPException(status_code=500, detail=f"Database error: {str(db_err)}")
    
    target_clean = party_name.strip().lower()
    transactions = []
    net_due = 0.0
    
    for row in rows:
        r_date = row[0]
        r_party_enc = row[1]
        r_vehicle = row[2] or "N/A"
        r_amount_enc = row[3]
        r_type = row[4]
        r_remarks = row[5] or ""
        
        # Decrypt fields
        try:
            r_party = decrypt_field(r_party_enc, return_type=str)
            r_amount = decrypt_field(r_amount_enc, return_type=float)
        except Exception:
            r_party = str(r_party_enc or "")
            r_amount = float(r_amount_enc or 0.0)
        
        if not r_party or r_party.strip().lower() != target_clean:
            continue
        
        # Accumulate running balance
        if r_type == "udhaar":
            net_due += r_amount
        elif r_type in ("payment", "deposit", "receipt"):
            net_due -= r_amount
        elif r_amount < 0:
            net_due += r_amount
        else:
            net_due += r_amount
        
        transactions.append({
            "date": r_date,
            "vehicle_no": r_vehicle,
            "amount": r_amount,
            "type": r_type,
            "remarks": r_remarks
        })
    
    if not transactions:
        logger.warning(f"[ThermalPrint] No ledger entries found for customer: {party_name}")
        raise HTTPException(status_code=404, detail=f"No ledger entries found for customer: {party_name}")
    
    # Dispatch to thermal printer driver
    result = print_credit_ledger_slip(
        party_name=party_name,
        transactions=transactions,
        net_due=net_due,
        dry_run=req.dry_run
    )
    
    logger.info(f"[ThermalPrint] Slip result: dry_run={result.get('dry_run')}, printer={result.get('printer')}")
    return result


@app.get("/api/printer-status")
def api_printer_status():
    """
    Returns the current thermal printer discovery status for the frontend UI.
    """
    from thermal_printer import connect_thermal_printer
    return connect_thermal_printer()


# ─────────────────────────────────────────────────────────────
# FIFO Credit Balancing Engine
# ─────────────────────────────────────────────────────────────

class FIFOSettleRequest(BaseModel):
    party_name: str
    payment_amount: float
    payment_date: str

@app.post("/api/fifo-settle")
def api_fifo_settle(req: FIFOSettleRequest):
    """
    Runs the FIFO credit settlement algorithm against a customer's outstanding debts.
    Allocates payment oldest-first, splitting partial rows as needed.
    """
    from fifo_settler import allocate_realization_fifo
    try:
        result = allocate_realization_fifo(
            party_name=req.party_name,
            payment_amount=req.payment_amount,
            payment_date=req.payment_date,
            db_path=DB_PATH
        )
        return result
    except Exception as e:
        logger.error(f"FIFO settlement failed: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/fifo-status")
def api_fifo_status(party_name: str):
    """
    Returns the FIFO settlement status for a specific customer account.
    """
    from fifo_settler import get_customer_fifo_status
    try:
        status = get_customer_fifo_status(party_name=party_name, db_path=DB_PATH)
        return {"status": "success", **status}
    except Exception as e:
        logger.error(f"FIFO status query failed: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ─────────────────────────────────────────────────────────────
# Credit Ledger Aging Analysis & Risk Evaluation Engine
# ─────────────────────────────────────────────────────────────

@app.get("/api/aging/customer")
def api_aging_customer(party_name: str, reference_date: Optional[str] = None):
    """
    Returns the credit ledger aging breakdown and risk profile for a specific customer account.
    """
    from aging_analysis import compile_customer_debt_aging
    try:
        result = compile_customer_debt_aging(
            party_name=party_name,
            reference_date=reference_date,
            db_path=DB_PATH
        )
        return result
    except Exception as e:
        logger.error(f"Aging customer query failed for {party_name}: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/aging/summary")
def api_aging_summary(reference_date: Optional[str] = None):
    """
    Returns the aggregated aging and risk profiles for all unique active credit customers.
    """
    from aging_analysis import compile_all_customers_aging
    try:
        results = compile_all_customers_aging(
            reference_date=reference_date,
            db_path=DB_PATH
        )
        return {"status": "success", "count": len(results), "summary": results}
    except Exception as e:
        logger.error(f"Aging summary query failed: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/aging/export-pdf")
def api_aging_export_pdf(reference_date: Optional[str] = None):
    """
    Generates and streams the Credit Aging & Risk Statement PDF.
    """
    from aging_analysis import export_aging_summary_pdf
    import datetime
    try:
        pdf_path = export_aging_summary_pdf(
            output_path=None,
            reference_date=reference_date,
            db_path=DB_PATH
        )
        ref_date = reference_date or datetime.date.today().strftime("%Y-%m-%d")
        filename = f"credit_aging_summary_{ref_date}.pdf"
        return FileResponse(
            path=pdf_path,
            media_type="application/pdf",
            filename=filename
        )
    except Exception as e:
        logger.error(f"Aging PDF export failed: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ─────────────────────────────────────────────────────────────
# Staff Cash Advance & Salary Deduction Registry Endpoints
# ─────────────────────────────────────────────────────────────

class StaffAdvanceRequest(BaseModel):
    date: str
    employee_name: str
    amount_drawn: float
    type: str  # 'CASH_ADVANCE' or 'FUEL_DRAWN'
    remarks: str = ""
    settlement_status: str = "PENDING_DEDUCTION"

class StaffAdvanceSettleRequest(BaseModel):
    advance_id: int
    settlement_status: str  # 'PENDING_DEDUCTION' or 'SETTLED_FROM_SALARY'

@app.post("/api/staff-advances")
def api_record_staff_advance(req: StaffAdvanceRequest):
    """
    Logs an employee cash advance or fuel drawing in the bookkeeping ledger.
    """
    from staff_ledger import record_staff_advance
    try:
        adv_id = record_staff_advance(
            date=req.date,
            employee_name=req.employee_name,
            amount_drawn=req.amount_drawn,
            atype=req.type,
            remarks=req.remarks,
            settlement_status=req.settlement_status,
            db_path=DB_PATH
        )
        return {"status": "success", "advance_id": adv_id}
    except Exception as e:
        logger.error(f"Failed to record staff advance: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/staff-advances")
def api_get_daily_staff_advances(date: str):
    """
    Retrieves all staff advances logged on a specific operations date.
    """
    from staff_ledger import get_daily_staff_advances
    try:
        advances = get_daily_staff_advances(date_str=date, db_path=DB_PATH)
        return {"status": "success", "date": date, "advances": advances}
    except Exception as e:
        logger.error(f"Failed to fetch daily staff advances: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/staff-advances/settle")
def api_settle_staff_advance(req: StaffAdvanceSettleRequest):
    """
    Updates the salary deduction status of a staff drawing.
    """
    from staff_ledger import settle_staff_advance
    try:
        settle_staff_advance(
            advance_id=req.advance_id,
            new_status=req.settlement_status,
            db_path=DB_PATH
        )
        return {"status": "success", "advance_id": req.advance_id, "settlement_status": req.settlement_status}
    except Exception as e:
        logger.error(f"Failed to settle staff advance: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/staff-advances/payroll")
def api_generate_payroll_deductions(employee_name: str, target_month: str):
    """
    Compiles and aggregates outstanding advances into a payroll deduction receipt for a month.
    """
    from staff_ledger import generate_monthly_payroll_deductions
    try:
        report = generate_monthly_payroll_deductions(
            employee_name=employee_name,
            target_month=target_month,
            db_path=DB_PATH
        )
        return report
    except Exception as e:
        logger.error(f"Failed to compile payroll deductions: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# Corporate Fleet Card Transaction Reconciliation Module Endpoints
# ===========================================================================

@app.post("/api/fleet/import")
async def api_import_fleet_portal(
    file: UploadFile = File(...),
    provider: str = "IOCL"
):
    """
    Accepts a raw oil company fleet portal statement CSV upload,
    and stores all extracted transaction rows in SQLite.
    """
    from fleet_cards import import_fleet_portal_csv
    if not file.filename.lower().endswith(".csv"):
        raise HTTPException(status_code=400, detail="Only CSV files are accepted.")

    with tempfile.NamedTemporaryFile(suffix=".csv", delete=False) as tmp:
        tmp_path = tmp.name
        shutil.copyfileobj(file.file, tmp)

    try:
        imported = import_fleet_portal_csv(tmp_path, provider, db_path=DB_PATH)
        return {
            "status": "success",
            "provider": provider,
            "rows_imported": imported,
            "message": f"Successfully imported {imported} fleet portal records from {file.filename}."
        }
    except Exception as e:
        logger.error(f"Fleet portal import failed: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        try:
            os.unlink(tmp_path)
        except Exception:
            pass


@app.post("/api/fleet/reconcile")
def api_reconcile_fleet(date: str):
    """
    Runs the automated cross-matching loop comparing notebook credit sales against portal statements.
    """
    from fleet_cards import reconcile_fleet_transactions
    try:
        result = reconcile_fleet_transactions(target_date=date, db_path=DB_PATH)
        return result
    except Exception as e:
        logger.error(f"Fleet reconciliation failed for {date}: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/fleet/status")
def api_get_fleet_status(date: str):
    """
    Returns the compiled fleet card reconciliation status records for a specific date.
    """
    from fleet_cards import get_fleet_reconciliation_status
    try:
        status_records = get_fleet_reconciliation_status(date_str=date, db_path=DB_PATH)
        return {"status": "success", "date": date, "records": status_records}
    except Exception as e:
        logger.error(f"Failed to fetch fleet status for {date}: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# Coordinated Fuel Tanker Decanting & transit shortages auditor Endpoints
# ===========================================================================

class TankerReceiptSaveRequest(BaseModel):
    invoice_no: str
    date: str
    tank_lorry_no: str
    product_type: str
    invoice_volume_liters: float
    invoice_density_at_15c: float
    observed_compartment_dips_mm: str
    observed_density_raw: float
    observed_temperature_celsius: float
    raw_observed_volume_liters: Optional[float] = None
    current_dip_mm: Optional[float] = None


@app.post("/api/decanting/receipt")
def api_save_tanker_receipt(req: TankerReceiptSaveRequest):
    """
    Saves tanker receipt, performs ambient-to-15C ASTM standard temperature corrections,
    checks underground ground capacity safety, and records transit shortage claims ledger.
    """
    from decanting_auditor import save_tanker_receipt
    try:
        res = save_tanker_receipt(
            invoice_no=req.invoice_no,
            date_str=req.date,
            tank_lorry_no=req.tank_lorry_no,
            product_type=req.product_type,
            invoice_volume_liters=req.invoice_volume_liters,
            invoice_density_at_15c=req.invoice_density_at_15c,
            observed_compartment_dips_mm=req.observed_compartment_dips_mm,
            observed_density_raw=req.observed_density_raw,
            observed_temperature_celsius=req.observed_temperature_celsius,
            raw_observed_volume_liters=req.raw_observed_volume_liters,
            current_dip_mm=req.current_dip_mm,
            db_path=DB_PATH
        )
        return res
    except Exception as e:
        logger.error(f"Failed to save tanker receipt: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/decanting/receipts")
def api_get_tanker_receipts():
    """
    Queries and returns the chronological tanker receipts audit log list.
    """
    from decanting_auditor import get_tanker_receipts
    try:
        receipts = get_tanker_receipts(db_path=DB_PATH)
        return {"status": "success", "receipts": receipts}
    except Exception as e:
        logger.error(f"Failed to fetch tanker receipts: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/decanting/check-space")
def api_check_decanting_space(
    product_type: str,
    actual_received_volume: float,
    current_dip_mm: Optional[float] = None
):
    """
    Runs a pre-decanting ground capacity safety verification check.
    """
    from decanting_auditor import validate_decanting_space
    try:
        res = validate_decanting_space(
            product_type=product_type,
            actual_received_volume=actual_received_volume,
            current_dip_mm=current_dip_mm,
            db_path=DB_PATH
        )
        return res
    except Exception as e:
        logger.error(f"Failed to validate decanting space: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# Monthly Lubricant Stock Book Reconciliation Endpoints
# ===========================================================================

class LubeReconcileRequest(BaseModel):
    item_sku: str
    target_month: str

class LubePhysicalCountRequest(BaseModel):
    item_sku: str
    current_physical_count: float

@app.get("/api/lube/ledger")
def api_get_lube_ledger():
    """
    Retrieves the entire lubricant inventory stock ledger.
    """
    from lube_stock_book import get_lube_inventory_ledger
    try:
        ledger = get_lube_inventory_ledger(db_path=DB_PATH)
        return {"status": "success", "ledger": ledger}
    except Exception as e:
        logger.error(f"Failed to fetch lubricant stock ledger: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/lube/reconcile")
def api_reconcile_lube_item(req: LubeReconcileRequest):
    """
    Runs monthly sales rollup and expected stock calculations for a specific lubricant item.
    """
    from lube_stock_book import compute_running_lube_book
    try:
        res = compute_running_lube_book(
            item_sku=req.item_sku,
            target_month=req.target_month,
            db_path=DB_PATH
        )
        return res
    except Exception as e:
        logger.error(f"Failed to compute lube book reconciliation: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/lube/physical-count")
def api_commit_lube_physical_count(req: LubePhysicalCountRequest):
    """
    Saves physical shelf inventory count and updates delta shortage variance.
    """
    from lube_stock_book import commit_physical_stock_count
    try:
        res = commit_physical_stock_count(
            item_sku=req.item_sku,
            current_physical_count=req.current_physical_count,
            db_path=DB_PATH
        )
        return res
    except Exception as e:
        logger.error(f"Failed to commit lube physical stock count: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ---------------------------------------------------------------------------
# Local Network Multi-Device Asset Upload Receiver (Prompt 100)
# ---------------------------------------------------------------------------

class ProcessSyncFileRequest(BaseModel):
    filename: str

@app.post("/api/sync/lan-upload")
async def post_lan_sync_upload(image: UploadFile = File(...)):
    """
    Ingests multi-device media streams, computes unique SHA-256 signature,
    saves the file to the raw backlog folder, and broadcasts a real-time event.
    """
    try:
        from lan_sync import save_to_backlog, sse_manager
        file_bytes = await image.read()
        filename, file_hash, file_size = save_to_backlog(file_bytes, original_filename=image.filename)
        
        # Broadcast real-time upload event
        event_payload = {
            "event": "NEW_UPLOAD",
            "filename": filename,
            "hash": file_hash,
            "size": file_size,
            "url": f"/raw_backlog/{filename}"
        }
        await sse_manager.broadcast(event_payload)
        
        return {
            "status": "success",
            "message": "Asset synced successfully",
            "filename": filename,
            "hash": file_hash,
            "size": file_size
        }
    except Exception as e:
        logger.error(f"Failed to ingest LAN upload: {str(e)}")
        raise HTTPException(status_code=500, detail=f"LAN Sync Ingest Error: {str(e)}")

@app.get("/api/sync/events")
async def get_lan_sync_events():
    """
    SSE stream for notifying connected laptop desktop clients in real time.
    """
    from lan_sync import sse_manager
    from fastapi.responses import StreamingResponse
    
    async def sse_event_generator():
        q = sse_manager.add_listener()
        try:
            # Yield connection establishment comment to prevent immediate client timeouts
            yield "comment: connection established\n\n"
            while True:
                data = await q.get()
                yield f"data: {json.dumps(data)}\n\n"
        except asyncio.CancelledError:
            pass
        finally:
            sse_manager.remove_listener(q)
            
    headers = {
        "Cache-Control": "no-cache",
        "Connection": "keep-alive",
        "X-Accel-Buffering": "no"
    }
    return StreamingResponse(sse_event_generator(), media_type="text/event-stream", headers=headers)

@app.post("/api/sync/lan-process")
def post_lan_process_sync_file(req: ProcessSyncFileRequest):
    """
    Processes a file already stored in the raw backlog directory using the 
    AI engine and returns audited JSON, indexing transcription and continuity checks.
    """
    import cv2
    from lan_sync import BACKLOG_DIR
    
    filename = os.path.basename(req.filename)
    filepath = os.path.join(BACKLOG_DIR, filename)
    
    if not os.path.exists(filepath):
        raise HTTPException(status_code=404, detail="Sync file not found in backlog.")
        
    try:
        # Read the file
        cv2_matrix = cv2.imread(filepath)
        if cv2_matrix is None:
            raise HTTPException(status_code=400, detail="Failed to parse image from file.")
            
        from intake import generate_matrix_hash
        page_hash = generate_matrix_hash(cv2_matrix)
        
        # Save raw copy in uploaded_raw_photos for static access compatibility
        upload_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "uploaded_raw_photos")
        os.makedirs(upload_dir, exist_ok=True)
        
        ext = os.path.splitext(filename)[1] or ".jpg"
        local_filename = f"raw_{page_hash}{ext}"
        local_path = os.path.join(upload_dir, local_filename)
        cv2.imwrite(local_path, cv2_matrix)
        
        # Double-page spread detection & splitting
        from stitcher import detect_and_split_double_pages
        split_paths = detect_and_split_double_pages(local_path)
        
        results = []
        for split_idx, split_path in enumerate(split_paths):
            from image_guard import validate_image_clarity
            from state_tracker import calculate_file_hash, upsert_job, mark_preprocessing_failed
            from archiver import archive_processed_file
            
            clarity_res = validate_image_clarity(split_path)
            if not clarity_res["success"]:
                split_hash = calculate_file_hash(split_path)
                split_filename = os.path.basename(split_path)
                
                upsert_job(split_hash, split_filename, db_path=DB_PATH)
                reason_msg = f"Clarity check failed: {clarity_res['status']} (Focus score: {clarity_res['focus_score']:.2f}, Contrast: {clarity_res['contrast_score']:.2f})"
                mark_preprocessing_failed(split_hash, reason_msg, db_path=DB_PATH)
                archived_path = archive_processed_file(split_path, status="FAILED_PREPROCESSING")
                
                failed_json = {
                    "date": "N/A",
                    "validation_status": "FAILED_PREPROCESSING",
                    "mathematical_warnings": [f"FAILED_PREPROCESSING: Clarity check failed: {clarity_res['status']}"],
                    "image_url": f"http://localhost:8000/uploaded_raw_photos/{local_filename}",
                    "split_image_url": f"http://localhost:8000/uploaded_raw_photos/{os.path.basename(archived_path)}",
                    "page_hash": split_hash,
                    "page_index": 0,
                    "split_index": split_idx,
                    "split_total": len(split_paths),
                    "original_filename": filename,
                    "error": f"Image clarity check failed: {clarity_res['status']}"
                }
                results.append(failed_json)
                continue
                
            from processor import optimize_register_image
            optimized_path = optimize_register_image(split_path)
            
            from ai_engine import analyze_register_sheet
            audited_json = analyze_register_sheet(optimized_path)
            
            # FTS Text Indexing
            try:
                raw_text = audited_json.get("raw_transcription_text")
                date_val = audited_json.get("date")
                if date_val and date_val != "N/A" and raw_text:
                    from text_search import index_ledger_text
                    index_ledger_text(DB_PATH, date_val, raw_text)
            except Exception as fts_err:
                logger.warning(f"FTS indexing failed for sync file: {fts_err}")
                
            # Chronological Continuity check
            try:
                from interlock_checker import verify_chronological_continuity
                date_extracted = audited_json.get("date")
                nozzles_extracted = audited_json.get("nozzles") or []
                if date_extracted:
                    continuity = verify_chronological_continuity(
                        target_date_string=date_extracted,
                        current_nozzles=nozzles_extracted,
                        db_path=DB_PATH
                    )
                    audited_json["continuity_check"] = continuity
            except Exception as continuity_err:
                logger.warning(f"Chronological check failed for sync file: {continuity_err}")
                
            # Inject static retrieval elements
            split_filename = os.path.basename(split_path)
            audited_json["image_url"] = f"http://localhost:8000/uploaded_raw_photos/{local_filename}"
            audited_json["split_image_url"] = f"http://localhost:8000/processed_images/{split_filename}"
            audited_json["page_hash"] = calculate_file_hash(split_path)
            audited_json["page_index"] = 0
            audited_json["split_index"] = split_idx
            audited_json["split_total"] = len(split_paths)
            audited_json["original_filename"] = filename
            
            results.append(audited_json)
            
        if not results:
            raise HTTPException(status_code=400, detail="No valid sheets parsed from sync file.")
            
        try:
            from db_hardener import execute_wal_checkpoint
            execute_wal_checkpoint()
        except Exception as checkpoint_err:
            logger.warning(f"WAL checkpoint failed after sync processing: {checkpoint_err}")
            
        if len(results) == 1:
            return results[0]
        else:
            return results
    except Exception as e:
        logger.error(f"Failed to process sync file {filename}: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))






