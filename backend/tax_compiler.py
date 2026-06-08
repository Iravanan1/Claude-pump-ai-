"""
Statutory Tax and Filing Compiler Pre-Formatter.

Hardcodes local VAT and GST statutory tax templates, queries monthly operational fuel 
volumes and lubricant inventory sales, and compiles a pre-formatted Excel returns template.
"""

import os
import sqlite3
import logging
from typing import Dict, Any, List
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger("TaxCompiler")

BACKEND_DIR = os.path.dirname(os.path.abspath(__file__))
WORKSPACE_DIR = os.path.dirname(BACKEND_DIR)
DB_PATH = os.path.join(BACKEND_DIR, "ledger.db")
DEFAULT_EXPORTS_DIR = os.path.join(WORKSPACE_DIR, "pump_exports", "tax_filings")

# Hardcoded State Statutory Reference Tax Rates
HSD_VAT_RATE = 0.1675  # 16.75% local state VAT for high speed diesel
MS_VAT_RATE = 0.1948   # 19.48% local state VAT for motor spirit (petrol)
LUBE_GST_RATE = 0.18   # 18.00% inclusive GST for lubricant inventory products


def compile_monthly_tax_summary(year: int, month: int, db_path: str = DB_PATH) -> Dict[str, Any]:
    """
    Parses SQLite historical tables across the calendar month and aggregates
    unrounded VAT and GSTR-1 GST statutory tax obligation breakdowns.
    """
    date_pattern = f"{year:04d}-{month:02d}-%"
    logger.info(f"Aggregating monthly tax details for calendar window: {year}-{month:02d}")
    
    # ── Segment 1: Fuel VAT Turnover Calculations ──
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute("""
        SELECT date, total_hsd_liters, total_ms_liters 
        FROM daily_summary 
        WHERE date LIKE ? 
        ORDER BY date ASC
    """, (date_pattern,))
    summary_rows = cursor.fetchall()
    
    fuel_records = []
    for row in summary_rows:
        date_str = row[0]
        hsd_liters = float(row[1] or 0.0)
        ms_liters = float(row[2] or 0.0)
        
        # Retrieve fuel rates for the date
        cursor.execute("SELECT hsd_rate, ms_rate FROM fuel_rates WHERE date = ?", (date_str,))
        rate_row = cursor.fetchone()
        
        # Backward-compatible reference rate fallbacks
        hsd_rate = float(rate_row[0] or 94.27) if rate_row else 94.27
        ms_rate = float(rate_row[1] or 106.31) if rate_row else 106.31
        
        # Calculations (Unrounded raw decimals for practitioner audit audits)
        hsd_turnover = hsd_liters * hsd_rate
        hsd_vat = hsd_turnover * HSD_VAT_RATE
        
        ms_turnover = ms_liters * ms_rate
        ms_vat = ms_turnover * MS_VAT_RATE
        
        total_vat = hsd_vat + ms_vat
        
        fuel_records.append({
            "Date": date_str,
            "Diesel (HSD) Liters": hsd_liters,
            "Diesel (HSD) Rate (INR)": hsd_rate,
            "Diesel (HSD) Turnover (INR)": hsd_turnover,
            "Diesel (HSD) VAT (INR)": hsd_vat,
            "Petrol (MS) Liters": ms_liters,
            "Petrol (MS) Rate (INR)": ms_rate,
            "Petrol (MS) Turnover (INR)": ms_turnover,
            "Petrol (MS) VAT (INR)": ms_vat,
            "Total VAT Obligation (INR)": total_vat
        })
        
    # ── Segment 2: Lubricants GST GSTR-1 Breakdown ──
    # Ensure inventory_sales table exists
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS inventory_sales (
        date TEXT,
        item_name TEXT,
        quantity_sold REAL DEFAULT 0.0,
        unit_price REAL DEFAULT 0.0,
        total_item_revenue REAL DEFAULT 0.0,
        UNIQUE(date, item_name)
    )
    """)
    
    cursor.execute("""
        SELECT date, item_name, quantity_sold, unit_price, total_item_revenue 
        FROM inventory_sales 
        WHERE date LIKE ? 
        ORDER BY date ASC
    """, (date_pattern,))
    lube_rows = cursor.fetchall()
    conn.close()
    
    lube_records = []
    for row in lube_rows:
        date_str = row[0]
        item_name = str(row[1] or "")
        qty = float(row[2] or 0.0)
        price = float(row[3] or 0.0)
        gross = float(row[4] or 0.0)
        
        # Standard fallback if total revenue isn't recorded directly
        if gross == 0.0:
            gross = qty * price
            
        # Strip 18.00% inclusive GST to isolate CGST/SGST/Base Taxable Value
        # base_value = gross / 1.18
        # total_gst = gross - base_value
        base_value = gross / (1.0 + LUBE_GST_RATE)
        total_gst = gross - base_value
        cgst = total_gst / 2.0
        sgst = total_gst / 2.0
        
        lube_records.append({
            "Date": date_str,
            "Lubricant Item Name": item_name,
            "Quantity Sold": qty,
            "Unit Price (INR)": price,
            "Gross Revenue (INR)": gross,
            "Base Taxable Value (INR)": base_value,
            "CGST (9%) (INR)": cgst,
            "SGST (9%) (INR)": sgst,
            "Total GST (18%) (INR)": total_gst
        })
        
    # Compute unrounded totals for statutory verification
    summary = {
        "year": year,
        "month": month,
        "fuel_vat_rate_diesel": HSD_VAT_RATE,
        "fuel_vat_rate_petrol": MS_VAT_RATE,
        "lube_gst_rate": LUBE_GST_RATE,
        
        # Totals
        "total_hsd_liters": sum(x["Diesel (HSD) Liters"] for x in fuel_records),
        "total_hsd_turnover": sum(x["Diesel (HSD) Turnover (INR)"] for x in fuel_records),
        "total_hsd_vat": sum(x["Diesel (HSD) VAT (INR)"] for x in fuel_records),
        
        "total_ms_liters": sum(x["Petrol (MS) Liters"] for x in fuel_records),
        "total_ms_turnover": sum(x["Petrol (MS) Turnover (INR)"] for x in fuel_records),
        "total_ms_vat": sum(x["Petrol (MS) VAT (INR)"] for x in fuel_records),
        
        "total_fuel_vat_obligation": sum(x["Total VAT Obligation (INR)"] for x in fuel_records),
        
        "total_lube_qty": sum(x["Quantity Sold"] for x in lube_records),
        "total_lube_gross": sum(x["Gross Revenue (INR)"] for x in lube_records),
        "total_lube_base": sum(x["Base Taxable Value (INR)"] for x in lube_records),
        "total_lube_cgst": sum(x["CGST (9%) (INR)"] for x in lube_records),
        "total_lube_sgst": sum(x["SGST (9%) (INR)"] for x in lube_records),
        "total_lube_gst": sum(x["Total GST (18%) (INR)"] for x in lube_records),
        
        "fuel_records": fuel_records,
        "lube_records": lube_records
    }
    
    return summary


def export_tax_filing_template(
    year: int, 
    month: int, 
    db_path: str = DB_PATH, 
    export_dir: str = DEFAULT_EXPORTS_DIR
) -> str:
    """
    Compiles monthly summaries into a clean GSTR / VAT pre-formatted Excel book.
    Saves to /pump_exports/tax_filings/Tax_Filing_Template_[Month].xlsx.
    """
    os.makedirs(export_dir, exist_ok=True)
    summary = compile_monthly_tax_summary(year, month, db_path=db_path)
    
    # ── 1. Fuel VAT Table compilations ──
    fuel_list = summary["fuel_records"]
    df_fuel = pd.DataFrame(fuel_list)
    
    # Append totals row (unrounded decimals) if data exists
    if not df_fuel.empty:
        totals_row_fuel = {
            "Date": "TOTALS",
            "Diesel (HSD) Liters": summary["total_hsd_liters"],
            "Diesel (HSD) Rate (INR)": None,
            "Diesel (HSD) Turnover (INR)": summary["total_hsd_turnover"],
            "Diesel (HSD) VAT (INR)": summary["total_hsd_vat"],
            "Petrol (MS) Liters": summary["total_ms_liters"],
            "Petrol (MS) Rate (INR)": None,
            "Petrol (MS) Turnover (INR)": summary["total_ms_turnover"],
            "Petrol (MS) VAT (INR)": summary["total_ms_vat"],
            "Total VAT Obligation (INR)": summary["total_fuel_vat_obligation"]
        }
        df_fuel = pd.concat([df_fuel, pd.DataFrame([totals_row_fuel])], ignore_index=True)
    else:
        # Construct empty DataFrame structure
        df_fuel = pd.DataFrame(columns=[
            "Date", "Diesel (HSD) Liters", "Diesel (HSD) Rate (INR)", "Diesel (HSD) Turnover (INR)",
            "Diesel (HSD) VAT (INR)", "Petrol (MS) Liters", "Petrol (MS) Rate (INR)", 
            "Petrol (MS) Turnover (INR)", "Petrol (MS) VAT (INR)", "Total VAT Obligation (INR)"
        ])
        
    # ── 2. Lubricant GST Table compilations ──
    lube_list = summary["lube_records"]
    df_lube = pd.DataFrame(lube_list)
    
    if not df_lube.empty:
        totals_row_lube = {
            "Date": "TOTALS",
            "Lubricant Item Name": None,
            "Quantity Sold": summary["total_lube_qty"],
            "Unit Price (INR)": None,
            "Gross Revenue (INR)": summary["total_lube_gross"],
            "Base Taxable Value (INR)": summary["total_lube_base"],
            "CGST (9%) (INR)": summary["total_lube_cgst"],
            "SGST (9%) (INR)": summary["total_lube_sgst"],
            "Total GST (18%) (INR)": summary["total_lube_gst"]
        }
        df_lube = pd.concat([df_lube, pd.DataFrame([totals_row_lube])], ignore_index=True)
    else:
        df_lube = pd.DataFrame(columns=[
            "Date", "Lubricant Item Name", "Quantity Sold", "Unit Price (INR)",
            "Gross Revenue (INR)", "Base Taxable Value (INR)", "CGST (9%) (INR)",
            "SGST (9%) (INR)", "Total GST (18%) (INR)"
        ])
        
    # Filename configuration
    filename = f"Tax_Filing_Template_{month:02d}.xlsx"
    file_path = os.path.join(export_dir, filename)
    
    # Write Excel with openpyxl engine
    writer = pd.ExcelWriter(file_path, engine='openpyxl')
    df_fuel.to_excel(writer, sheet_name="Fuel VAT Returns", index=False)
    df_lube.to_excel(writer, sheet_name="Lubricant GST GSTR-1 Data Breakdown", index=False)
    
    # ── openpyxl Visual Styling Engine ──
    workbook = writer.book
    
    # Visual Styles
    navy_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    white_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Segoe UI", size=10, bold=True)
    regular_font = Font(name="Segoe UI", size=10)
    
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    double_bottom_border = Border(
        top=Side(style='thin', color='1F497D'),
        bottom=Side(style='double', color='1F497D')
    )
    
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        
        # 1. Format headers
        for col_idx in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = navy_fill
            cell.font = white_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
        # 2. Format data cells
        for row_idx in range(2, worksheet.max_row + 1):
            is_totals_row = (worksheet.cell(row=row_idx, column=1).value == "TOTALS")
            
            for col_idx in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                
                # Alignments and values format
                if col_idx == 1:
                    cell.alignment = Alignment(horizontal="center")
                elif isinstance(cell.value, (int, float)):
                    cell.alignment = Alignment(horizontal="right")
                    # unrounded decimal output format to keep full float values audit-stable
                    cell.number_format = "#,##0.0000"
                else:
                    cell.alignment = Alignment(horizontal="left")
                    
                # Fonts and borders
                if is_totals_row:
                    cell.font = bold_font
                    cell.border = double_bottom_border
                else:
                    cell.font = regular_font
                    cell.border = thin_border
                    
        # 3. Autofit columns
        for col in worksheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or "")
                if len(val_str) > max_len:
                    max_len = len(val_str)
            worksheet.column_dimensions[col_letter].width = max(max_len + 3, 14)
            
        # Freeze top row
        worksheet.freeze_panes = "A2"
        
    writer.close()
    logger.info(f"Successfully generated styled statutory Excel template: {file_path}")
    return file_path
