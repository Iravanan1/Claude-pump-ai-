"""
Unit tests for the repair_kit.py database recovery module.
"""

import os
import sys
import json
import sqlite3
import shutil
import tempfile
import unittest
from datetime import datetime
from unittest.mock import patch, MagicMock

import pandas as pd
from openpyxl import load_workbook, Workbook

# Add backend directory to system path
BACKEND_DIR = os.path.dirname(os.path.abspath(__file__))
if BACKEND_DIR not in sys.path:
    sys.path.append(BACKEND_DIR)

import repair_kit
from crypto_vault import encrypt_field

class TestRepairKit(unittest.TestCase):
    def setUp(self):
        # Create unique temporary workspace directories
        self.test_dir = tempfile.mkdtemp()
        self.db_path = os.path.join(self.test_dir, "test_ledger.db")
        self.exports_dir = os.path.join(self.test_dir, "pump_exports")
        os.makedirs(self.exports_dir, exist_ok=True)
        self.output_path = os.path.join(self.exports_dir, "Pump_Accounts.xlsx")

        # Initialize mock database
        self.conn = sqlite3.connect(self.db_path)
        self.cursor = self.conn.cursor()
        
        # Create required tables
        self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS daily_summary (
                date TEXT PRIMARY KEY,
                total_hsd_liters REAL DEFAULT 0.0,
                total_ms_liters REAL DEFAULT 0.0,
                total_cash_calculated REAL DEFAULT 0.0,
                total_credit_sales REAL DEFAULT 0.0,
                total_testing_deductions REAL DEFAULT 0.0,
                is_verified INTEGER DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        
        self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS ledger_entries (
                entry_id INTEGER PRIMARY KEY AUTOINCREMENT,
                date TEXT,
                party_name TEXT,
                vehicle_wheel_no TEXT,
                amount REAL DEFAULT 0.0,
                type TEXT,
                remarks TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        
        self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS daily_ledger (
                date TEXT PRIMARY KEY,
                raw_data TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        
        self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS dsm_shifts (
                date TEXT,
                shift_type TEXT,
                dsm_name TEXT,
                assigned_nozzles TEXT,
                cash_handed_over REAL,
                digital_slips_value REAL,
                calculated_shortage_or_excess REAL
            )
        """)
        self.conn.commit()

    def tearDown(self):
        self.conn.close()
        try:
            shutil.rmtree(self.test_dir)
        except Exception:
            pass

    def test_file_isolation_missing(self):
        # If output file is missing, isolation should return False and not raise error
        isolated = repair_kit.isolate_master_spreadsheet(self.output_path)
        self.assertFalse(isolated)
        self.assertFalse(os.path.exists(self.output_path))

    def test_file_isolation_healthy(self):
        # Create a healthy placeholder spreadsheet
        wb = Workbook()
        wb.save(self.output_path)
        self.assertTrue(os.path.exists(self.output_path))
        
        # Trigger isolation (healthy gets isolated as backup by default)
        isolated = repair_kit.isolate_master_spreadsheet(self.output_path)
        self.assertTrue(isolated)
        self.assertFalse(os.path.exists(self.output_path))
        
        # Verify backup exists inside temporary exports folder
        files = os.listdir(self.exports_dir)
        self.assertTrue(any(f.startswith("Pump_Accounts_backup_") and f.endswith(".xlsx") for f in files))

    def test_file_isolation_corrupted(self):
        # Create a corrupted/unparseable file
        with open(self.output_path, "w") as f:
            f.write("corrupted data not xlsx format")
            
        # Trigger isolation
        isolated = repair_kit.isolate_master_spreadsheet(self.output_path)
        self.assertTrue(isolated)
        self.assertFalse(os.path.exists(self.output_path))
        
        # Verify isolated corrupted backup exists
        files = os.listdir(self.exports_dir)
        self.assertTrue(any(f.startswith("Pump_Accounts_corrupted_") and f.endswith(".xlsx") for f in files))

    @patch("repair_kit.open")
    def test_file_isolation_open(self, mock_open):
        # Simulate open/locked state by raising PermissionError/IOError on file checks
        wb = Workbook()
        wb.save(self.output_path)
        
        # Mock file open to throw permission error simulating locked file
        mock_open.side_effect = PermissionError("File is open in Microsoft Excel")
        
        # Trigger isolation
        isolated = repair_kit.isolate_master_spreadsheet(self.output_path)
        self.assertTrue(isolated)
        self.assertFalse(os.path.exists(self.output_path))
        
        # Verify isolated open backup exists
        files = os.listdir(self.exports_dir)
        self.assertTrue(any(f.startswith("Pump_Accounts_open_") and f.endswith(".xlsx") for f in files))

    def test_rebuild_chronological_sorting_and_verified_filtering(self):
        # 1. Populate mock database data (verified and unverified entries, out of order)
        # 2026-06-10 (Verified)
        self.cursor.execute("""
            INSERT INTO daily_summary (date, total_hsd_liters, total_ms_liters, total_cash_calculated, total_credit_sales, total_testing_deductions, is_verified)
            VALUES ('2026-06-10', 400.0, 200.0, 42000.0, 10000.0, 5.0, 1)
        """)
        # 2026-06-05 (Verified - Earlier date added second to check sorting!)
        self.cursor.execute("""
            INSERT INTO daily_summary (date, total_hsd_liters, total_ms_liters, total_cash_calculated, total_credit_sales, total_testing_deductions, is_verified)
            VALUES ('2026-06-05', 300.0, 150.0, 31500.0, 5000.0, 5.0, 1)
        """)
        # 2026-06-12 (Unverified - Should be skipped in main scan!)
        self.cursor.execute("""
            INSERT INTO daily_summary (date, total_hsd_liters, total_ms_liters, total_cash_calculated, total_credit_sales, total_testing_deductions, is_verified)
            VALUES ('2026-06-12', 500.0, 300.0, 55000.0, 20000.0, 5.0, 0)
        """)
        
        # Populating ledger entries (with encrypted fields)
        enc_party1 = encrypt_field("Rahul Transports")
        enc_amt1 = encrypt_field(5000.0)
        self.cursor.execute("""
            INSERT INTO ledger_entries (date, party_name, vehicle_wheel_no, amount, type, remarks)
            VALUES ('2026-06-05', ?, 'HR-55-A-1111', ?, 'udhaar', 'HSD 50L')
        """, (enc_party1, enc_amt1))
        
        enc_party2 = encrypt_field("Office Tea Vendor")
        enc_amt2 = encrypt_field(150.0)
        self.cursor.execute("""
            INSERT INTO ledger_entries (date, party_name, vehicle_wheel_no, amount, type, remarks)
            VALUES ('2026-06-10', ?, 'N/A', ?, 'expense', 'Evening snacks')
        """, (enc_party2, enc_amt2))
        
        # This belongs to unverified date 2026-06-12 - should be excluded
        enc_party3 = encrypt_field("Unverified Customer")
        enc_amt3 = encrypt_field(20000.0)
        self.cursor.execute("""
            INSERT INTO ledger_entries (date, party_name, vehicle_wheel_no, amount, type, remarks)
            VALUES ('2026-06-12', ?, 'DL-1C-9999', ?, 'udhaar', 'Bulk MS')
        """, (enc_party3, enc_amt3))
        
        # DSM Shift allocations for responsible staff
        self.cursor.execute("""
            INSERT INTO dsm_shifts (date, shift_type, dsm_name, assigned_nozzles, cash_handed_over, digital_slips_value, calculated_shortage_or_excess)
            VALUES ('2026-06-05', 'Day', 'Suresh', 'HSD-1', 15000.0, 0.0, 0.0)
        """)
        
        # Daily ledger nozzle data
        raw_ledger_6_05 = {
            "date": "2026-06-05",
            "nozzles": [
                {"nozzle_name": "HSD-1", "sales_liters_calculated": 300.0}
            ]
        }
        self.cursor.execute("""
            INSERT INTO daily_ledger (date, raw_data)
            VALUES ('2026-06-05', ?)
        """, (json.dumps(raw_ledger_6_05),))
        
        self.conn.commit()
        
        # Run reconstruction utility
        repair_kit.rebuild_master_spreadsheet(db_path=self.db_path, output_path=self.output_path)
        
        # 2. Verify Output File Generated
        self.assertTrue(os.path.exists(self.output_path))
        
        # 3. Read compiled sheets
        wb = load_workbook(self.output_path)
        self.assertIn("Shift Readings", wb.sheetnames)
        self.assertIn("Credit Ledger", wb.sheetnames)
        self.assertIn("Expenses", wb.sheetnames)
        
        # Assert Shift Readings chronological sorting
        ws_shift = wb["Shift Readings"]
        self.assertEqual(ws_shift.cell(row=2, column=1).value, "2026-06-05")  # Earliest first!
        self.assertEqual(ws_shift.cell(row=3, column=1).value, "2026-06-10")  # Latest second!
        self.assertEqual(ws_shift.max_row, 3) # Excluded unverified 2026-06-12!
        
        # Nozzle flow cell check
        self.assertIn("HSD-1: 300.0 L", str(ws_shift.cell(row=2, column=2).value))
        
        # Active staff check
        self.assertEqual(ws_shift.cell(row=2, column=5).value, "HSD-1: Suresh")
        
        # Assert Credit Ledger decryption and sorting
        ws_credit = wb["Credit Ledger"]
        self.assertEqual(ws_credit.max_row, 2) # Header + 1 credit record (from 2026-06-05)
        self.assertEqual(ws_credit.cell(row=2, column=1).value, "2026-06-05")
        self.assertEqual(ws_credit.cell(row=2, column=2).value, "Rahul Transports")  # Decrypted!
        self.assertEqual(float(ws_credit.cell(row=2, column=4).value), 5000.0)      # Decrypted!
        
        # Assert Expenses decryption and sorting
        ws_exp = wb["Expenses"]
        self.assertEqual(ws_exp.max_row, 2) # Header + 1 expense record (from 2026-06-10)
        self.assertEqual(ws_exp.cell(row=2, column=1).value, "2026-06-10")
        self.assertEqual(ws_exp.cell(row=2, column=2).value, "Office Tea Vendor")  # Decrypted!
        self.assertEqual(float(ws_exp.cell(row=2, column=3).value), 150.0)         # Decrypted!

        # Assert visual style formatting applied (Navy header text font size, borders)
        self.assertIn("FFFFFF", str(ws_shift.cell(row=1, column=1).font.color.value))  # White text headers
        self.assertEqual(ws_shift.cell(row=1, column=1).fill.start_color.value, "001F497D") # Navy blue fill
        self.assertTrue(ws_shift.views.sheetView[0].showGridLines)  # Show Gridlines forced!

if __name__ == "__main__":
    unittest.main()
