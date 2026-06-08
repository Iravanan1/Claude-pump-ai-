"""
Database Recovery & Spreadsheet Reconstruction Utility.
Performs a complete system refresh by checking, isolating, and rebuilding
the master accounts spreadsheet 'Pump_Accounts.xlsx' chronologically from SQLite.
"""

import os
import re
import json
import sqlite3
import shutil
import logging
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Setup logger
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("RepairKit")

from pathlib import Path
from dotenv import load_dotenv

# Explicitly find root directory .env
_root_dir = Path(__file__).resolve().parent.parent
_root_env = _root_dir / ".env"
if _root_env.exists():
    load_dotenv(dotenv_path=_root_env)
else:
    load_dotenv() # Fallback to local

BACKEND_DIR = os.path.dirname(os.path.abspath(__file__))
WORKSPACE_DIR = os.path.dirname(BACKEND_DIR)
EXPORTS_DIR = os.path.join(WORKSPACE_DIR, "pump_exports")

_export_path = os.getenv("EXPORT_EXCEL_PATH")
if not _export_path:
    OUTPUT_PATH = os.path.join(EXPORTS_DIR, "Pump_Accounts.xlsx")
else:
    if not os.path.isabs(_export_path):
        OUTPUT_PATH = os.path.abspath(os.path.join(WORKSPACE_DIR, _export_path))
    else:
        OUTPUT_PATH = _export_path

DB_PATH = os.path.abspath(os.path.join(BACKEND_DIR, "ledger.db"))

from crypto_vault import decrypt_field, decrypt_raw_data

def isolate_master_spreadsheet(filepath: str = OUTPUT_PATH) -> bool:
    """
    Check if the master spreadsheet exists. If it is open, missing, or corrupted,
    isolate it safely by renaming it to avoid overwrite conflicts and preserve data.
    Returns True if an isolation was performed.
    """
    if not os.path.exists(filepath):
        logger.info("Master spreadsheet is missing. No isolation needed.")
        return False

    is_corrupted = False
    try:
        # Check if file can be opened and parsed as a workbook
        wb = load_workbook(filepath, read_only=True)
        wb.close()
    except Exception as e:
        logger.warning(f"Existing spreadsheet appears to be corrupted: {str(e)}")
        is_corrupted = True

    is_open = False
    if not is_corrupted:
        try:
            # Check if file is open/locked (try opening in write exclusive mode)
            with open(filepath, "r+"):
                pass
        except (IOError, OSError) as e:
            logger.warning(f"Existing spreadsheet appears to be open or locked: {str(e)}")
            is_open = True

    # Isolate safely
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    reason = "corrupted" if is_corrupted else ("open" if is_open else "backup")
    isolated_dir = os.path.dirname(filepath) or "."
    isolated_path = os.path.join(isolated_dir, f"Pump_Accounts_{reason}_{timestamp}.xlsx")

    try:
        shutil.move(filepath, isolated_path)
        logger.info(f"Isolated existing {reason} spreadsheet to: {isolated_path}")
        return True
    except Exception as move_err:
        logger.error(f"Failed to isolate master spreadsheet: {str(move_err)}")
        # If move fails (e.g. strict OS file locks on Windows), try copy and overwrite
        try:
            shutil.copy2(filepath, isolated_path)
            os.remove(filepath)
            logger.info(f"Copied and removed master spreadsheet to isolate: {isolated_path}")
            return True
        except Exception as fallback_err:
            logger.critical(f"Critical failure isolating master spreadsheet: {str(fallback_err)}")
            raise fallback_err

def apply_premium_styles(excel_path: str):
    """
    Applies professional styling, gridlines, auto-fits, and number formatting to the Excel sheets.
    """
    logger.info(f"Applying premium openpyxl styles to {excel_path}...")
    try:
        wb = load_workbook(excel_path)
        
        # Premium design tokens (matching standard exporter styling)
        header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Navy Blue
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        regular_font = Font(name="Segoe UI", size=11)
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Force gridlines visibility
            ws.views.sheetView[0].showGridLines = True
            
            # 1. Headers Row formatting
            ws.row_dimensions[1].height = 28
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border
                
            # 2. Data Rows formatting
            for row in range(2, ws.max_row + 1):
                ws.row_dimensions[row].height = 20
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.font = regular_font
                    cell.border = thin_border
                    
                    col_name = str(ws.cell(row=1, column=col).value)
                    
                    # Formatting values
                    if isinstance(cell.value, (int, float)):
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                        
                        # INR formatting for cash and amounts
                        if any(term in col_name for term in ["cash", "amount", "credit", "sales", "INR", "Amount"]):
                            if "liter" not in col_name.lower():
                                cell.number_format = "[$₹-4009] #,##0.00"
                            else:
                                cell.number_format = "#,##0.00"
                        else:
                            cell.number_format = "#,##0.00"
                    else:
                        # Align center for dates
                        if "Date" in col_name:
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        else:
                            cell.alignment = Alignment(horizontal="left", vertical="center")
            
            # 3. Dynamic Column widths
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    val_str = str(cell.value or '')
                    if len(val_str) > max_len:
                        max_len = len(val_str)
                ws.column_dimensions[col_letter].width = max(max_len + 4, 14)
                
        wb.save(excel_path)
        logger.info("Premium openpyxl styling applied successfully!")
        
    except Exception as e:
        logger.error(f"Failed to apply Excel styling to reconstructed sheet: {str(e)}")
        raise e

def rebuild_master_spreadsheet(db_path: str = DB_PATH, output_path: str = OUTPUT_PATH) -> str:
    """
    Scans the SQLite database, extracts verified operational summaries and ledger entries
    across the entire history, sorts them chronologically, and compiles a fresh Excel workbook.
    """
    logger.info("Initializing complete spreadsheet reconstruction and database recovery...")
    
    # 1. Create target directory
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    
    # 2. Isolate existing spreadsheet
    isolate_master_spreadsheet(output_path)
    
    if not os.path.exists(db_path):
        logger.error(f"Database file not found at {db_path}")
        raise FileNotFoundError(f"Database file not found at {db_path}")
        
    try:
        conn = sqlite3.connect(db_path)
        
        # 3. Complete Table Scan: Query verified daily_summaries (with is_verified = 1)
        summary_query = "SELECT * FROM daily_summary WHERE is_verified = 1 ORDER BY date ASC"
        df_summary = pd.read_sql_query(summary_query, conn)
        
        if df_summary.empty:
            logger.warning("No verified daily summaries found (is_verified = 1). Performing fallback to extract all summaries.")
            summary_query = "SELECT * FROM daily_summary ORDER BY date ASC"
            df_summary = pd.read_sql_query(summary_query, conn)
            
        # Ensure summaries are sorted chronologically
        if not df_summary.empty:
            df_summary["date_parsed"] = pd.to_datetime(df_summary["date"], errors="coerce")
            df_summary = df_summary.sort_values(by="date_parsed", ascending=True)
            df_summary = df_summary.drop(columns=["date_parsed"])
            
        # 4. Extract verified ledger entries corresponding to verified dates
        verified_dates = list(df_summary["date"].unique())
        
        if verified_dates:
            # Construct parameterized query for safety
            placeholders = ",".join("?" for _ in verified_dates)
            entries_query = f"SELECT * FROM ledger_entries WHERE date IN ({placeholders}) ORDER BY date ASC, entry_id ASC"
            df_entries = pd.read_sql_query(entries_query, conn, params=verified_dates)
        else:
            entries_query = "SELECT * FROM ledger_entries ORDER BY date ASC, entry_id ASC"
            df_entries = pd.read_sql_query(entries_query, conn)
            
        # Decrypt ledger fields transparently
        if not df_entries.empty:
            df_entries["party_name"] = df_entries["party_name"].apply(lambda v: decrypt_field(v, return_type=str))
            df_entries["amount"] = df_entries["amount"].apply(lambda v: decrypt_field(v, return_type=float))
            
        # Fetch raw ledger data for nozzle flow parsing
        df_ledger = pd.read_sql_query("SELECT date, raw_data FROM daily_ledger", conn)
        
        # Fetch DSM nozzle staffing mapping
        dsm_mapping = {}
        try:
            cursor_dsm = conn.cursor()
            cursor_dsm.execute("SELECT date, dsm_name, assigned_nozzles FROM dsm_shifts")
            dsm_rows = cursor_dsm.fetchall()
            for d_str_dsm, name, nozzles_str in dsm_rows:
                if nozzles_str:
                    if d_str_dsm not in dsm_mapping:
                        dsm_mapping[d_str_dsm] = {}
                    nozzles_list = [n.strip() for n in nozzles_str.split(",") if n.strip()]
                    for n in nozzles_list:
                        n_clean = n.lower()
                        if " (" in n_clean:
                            n_clean = n_clean.split(" (")[0]
                        dsm_mapping[d_str_dsm][n_clean] = name
        except Exception as dsm_err:
            logger.warning(f"Failed to load DSM mapping: {str(dsm_err)}")
            
        conn.close()
        
        # =====================================================================
        # Tab 1: "Shift Readings" compilation
        # =====================================================================
        shift_readings = []
        for idx, row in df_summary.iterrows():
            d_str = row["date"]
            total_liters = float(row["total_hsd_liters"] or 0.0) + float(row["total_ms_liters"] or 0.0)
            fuel_cash = float(row["total_cash_calculated"] or 0.0)
            
            # Format Nozzle Flows
            flows = "N/A"
            matching_ledger = df_ledger[df_ledger["date"] == d_str]
            if not matching_ledger.empty:
                try:
                    decrypted_raw = decrypt_raw_data(json.loads(matching_ledger.iloc[0]["raw_data"]))
                    flow_items = []
                    for n in decrypted_raw.get("nozzles", []):
                        flow_items.append(f"{n.get('nozzle_name')}: {n.get('sales_liters_calculated') or n.get('net_sales_liters') or 0.0} L")
                    flows = " | ".join(flow_items)
                except Exception:
                    pass
            
            if flows == "N/A":
                flows = f"HSD: {row['total_hsd_liters']} L, MS: {row['total_ms_liters']} L"
                
            # DSM Nozzle staff responsible
            staff_list = []
            date_mapping = dsm_mapping.get(d_str, {})
            if date_mapping:
                staff_list = [f"{n.upper()}: {staff}" for n, staff in sorted(date_mapping.items())]
            staff_str = " | ".join(staff_list) if staff_list else "N/A"
            
            shift_readings.append({
                "Date": d_str,
                "Nozzle Flows": flows,
                "Total Liters Sold": total_liters,
                "Calculated Fuel Cash": fuel_cash,
                "Active Nozzle Staff": staff_str
            })
            
        df_shift_nice = pd.DataFrame(shift_readings) if shift_readings else pd.DataFrame(
            columns=["Date", "Nozzle Flows", "Total Liters Sold", "Calculated Fuel Cash", "Active Nozzle Staff"]
        )
        
        # Ensure chronological sorting
        if not df_shift_nice.empty:
            df_shift_nice["date_parsed"] = pd.to_datetime(df_shift_nice["Date"], errors="coerce")
            df_shift_nice = df_shift_nice.sort_values(by="date_parsed", ascending=True)
            df_shift_nice = df_shift_nice.drop(columns=["date_parsed"])

        # =====================================================================
        # Tab 2: "Credit Ledger" compilation
        # =====================================================================
        df_credit_raw = df_entries[df_entries["type"] == "udhaar"].copy()
        if not df_credit_raw.empty:
            df_credit_nice = pd.DataFrame({
                "Date": df_credit_raw["date"],
                "Customer / Party Name": df_credit_raw["party_name"],
                "Vehicle Wheel No": df_credit_raw["vehicle_wheel_no"].fillna("N/A"),
                "Amount (INR)": df_credit_raw["amount"],
                "Remarks / Details": df_credit_raw["remarks"].fillna("")
            })
            # Ensure chronological sorting
            df_credit_nice["date_parsed"] = pd.to_datetime(df_credit_nice["Date"], errors="coerce")
            df_credit_nice = df_credit_nice.sort_values(by="date_parsed", ascending=True)
            df_credit_nice = df_credit_nice.drop(columns=["date_parsed"])
        else:
            df_credit_nice = pd.DataFrame(columns=["Date", "Customer / Party Name", "Vehicle Wheel No", "Amount (INR)", "Remarks / Details"])

        # =====================================================================
        # Tab 3: "Expenses" compilation
        # =====================================================================
        df_expenses_raw = df_entries[df_entries["type"] == "expense"].copy()
        if not df_expenses_raw.empty:
            df_expenses_nice = pd.DataFrame({
                "Date": df_expenses_raw["date"],
                "Party / Payee": df_expenses_raw["party_name"],
                "Amount (INR)": df_expenses_raw["amount"],
                "Remarks / Details": df_expenses_raw["remarks"].fillna("")
            })
            # Ensure chronological sorting
            df_expenses_nice["date_parsed"] = pd.to_datetime(df_expenses_nice["Date"], errors="coerce")
            df_expenses_nice = df_expenses_nice.sort_values(by="date_parsed", ascending=True)
            df_expenses_nice = df_expenses_nice.drop(columns=["date_parsed"])
        else:
            df_expenses_nice = pd.DataFrame(columns=["Date", "Party / Payee", "Amount (INR)", "Remarks / Details"])

        # 5. Write to distinct sheet tabs
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df_shift_nice.to_excel(writer, sheet_name="Shift Readings", index=False)
            df_credit_nice.to_excel(writer, sheet_name="Credit Ledger", index=False)
            df_expenses_nice.to_excel(writer, sheet_name="Expenses", index=False)
            
        # 6. Apply professional visual styles
        apply_premium_styles(output_path)
        
        logger.info(f"Pristine master spreadsheet successfully reconstructed at {output_path}!")
        return output_path
        
    except Exception as e:
        logger.error(f"Spreadsheet reconstruction aborted: {str(e)}")
        raise e

if __name__ == "__main__":
    logger.info("Executing Standalone Spreadsheet Repair Kit Rebuild...")
    rebuild_master_spreadsheet()
