import sqlite3
import pandas as pd
import os
import logging
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Setup logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("Exporter")

from pathlib import Path
from dotenv import load_dotenv

# Explicitly find root directory .env
_root_dir = Path(__file__).resolve().parent.parent
_root_env = _root_dir / ".env"
if _root_env.exists():
    load_dotenv(dotenv_path=_root_env)
else:
    load_dotenv() # Fallback to local

# Dynamically resolve default master Excel path from environment
_export_path = os.getenv("EXPORT_EXCEL_PATH")
if not _export_path:
    _backend_dir = os.path.dirname(os.path.abspath(__file__))
    _workspace_dir = os.path.dirname(_backend_dir)
    DEFAULT_EXCEL_PATH = os.path.join(_workspace_dir, "pump_exports", "Pump_Accounts.xlsx")
else:
    if not os.path.isabs(_export_path):
        _backend_dir = os.path.dirname(os.path.abspath(__file__))
        _workspace_dir = os.path.dirname(_backend_dir)
        DEFAULT_EXCEL_PATH = os.path.abspath(os.path.join(_workspace_dir, _export_path))
    else:
        DEFAULT_EXCEL_PATH = _export_path

BACKEND_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BACKEND_DIR, "ledger.db")

from crypto_vault import decrypt_field, decrypt_raw_data

def apply_excel_styling(excel_path: str):
    """
    Applies professional styling, gridlines, auto-fits, and number formatting to the Excel sheets.
    """
    logger.info(f"Applying premium openpyxl styles to {excel_path}...")
    try:
        wb = load_workbook(excel_path)
        
        # Style tokens
        header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Navy Blue
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        regular_font = Font(name="Segoe UI", size=11)
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Explicitly force Excel to display gridlines
            ws.views.sheetView[0].showGridLines = True
            
            # Identify secondary headers row if "Internal Consumption" sheet
            second_header_row = None
            section_title_row = None
            if sheet_name == "Internal Consumption":
                for r in range(2, ws.max_row + 1):
                    val = ws.cell(row=r, column=1).value
                    if val == "Entry ID":
                        second_header_row = r
                        if r > 1:
                            section_title_row = r - 1
                        break
            
            # 1. Format Headers row
            ws.row_dimensions[1].height = 28
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border
                
            # Style second header if present
            if second_header_row:
                ws.row_dimensions[second_header_row].height = 28
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=second_header_row, column=col)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = thin_border
            
            # Style section title row if present
            if section_title_row:
                ws.row_dimensions[section_title_row].height = 24
                # Merge across all columns for a beautiful banner
                ws.merge_cells(start_row=section_title_row, start_column=1, end_row=section_title_row, end_column=ws.max_column)
                title_cell = ws.cell(row=section_title_row, column=1)
                title_cell.font = Font(name="Segoe UI", size=12, bold=True, color="1F497D")
                title_cell.alignment = Alignment(horizontal="left", vertical="center")
                # Clear border for title cell to keep it clean
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=section_title_row, column=col).border = Border()
                
            # 2. Format Data rows
            for row in range(2, ws.max_row + 1):
                # If row is empty, clear borders and skip
                is_empty_row = all(ws.cell(row=row, column=c).value in (None, "") for c in range(1, ws.max_column + 1))
                if is_empty_row:
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=row, column=col).border = Border()
                    continue
                
                if second_header_row and row == second_header_row:
                    continue
                    
                if section_title_row and row == section_title_row:
                    continue
                    
                ws.row_dimensions[row].height = 20
                is_totals_row = (str(ws.cell(row=row, column=1).value) == "Profit Accounting Totals")
                
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    
                    if is_totals_row:
                        cell.font = Font(name="Segoe UI", size=11, bold=True)
                        cell.border = Border(
                            top=Side(style='thin', color='A0A0A0'),
                            bottom=Side(style='double', color='1F497D'),
                            left=Side(style='thin', color='D9D9D9'),
                            right=Side(style='thin', color='D9D9D9')
                        )
                    else:
                        cell.font = regular_font
                        cell.border = thin_border
                    
                    # For custom tables, resolve header row name
                    # If row > second_header_row, look at second_header_row for col_name
                    h_row = second_header_row if (second_header_row and row > second_header_row) else 1
                    col_name = str(ws.cell(row=h_row, column=col).value)
                    
                    # Formatting logic based on column contents
                    if isinstance(cell.value, (int, float)):
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                        
                        # INR Currency columns (Rupees)
                        if any(term in col_name for term in ["cash", "amount", "credit", "sales", "Tender", "INR", "calculated", "Profit", "Spread"]) and not any(term in col_name for term in ["USD", "usd"]):
                            if "liter" not in col_name.lower():
                                cell.number_format = "[$₹-4009] #,##0.00"
                            else:
                                cell.number_format = "#,##0.00"
                        # USD Currency columns (Dollars)
                        elif any(term in col_name for term in ["USD", "usd"]):
                            cell.number_format = "$#,##0.00"
                        # Liters or generic float columns
                        else:
                            cell.number_format = "#,##0.00"
                    else:
                        # Center status, booleans, and dates
                        if any(term in col_name for term in ["Date", "verified", "status", "type", "Month", "Product Type"]):
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        else:
                            cell.alignment = Alignment(horizontal="left", vertical="center")
                        
                        # Inventory Reconciliation status column — traffic-light colouring
                        if sheet_name == "Inventory Reconciliation" and col_name == "Status":
                            if cell.value == "Abnormal Product Leakage Alert":
                                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                                cell.font = Font(name="Segoe UI", size=11, bold=True, color="9C0006")
                            elif cell.value == "Normal Handling Shrinkage":
                                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                                cell.font = Font(name="Segoe UI", size=11, color="276221")
            
            # 3. Dynamic Column Auto-fitting
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    val_str = str(cell.value or '')
                    # Avoid long merged title strings warping column widths
                    if cell.coordinate in ws.merged_cells:
                        continue
                    if len(val_str) > max_len:
                        max_len = len(val_str)
                ws.column_dimensions[col_letter].width = max(max_len + 4, 14)
                
        wb.save(excel_path)
        logger.info("Excel premium styling applied triumphantly!")
        
    except Exception as e:
        logger.error(f"Failed to apply Excel styling: {str(e)}")
        raise e

def export_db_to_excel(excel_path: str = DEFAULT_EXCEL_PATH) -> str:
    """
    Reads the sqlite3 tables daily_summary and ledger_entries, writes them 
    into an Excel workbook, and styles them.
    """
    logger.info(f"Exporting database to Excel file: {excel_path}...")
    
    if not os.path.exists(DB_PATH):
        logger.error(f"Database file '{DB_PATH}' does not exist!")
        raise FileNotFoundError(f"Database ledger.db not found.")
        
    try:
        # Connect to sqlite
        conn = sqlite3.connect(DB_PATH)
        
        # Load tables into Pandas DataFrames
        df_summary = pd.read_sql_query("SELECT * FROM daily_summary ORDER BY date DESC", conn)
        df_entries = pd.read_sql_query("SELECT * FROM ledger_entries ORDER BY date DESC, entry_id DESC", conn)
        
        # Ensure internal_consumption table exists
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='internal_consumption'")
        if not cursor.fetchone():
            cursor.execute("""
            CREATE TABLE IF NOT EXISTS internal_consumption (
                entry_id INTEGER PRIMARY KEY AUTOINCREMENT,
                date TEXT,
                product_type TEXT CHECK(product_type IN ('HSD', 'MS')),
                liters_drawn REAL DEFAULT 0.0,
                purpose_head TEXT,
                authorized_by TEXT
            )
            """)
            conn.commit()
            
        # Load internal consumption logs and monthly cumulative summary
        df_detailed = pd.read_sql_query("SELECT * FROM internal_consumption ORDER BY date DESC, entry_id DESC", conn)
        df_monthly = pd.read_sql_query("""
            SELECT 
                strftime('%Y-%m', date) AS month,
                product_type,
                SUM(liters_drawn) AS cumulative_liters_drawn,
                COUNT(*) AS total_transactions
            FROM internal_consumption
            GROUP BY month, product_type
            ORDER BY month DESC, product_type ASC
        """, conn)

        # Check and load tanker receipts if table exists
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='tanker_receipts'")
        has_receipts = cursor.fetchone() is not None
        if has_receipts:
            df_receipts = pd.read_sql_query("SELECT * FROM tanker_receipts ORDER BY date DESC, invoice_no DESC", conn)
        else:
            df_receipts = pd.DataFrame()
            
        # Check and load lube stock ledger if table exists
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='lube_inventory_ledger'")
        has_lube_ledger = cursor.fetchone() is not None
        if has_lube_ledger:
            df_lube_ledger = pd.read_sql_query("SELECT * FROM lube_inventory_ledger ORDER BY item_sku ASC", conn)
        else:
            df_lube_ledger = pd.DataFrame()
        
        conn.close()
        
        # Seamlessly decrypt sensitive fields for unencrypted Excel generation
        if not df_entries.empty:
            df_entries["party_name"] = df_entries["party_name"].apply(lambda val: decrypt_field(val, return_type=str))
            df_entries["amount"] = df_entries["amount"].apply(lambda val: decrypt_field(val, return_type=float))
        
        # Map nice column headers for Daily Summary
        rename_map = {
            "date": "Date",
            "total_hsd_liters": "HSD Sold (Liters)",
            "total_ms_liters": "MS Sold (Liters)",
            "total_cash_calculated": "Cash Calculated (INR)",
            "total_credit_sales": "Credit Sales (INR)",
            "total_testing_deductions": "Testing Deductions (Liters)",
            "is_verified": "Is Verified (1=Yes)",
            "created_at": "Logged Timestamp"
        }
        if "total_regular_hsd_liters" in df_summary.columns:
            rename_map["total_regular_hsd_liters"] = "Regular HSD Sold (L)"
        if "total_premium_hsd_liters" in df_summary.columns:
            rename_map["total_premium_hsd_liters"] = "Premium HSD Sold (L)"
        if "total_regular_ms_liters" in df_summary.columns:
            rename_map["total_regular_ms_liters"] = "Regular MS Sold (L)"
        if "total_premium_ms_liters" in df_summary.columns:
            rename_map["total_premium_ms_liters"] = "Premium MS Sold (L)"
        if "gross_spread_usd" in df_summary.columns:
            rename_map["gross_spread_usd"] = "Gross Spread (USD)"
        if "realized_profit" in df_summary.columns:
            rename_map["realized_profit"] = "Realized Profit"
            
        df_summary_nice = df_summary.rename(columns=rename_map)

        # Calculate totals for daily summary sheet
        if not df_summary_nice.empty:
            totals = {col: "" for col in df_summary_nice.columns}
            totals[df_summary_nice.columns[0]] = "Profit Accounting Totals"
            
            for col in df_summary_nice.columns:
                if col in [
                    "HSD Sold (Liters)", "MS Sold (Liters)",
                    "Regular HSD Sold (L)", "Premium HSD Sold (L)",
                    "Regular MS Sold (L)", "Premium MS Sold (L)",
                    "Cash Calculated (INR)", "Credit Sales (INR)",
                    "Gross Spread (USD)", "Realized Profit"
                ]:
                    totals[col] = df_summary_nice[col].sum()
            
            totals_df = pd.DataFrame([totals])
            df_summary_nice = pd.concat([df_summary_nice, totals_df], ignore_index=True)
        
        # Map nice column headers for Credit Ledger & Expenses
        df_entries_nice = df_entries.rename(columns={
            "entry_id": "Entry ID",
            "date": "Date",
            "party_name": "Party / Customer Name",
            "vehicle_wheel_no": "Vehicle Wheel No",
            "amount": "Amount (INR)",
            "type": "Entry Type",
            "remarks": "Remarks / Details",
            "created_at": "Logged Timestamp"
        })

        # Map nice column headers for Internal Consumption Monthly Cumulative
        df_monthly_nice = df_monthly.rename(columns={
            "month": "Month",
            "product_type": "Product Type",
            "cumulative_liters_drawn": "Cumulative Liters Drawn (Liters)",
            "total_transactions": "Total Transactions"
        })
        
        # Map nice column headers for Internal Consumption Detailed Log
        df_detailed_nice = df_detailed.rename(columns={
            "entry_id": "Entry ID",
            "date": "Date",
            "product_type": "Product Type",
            "liters_drawn": "Liters Drawn (Liters)",
            "purpose_head": "Purpose Head",
            "authorized_by": "Authorized By"
        })
        
        # Write to separate Excel sheets using Pandas ExcelWriter
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df_summary_nice.to_excel(writer, sheet_name="Daily Sales Summaries", index=False)
            df_entries_nice.to_excel(writer, sheet_name="Credit Ledger & Expenses", index=False)
            
            # Write first table
            df_monthly_nice.to_excel(writer, sheet_name="Internal Consumption", index=False, startrow=0, startcol=0)
            
            # Write second table
            start_row_detailed = len(df_monthly_nice) + 4
            df_detailed_nice.to_excel(writer, sheet_name="Internal Consumption", index=False, startrow=start_row_detailed, startcol=0)
            
            # Get sheet to write section banner
            ws = writer.sheets["Internal Consumption"]
            ws.cell(row=start_row_detailed, column=1, value="DETAILED INTERNAL FUEL DRAW LOGS")

            # Write Transporter Transit Claims Ledger if present
            if has_receipts and not df_receipts.empty:
                df_receipts_nice = df_receipts.rename(columns={
                    "invoice_no": "Invoice No",
                    "date": "Date",
                    "tank_lorry_no": "Tank Lorry No",
                    "product_type": "Product Type",
                    "invoice_volume_liters": "Invoice Volume (Liters)",
                    "invoice_density_at_15c": "Invoice Density @ 15°C (kg/m³)",
                    "observed_compartment_dips_mm": "Observed Compartment Dips (mm)",
                    "observed_density_raw": "Observed Density Raw (kg/m³)",
                    "observed_temperature_celsius": "Observed Temperature (°C)",
                    "actual_received_volume_liters": "Actual Received Volume (Liters)",
                    "transit_shortage_liters": "Transit Shortage (Liters)"
                })
                df_receipts_nice.to_excel(writer, sheet_name="Transporter Transit Claims Ledger", index=False)
                
            # Write Lubricant Stock Inventory Book if present
            if has_lube_ledger and not df_lube_ledger.empty:
                df_lube_ledger_nice = df_lube_ledger.rename(columns={
                    "item_sku": "Item SKU",
                    "item_name": "Item Name",
                    "opening_stock_units": "Opening Stock (Units)",
                    "inward_receipt_units": "Inward Receipts (Units)",
                    "outward_sold_units": "Outward Sold (Units)",
                    "expected_closing_stock": "Expected Closing Stock (Units)",
                    "actual_physical_audit_stock": "Actual Physical Audit Stock (Units)",
                    "inventory_shortage_variance": "Inventory Shortage Variance (Units)"
                })
                df_lube_ledger_nice.to_excel(writer, sheet_name="Lubricant Stock Inventory Book", index=False)

            # Write Inventory Reconciliation tab from wet_stock_recon
            try:
                from wet_stock_recon import generate_reconciliation_report_data
                df_inv_recon = generate_reconciliation_report_data(db_path=DB_PATH)
                if not df_inv_recon.empty:
                    df_inv_recon_nice = df_inv_recon.rename(columns={
                        "date": "Date",
                        "tank_id": "Tank ID",
                        "product_type": "Product Type",
                        "opening_volume": "Opening Volume (Liters)",
                        "inbound_receipts": "Inbound Receipts (Liters)",
                        "meter_sales_volume": "Meter Sales Volume (Liters)",
                        "expected_closing_volume": "Expected Closing Volume (Liters)",
                        "evening_dip_mm": "Evening Dip (mm)",
                        "actual_closing_volume": "Actual Closing Volume (Liters)",
                        "variance": "Variance (Liters)",
                        "shrinkage_limit_pct": "Shrinkage Limit (%)",
                        "shrinkage_limit_liters": "Shrinkage Limit (Liters)",
                        "status": "Status"
                    })
                    df_inv_recon_nice.to_excel(writer, sheet_name="Inventory Reconciliation", index=False)
                    logger.info("Inventory Reconciliation tab written to Excel workbook.")
                else:
                    logger.info("No inventory reconciliation data found — skipping Inventory Reconciliation tab.")
            except Exception as inv_err:
                logger.warning(f"Failed to write Inventory Reconciliation tab: {inv_err}")
            
        # Apply premium formatting styling
        apply_excel_styling(excel_path)
        
        logger.info(f"Database fully exported to {excel_path}!")
        return excel_path
        
    except Exception as e:
        logger.error(f"Failed to export database to Excel: {str(e)}")
        raise e

def generate_accounting_export(date_string: str = None) -> tuple[str, str]:
    """
    Extracts records from SQLite for a specific date (or all if None/all)
    and formats two distinct accounting exports inside /pump_exports folder:
    1. Excel workbook with sheets "Shift Readings" and "Ledger Entries" styled with openpyxl.
    2. PetroByte Sync CSV tailored for standard Indian pump accounting software imports.
    """
    import json
    from evaporation_handler import calculate_evaporation_allowances
    logger.info(f"Generating continuous accounting exports for date_string: {date_string}...")
    
    # Resolve folders
    backend_dir = os.path.dirname(os.path.abspath(__file__))
    WORKSPACE_DIR = os.path.dirname(backend_dir)
    EXPORTS_DIR = os.path.join(WORKSPACE_DIR, "pump_exports")
    os.makedirs(EXPORTS_DIR, exist_ok=True)
    
    d_suffix = date_string if date_string else "all"
    excel_path = os.path.join(EXPORTS_DIR, f"accounting_export_{d_suffix}.xlsx")
    csv_path = os.path.join(EXPORTS_DIR, f"petrobyte_sync_{d_suffix}.csv")
    
    try:
        conn = sqlite3.connect(DB_PATH)
        
        # 1. Fetch Daily Summary Data
        summary_query = "SELECT * FROM daily_summary"
        params = []
        if date_string and date_string.lower() != "all":
            summary_query += " WHERE date = ?"
            params.append(date_string)
        summary_query += " ORDER BY date DESC"
        df_summary = pd.read_sql_query(summary_query, conn, params=params)
        
        # 2. Fetch Ledger Entries Data
        entries_query = "SELECT * FROM ledger_entries"
        entry_params = []
        if date_string and date_string.lower() != "all":
            entries_query += " WHERE date = ?"
            entry_params.append(date_string)
        entries_query += " ORDER BY date DESC, entry_id DESC"
        df_entries = pd.read_sql_query(entries_query, conn, params=entry_params)
        
        # 3. Fetch Nozzle readings from daily_ledger table to format nozzle flows
        ledger_query = "SELECT date, raw_data FROM daily_ledger"
        ledger_params = []
        if date_string and date_string.lower() != "all":
            ledger_query += " WHERE date = ?"
            ledger_params.append(date_string)
        df_ledger = pd.read_sql_query(ledger_query, conn, params=ledger_params)
        
        # Ensure internal_consumption table exists
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='internal_consumption'")
        if not cursor.fetchone():
            cursor.execute("""
            CREATE TABLE IF NOT EXISTS internal_consumption (
                entry_id INTEGER PRIMARY KEY AUTOINCREMENT,
                date TEXT,
                product_type TEXT CHECK(product_type IN ('HSD', 'MS')),
                liters_drawn REAL DEFAULT 0.0,
                purpose_head TEXT,
                authorized_by TEXT
            )
            """)
            conn.commit()

        detailed_query = "SELECT * FROM internal_consumption"
        detailed_params = []
        if date_string and date_string.lower() != "all":
            detailed_query += " WHERE date = ?"
            detailed_params.append(date_string)
        detailed_query += " ORDER BY date DESC, entry_id DESC"

        monthly_query = """
            SELECT 
                strftime('%Y-%m', date) AS month,
                product_type,
                SUM(liters_drawn) AS cumulative_liters_drawn,
                COUNT(*) AS total_transactions
            FROM internal_consumption
        """
        monthly_params = []
        if date_string and date_string.lower() != "all":
            monthly_query += " WHERE strftime('%Y-%m', date) = strftime('%Y-%m', ?)"
            monthly_params.append(date_string)
        monthly_query += """
            GROUP BY month, product_type
            ORDER BY month DESC, product_type ASC
        """

        df_detailed = pd.read_sql_query(detailed_query, conn, params=detailed_params)
        df_monthly = pd.read_sql_query(monthly_query, conn, params=monthly_params)
        
        conn.close()
        
        # Decrypt fields for exports
        import json
        if not df_entries.empty:
            df_entries["party_name"] = df_entries["party_name"].apply(lambda val: decrypt_field(val, return_type=str))
            df_entries["amount"] = df_entries["amount"].apply(lambda val: decrypt_field(val, return_type=float))
            
        def decrypt_raw_json_str(raw_json_str):
            if not raw_json_str:
                return raw_json_str
            try:
                data = json.loads(raw_json_str)
                decrypted = decrypt_raw_data(data)
                return json.dumps(decrypted, ensure_ascii=False)
            except Exception:
                return raw_json_str
                
        if not df_ledger.empty:
            df_ledger["raw_data"] = df_ledger["raw_data"].apply(decrypt_raw_json_str)
        
        # =====================================================================
        # Pipeline 1: Excel Layout ("Shift Readings" & "Ledger Entries")
        # =====================================================================
        # Fetch all DSM shifts to build mapping in memory
        dsm_mapping = {}
        try:
            conn_dsm = sqlite3.connect(DB_PATH)
            cursor_dsm = conn_dsm.cursor()
            cursor_dsm.execute("SELECT date, dsm_name, assigned_nozzles FROM dsm_shifts")
            dsm_rows = cursor_dsm.fetchall()
            conn_dsm.close()
            
            for d_str_dsm, name, nozzles_str in dsm_rows:
                if nozzles_str:
                    if d_str_dsm not in dsm_mapping:
                        dsm_mapping[d_str_dsm] = {}
                    
                    nozzles_list = [n.strip() for n in nozzles_str.split(",") if n.strip()]
                    for n in nozzles_list:
                        # Clean and lowercase the nozzle name key (e.g. "ms-1")
                        n_clean = n.lower()
                        # Clean nozzle name to strip any trailing descriptions if present
                        if " (" in n_clean:
                            n_clean = n_clean.split(" (")[0]
                        dsm_mapping[d_str_dsm][n_clean] = name
        except Exception as dsm_err:
            logger.warning(f"Failed to load DSM mapping for Excel export: {str(dsm_err)}")

        shift_readings = []
        for idx, row in df_summary.iterrows():
            d_str = row["date"]
            total_liters = float(row["total_hsd_liters"] or 0.0) + float(row["total_ms_liters"] or 0.0)
            fuel_cash = float(row["total_cash_calculated"] or 0.0)
            
            # Format nozzle flow summaries
            flows = "N/A"
            matching_ledger = df_ledger[df_ledger["date"] == d_str]
            if not matching_ledger.empty:
                try:
                    raw_json = json.loads(matching_ledger.iloc[0]["raw_data"])
                    flow_items = []
                    for n in raw_json.get("nozzles", []):
                        flow_items.append(f"{n.get('nozzle_name')}: {n.get('sales_liters_calculated') or n.get('net_sales_liters') or 0.0} L")
                    flows = " | ".join(flow_items)
                except Exception:
                    pass
            
            if flows == "N/A":
                flows = f"HSD: {row['total_hsd_liters']} L, MS: {row['total_ms_liters']} L"
                
            # Resolve DSM staff responsible for nozzles on this date
            staff_list = []
            date_mapping = dsm_mapping.get(d_str, {})
            if date_mapping:
                staff_list = [f"{n.upper()}: {staff}" for n, staff in sorted(date_mapping.items())]
            staff_str = " | ".join(staff_list) if staff_list else "N/A"
            
            # Evaporation calculations
            try:
                evap = calculate_evaporation_allowances(d_str)
            except Exception as evap_err:
                logger.warning(f"Failed to calculate evaporation for {d_str}: {evap_err}")
                evap = {
                    "hsd_permissible_evaporation_liters": 0.0,
                    "hsd_actual_shortage_liters": 0.0,
                    "hsd_normal_evaporation_loss_liters": 0.0,
                    "hsd_abnormal_shortage_liters": 0.0,
                    "hsd_classification": "N/A",
                    "ms_permissible_evaporation_liters": 0.0,
                    "ms_actual_shortage_liters": 0.0,
                    "ms_normal_evaporation_loss_liters": 0.0,
                    "ms_abnormal_shortage_liters": 0.0,
                    "ms_classification": "N/A"
                }

            reg_hsd = float(row["total_regular_hsd_liters"]) if "total_regular_hsd_liters" in df_summary.columns else float(row["total_hsd_liters"] or 0.0)
            prem_hsd = float(row["total_premium_hsd_liters"]) if "total_premium_hsd_liters" in df_summary.columns else 0.0
            reg_ms = float(row["total_regular_ms_liters"]) if "total_regular_ms_liters" in df_summary.columns else float(row["total_ms_liters"] or 0.0)
            prem_ms = float(row["total_premium_ms_liters"]) if "total_premium_ms_liters" in df_summary.columns else 0.0
            
            # If the database columns exist but are NaN or None, fill with defaults
            import math
            if math.isnan(reg_hsd) or reg_hsd is None: reg_hsd = float(row["total_hsd_liters"] or 0.0)
            if math.isnan(prem_hsd) or prem_hsd is None: prem_hsd = 0.0
            if math.isnan(reg_ms) or reg_ms is None: reg_ms = float(row["total_ms_liters"] or 0.0)
            if math.isnan(prem_ms) or prem_ms is None: prem_ms = 0.0

            gross_usd = float(row["gross_spread_usd"] or 0.0) if "gross_spread_usd" in df_summary.columns else 0.0
            realized = float(row["realized_profit"] or 0.0) if "realized_profit" in df_summary.columns else 0.0

            shift_readings.append({
                "Date": d_str,
                "Nozzle Flows": flows,
                "Regular HSD Sold (L)": reg_hsd,
                "Premium HSD Sold (L)": prem_hsd,
                "Regular MS Sold (L)": reg_ms,
                "Premium MS Sold (L)": prem_ms,
                "Total Liters Sold": total_liters,
                "Calculated Fuel Cash": fuel_cash,
                "Gross Spread (USD)": gross_usd,
                "Realized Profit": realized,
                "Active Nozzle Staff": staff_str,
                "HSD Permissible Evaporation Loss (L)": evap["hsd_permissible_evaporation_liters"],
                "HSD Actual Shortage (L)": evap["hsd_actual_shortage_liters"],
                "HSD Normal Evaporation Loss (L)": evap["hsd_normal_evaporation_loss_liters"],
                "HSD Abnormal Shortage (L)": evap["hsd_abnormal_shortage_liters"],
                "HSD Loss Classification": evap["hsd_classification"],
                "MS Permissible Evaporation Loss (L)": evap["ms_permissible_evaporation_liters"],
                "MS Actual Shortage (L)": evap["ms_actual_shortage_liters"],
                "MS Normal Evaporation Loss (L)": evap["ms_normal_evaporation_loss_liters"],
                "MS Abnormal Shortage (L)": evap["ms_abnormal_shortage_liters"],
                "MS Loss Classification": evap["ms_classification"]
            })
            
        df_shift_nice = pd.DataFrame(shift_readings) if shift_readings else pd.DataFrame(columns=[
            "Date", "Nozzle Flows", "Regular HSD Sold (L)", "Premium HSD Sold (L)", "Regular MS Sold (L)", "Premium MS Sold (L)",
            "Total Liters Sold", "Calculated Fuel Cash", "Gross Spread (USD)", "Realized Profit", "Active Nozzle Staff",
            "HSD Permissible Evaporation Loss (L)", "HSD Actual Shortage (L)", "HSD Normal Evaporation Loss (L)", "HSD Abnormal Shortage (L)", "HSD Loss Classification",
            "MS Permissible Evaporation Loss (L)", "MS Actual Shortage (L)", "MS Normal Evaporation Loss (L)", "MS Abnormal Shortage (L)", "MS Loss Classification"
        ])
        
        # Calculate totals for shift readings sheet
        if not df_shift_nice.empty:
            totals_shift = {col: "" for col in df_shift_nice.columns}
            totals_shift[df_shift_nice.columns[0]] = "Profit Accounting Totals"
            
            for col in df_shift_nice.columns:
                if col in [
                    "Regular HSD Sold (L)", "Premium HSD Sold (L)",
                    "Regular MS Sold (L)", "Premium MS Sold (L)",
                    "Total Liters Sold", "Calculated Fuel Cash",
                    "Gross Spread (USD)", "Realized Profit"
                ]:
                    totals_shift[col] = df_shift_nice[col].sum()
            
            totals_shift_df = pd.DataFrame([totals_shift])
            df_shift_nice = pd.concat([df_shift_nice, totals_shift_df], ignore_index=True)
        
        ledger_nice = []
        for idx, row in df_entries.iterrows():
            ledger_nice.append({
                "Date": row["date"],
                "Account Head": row["party_name"] or "Unknown Head",
                "Wheel No": row["vehicle_wheel_no"] or "N/A",
                "Amount": float(row["amount"] or 0.0),
                "Transaction Type": row["type"]
            })
            
        df_ledger_nice = pd.DataFrame(ledger_nice) if ledger_nice else pd.DataFrame(columns=["Date", "Account Head", "Wheel No", "Amount", "Transaction Type"])
        
        # Map nice column headers for Internal Consumption Monthly Cumulative
        df_monthly_nice = df_monthly.rename(columns={
            "month": "Month",
            "product_type": "Product Type",
            "cumulative_liters_drawn": "Cumulative Liters Drawn (Liters)",
            "total_transactions": "Total Transactions"
        })
        
        # Map nice column headers for Internal Consumption Detailed Log
        df_detailed_nice = df_detailed.rename(columns={
            "entry_id": "Entry ID",
            "date": "Date",
            "product_type": "Product Type",
            "liters_drawn": "Liters Drawn (Liters)",
            "purpose_head": "Purpose Head",
            "authorized_by": "Authorized By"
        })
        
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df_shift_nice.to_excel(writer, sheet_name="Shift Readings", index=False)
            df_ledger_nice.to_excel(writer, sheet_name="Ledger Entries", index=False)
            
            # Write first table
            df_monthly_nice.to_excel(writer, sheet_name="Internal Consumption", index=False, startrow=0, startcol=0)
            
            # Write second table
            start_row_detailed = len(df_monthly_nice) + 4
            df_detailed_nice.to_excel(writer, sheet_name="Internal Consumption", index=False, startrow=start_row_detailed, startcol=0)
            
            # Get sheet to write section banner
            ws = writer.sheets["Internal Consumption"]
            ws.cell(row=start_row_detailed, column=1, value="DETAILED INTERNAL FUEL DRAW LOGS")
            
        apply_excel_styling(excel_path)
        logger.info(f"Excel Accounting Export saved and styled at: {excel_path}")
        
        # =====================================================================
        # Pipeline 2: PetroByte Sync Export CSV
        # =====================================================================
        petrobyte_records = []
        for idx, row in df_summary.iterrows():
            d_str = row["date"]
            total_cash = float(row["total_cash_calculated"] or 0.0)
            total_credit = float(row["total_credit_sales"] or 0.0)
            
            # Net Cash sales journal voucher receipt
            net_cash = max(0.0, total_cash - total_credit)
            if net_cash > 0:
                # Credit leg (Cash Sale account)
                petrobyte_records.append({
                    "Date": d_str,
                    "Ledger Name": "Cash Sale",
                    "Voucher Type": "Receipt",
                    "Account Debit": 0.0,
                    "Account Credit": net_cash,
                    "Narration": "Daily net cash fuel sales"
                })
                # Debit leg (Cash / Bank account)
                petrobyte_records.append({
                    "Date": d_str,
                    "Ledger Name": "Cash",
                    "Voucher Type": "Receipt",
                    "Account Debit": net_cash,
                    "Account Credit": 0.0,
                    "Narration": "Daily net cash fuel sales"
                })
                
        for idx, row in df_entries.iterrows():
            d_str = row["date"]
            amount = float(row["amount"] or 0.0)
            party = row["party_name"] or "Unknown Head"
            e_type = row["type"]
            v_no = row.get("vehicle_wheel_no") or "N/A"
            rem = row.get("remarks") or ""
            
            if e_type == "udhaar":
                # Debit leg (Customer Account)
                petrobyte_records.append({
                    "Date": d_str,
                    "Ledger Name": party,
                    "Voucher Type": "Sale",
                    "Account Debit": amount,
                    "Account Credit": 0.0,
                    "Narration": f"Credit sale to {party} - vehicle {v_no} - {rem}"
                })
                # Credit leg (Sales Account)
                petrobyte_records.append({
                    "Date": d_str,
                    "Ledger Name": "Sales",
                    "Voucher Type": "Sale",
                    "Account Debit": 0.0,
                    "Account Credit": amount,
                    "Narration": f"Credit sale to {party} - vehicle {v_no} - {rem}"
                })
            elif e_type == "expense":
                # Debit leg (Expense Head)
                petrobyte_records.append({
                    "Date": d_str,
                    "Ledger Name": party,
                    "Voucher Type": "Payment",
                    "Account Debit": amount,
                    "Account Credit": 0.0,
                    "Narration": f"Expense: {rem}"
                })
                # Credit leg (Cash Account)
                petrobyte_records.append({
                    "Date": d_str,
                    "Ledger Name": "Cash",
                    "Voucher Type": "Payment",
                    "Account Debit": 0.0,
                    "Account Credit": amount,
                    "Narration": f"Expense: {rem}"
                })
                
        temp_csv_path = csv_path + ".tmp"
        df_petrobyte = pd.DataFrame(petrobyte_records) if petrobyte_records else pd.DataFrame(
            columns=["Date", "Ledger Name", "Voucher Type", "Account Debit", "Account Credit", "Narration"]
        )
        df_petrobyte.to_csv(temp_csv_path, index=False, encoding="utf-8-sig")
        
        try:
            from petrobyte_validator import validate_petrobyte_csv_format
            validate_petrobyte_csv_format(temp_csv_path)
            
            if os.path.exists(csv_path):
                os.remove(csv_path)
            os.rename(temp_csv_path, csv_path)
            logger.info(f"PetroByte CSV Sync Export validated and saved at: {csv_path}")
        except Exception as validation_err:
            if os.path.exists(temp_csv_path):
                os.remove(temp_csv_path)
            logger.error(f"PetroByte compliance validation failed, export blocked: {str(validation_err)}")
            raise validation_err
            
        return excel_path, csv_path
    except Exception as e:
        logger.error(f"Failed to generate accounting export: {str(e)}")
        raise e

def export_density_register_pdf(output_path: str = None) -> str:
    """
    Generates a statutory Density Register PDF report formatted to match the standard
    layout verified during official oil company (IOC/BPCL/HPCL) inspections.
    Uses PyMuPDF (fitz) for precise, grid-based tabular PDF rendering.
    
    Output file: Statutory_Density_Register.pdf inside pump_exports folder.
    """
    try:
        import fitz  # PyMuPDF
    except ImportError:
        logger.error("PyMuPDF (fitz) is not installed. Run: pip install pymupdf")
        raise ImportError("PyMuPDF is required for PDF generation. Install it with: pip install pymupdf")

    from density_logger import get_density_records

    logger.info("Generating Statutory Density Register PDF report...")

    # Resolve output path
    if not output_path:
        backend_dir = os.path.dirname(os.path.abspath(__file__))
        workspace_dir = os.path.dirname(backend_dir)
        exports_dir = os.path.join(workspace_dir, "pump_exports")
        os.makedirs(exports_dir, exist_ok=True)
        output_path = os.path.join(exports_dir, "Statutory_Density_Register.pdf")

    # Fetch all density records
    records = get_density_records()

    # -----------------------------------------------------------------------
    # Page / layout constants (A4 portrait)
    # -----------------------------------------------------------------------
    PAGE_W, PAGE_H = 595, 842          # A4 points
    MARGIN_L, MARGIN_R = 30, 30
    MARGIN_T, MARGIN_B = 50, 40
    TABLE_W = PAGE_W - MARGIN_L - MARGIN_R
    ROW_H = 20                          # Data row height (pt)
    HEADER_H = 28                       # Column header row height
    TITLE_H = 40                        # Document title strip height

    # Column proportions  (must sum to 1.0)
    COL_PROPS = [0.10, 0.09, 0.13, 0.13, 0.14, 0.14, 0.12, 0.15]
    COL_HEADERS = [
        "Date", "Product", "Obs. Temp\n(°C)", "Obs. Density\n(kg/m³)",
        "Conv. Density\n@ 15°C (kg/m³)", "Invoice Ref.\n(kg/m³)",
        "Variation\n(kg/m³)", "Compliance\nStatus"
    ]

    COL_WIDTHS = [TABLE_W * p for p in COL_PROPS]

    # Colour palette
    C_NAVY      = (0.122, 0.29, 0.49)    # #1F4A7E  header background
    C_WHITE     = (1, 1, 1)
    C_BLACK     = (0, 0, 0)
    C_PASS      = (0.0, 0.55, 0.27)      # Green  ✓ PASS
    C_FAIL      = (0.75, 0.11, 0.11)     # Red    ✗ FAIL
    C_ALT_ROW   = (0.94, 0.97, 1.0)      # Light-blue alternating row fill
    C_GRID      = (0.65, 0.65, 0.65)     # Grid-line colour
    C_TITLE_BG  = (0.08, 0.18, 0.34)     # Dark navy title strip

    # -----------------------------------------------------------------------
    # Helper: draw a cell with optional fill, centred/left text and grid border
    # -----------------------------------------------------------------------
    def draw_cell(page, x, y, w, h, text, font_size=7.5,
                  bold=False, fill=None, text_color=C_BLACK, align="center"):
        rect = fitz.Rect(x, y, x + w, y + h)

        # Fill background
        if fill:
            page.draw_rect(rect, color=None, fill=fill, width=0)

        # Grid border (always drawn, thin line)
        page.draw_rect(rect, color=C_GRID, fill=None, width=0.5)

        # Text insertion – split on newlines for multi-line headers
        lines = str(text).split("\n")
        font_name = "Helvetica" if not bold else "Helvetica-Bold"
        line_h = font_size * 1.35
        total_text_h = line_h * len(lines)
        y_start = y + (h - total_text_h) / 2 + font_size * 0.85

        for i, line in enumerate(lines):
            line = line.strip()
            if not line:
                continue
            text_w = fitz.get_text_length(line, fontname=font_name, fontsize=font_size)
            if align == "center":
                x_text = x + (w - text_w) / 2
            elif align == "right":
                x_text = x + w - text_w - 3
            else:
                x_text = x + 4  # left with small indent

            page.insert_text(
                (x_text, y_start + i * line_h),
                line,
                fontname=font_name,
                fontsize=font_size,
                color=text_color
            )

    # -----------------------------------------------------------------------
    # Pagination helper
    # -----------------------------------------------------------------------
    ROWS_PER_PAGE = int((PAGE_H - MARGIN_T - MARGIN_B - TITLE_H - HEADER_H) // ROW_H)

    doc = fitz.open()
    total_pages = max(1, -(-len(records) // ROWS_PER_PAGE))  # ceiling division

    for page_num in range(total_pages):
        page = doc.new_page(width=PAGE_W, height=PAGE_H)
        page_records = records[page_num * ROWS_PER_PAGE:(page_num + 1) * ROWS_PER_PAGE]

        # ---------------------------------------------------------------
        # Title strip
        # ---------------------------------------------------------------
        title_rect = fitz.Rect(MARGIN_L, MARGIN_T, PAGE_W - MARGIN_R, MARGIN_T + TITLE_H)
        page.draw_rect(title_rect, color=None, fill=C_TITLE_BG, width=0)

        title_text = "STATUTORY DENSITY REGISTER"
        subtitle_text = "Daily Fuel Quality Compliance Log — ASTM D1250 / IS:1460"
        title_font_size = 13
        sub_font_size = 7.5

        tw = fitz.get_text_length(title_text, fontname="Helvetica-Bold", fontsize=title_font_size)
        page.insert_text(
            ((PAGE_W - tw) / 2, MARGIN_T + 16),
            title_text, fontname="Helvetica-Bold", fontsize=title_font_size, color=C_WHITE
        )
        sw = fitz.get_text_length(subtitle_text, fontname="Helvetica", fontsize=sub_font_size)
        page.insert_text(
            ((PAGE_W - sw) / 2, MARGIN_T + 30),
            subtitle_text, fontname="Helvetica", fontsize=sub_font_size, color=(0.8, 0.88, 1.0)
        )

        # Page number (bottom right)
        pg_text = f"Page {page_num + 1} of {total_pages}"
        pgw = fitz.get_text_length(pg_text, fontname="Helvetica", fontsize=7)
        page.insert_text(
            (PAGE_W - MARGIN_R - pgw, PAGE_H - MARGIN_B + 12),
            pg_text, fontname="Helvetica", fontsize=7, color=(0.4, 0.4, 0.4)
        )

        # Generation timestamp (bottom left)
        import datetime
        ts = datetime.datetime.now().strftime("Generated: %d-%b-%Y %H:%M")
        page.insert_text(
            (MARGIN_L, PAGE_H - MARGIN_B + 12),
            ts, fontname="Helvetica", fontsize=7, color=(0.4, 0.4, 0.4)
        )

        # ---------------------------------------------------------------
        # Column header row
        # ---------------------------------------------------------------
        header_y = MARGIN_T + TITLE_H + 4
        x_cursor = MARGIN_L
        for col_idx, (hdr, cw) in enumerate(zip(COL_HEADERS, COL_WIDTHS)):
            draw_cell(
                page, x_cursor, header_y, cw, HEADER_H,
                hdr, font_size=7, bold=True,
                fill=C_NAVY, text_color=C_WHITE, align="center"
            )
            x_cursor += cw

        # ---------------------------------------------------------------
        # Data rows
        # ---------------------------------------------------------------
        row_y = header_y + HEADER_H
        for row_idx, rec in enumerate(page_records):
            alt_fill = C_ALT_ROW if row_idx % 2 == 1 else C_WHITE
            x_cursor = MARGIN_L

            passed = rec.get("permissible_variation_passed", True)
            variation = rec.get("variation", 0.0)

            row_values = [
                rec.get("date", ""),
                rec.get("product_type", ""),
                f"{rec.get('observed_temperature_celsius', 0.0):.1f}",
                f"{rec.get('observed_density_raw', 0.0):.2f}",
                f"{rec.get('converted_density_at_15c', 0.0):.2f}",
                f"{rec.get('invoice_density_reference', 0.0):.2f}",
                f"{variation:+.2f}",
                "✓ PASS" if passed else "✗ FAIL"
            ]

            for col_idx, (val, cw) in enumerate(zip(row_values, COL_WIDTHS)):
                # Override text colour for compliance column
                if col_idx == 7:
                    txt_color = C_PASS if passed else C_FAIL
                    bold = True
                elif col_idx == 6 and not passed:
                    txt_color = C_FAIL
                    bold = True
                else:
                    txt_color = C_BLACK
                    bold = False

                text_align = "center" if col_idx in (0, 1, 7) else "right"
                draw_cell(
                    page, x_cursor, row_y, cw, ROW_H,
                    val, font_size=7.5, bold=bold,
                    fill=alt_fill, text_color=txt_color, align=text_align
                )
                x_cursor += cw

            row_y += ROW_H

        # Empty-state notice if no records exist on this page
        if not page_records:
            empty_y = header_y + HEADER_H + ROW_H
            empty_txt = "No density compliance records found in the database."
            ew = fitz.get_text_length(empty_txt, fontname="Helvetica", fontsize=9)
            page.insert_text(
                ((PAGE_W - ew) / 2, empty_y + 15),
                empty_txt, fontname="Helvetica", fontsize=9, color=(0.5, 0.5, 0.5)
            )

        # ---------------------------------------------------------------
        # Summary legend strip at bottom of last page
        # ---------------------------------------------------------------
        if page_num == total_pages - 1:
            legend_y = PAGE_H - MARGIN_B - 20
            pass_count = sum(1 for r in records if r.get("permissible_variation_passed"))
            fail_count = len(records) - pass_count
            legend = (
                f"Total Records: {len(records)}     "
                f"PASS: {pass_count}     FAIL: {fail_count}     "
                f"Permissible Variation Threshold: ±3.0 kg/m³"
            )
            lw = fitz.get_text_length(legend, fontname="Helvetica", fontsize=7.5)
            page.insert_text(
                ((PAGE_W - lw) / 2, legend_y),
                legend, fontname="Helvetica", fontsize=7.5, color=(0.25, 0.25, 0.25)
            )

    doc.save(output_path)
    doc.close()
    logger.info(f"Statutory Density Register PDF saved at: {output_path}")
    return output_path


if __name__ == "__main__":
    # If run directly, perform a self-test by populating mock data and exporting
    logger.info("Running exporter self-test...")
    
    # 1. Initialize DB tables unconditionally
    from init_db import initialize_database
    initialize_database()
        
    # 2. Add mock data for testing
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # Insert test daily summary
        logger.info("Inserting mock entries for self-test...")
        cursor.execute("""
            INSERT OR REPLACE INTO daily_summary (date, total_hsd_liters, total_ms_liters, total_cash_calculated, total_credit_sales, total_testing_deductions, is_verified)
            VALUES ('2026-05-30', 450.50, 310.20, 48250.00, 12500.00, 5.0, 1)
        """)
        cursor.execute("""
            INSERT OR REPLACE INTO daily_summary (date, total_hsd_liters, total_ms_liters, total_cash_calculated, total_credit_sales, total_testing_deductions, is_verified)
            VALUES ('2026-05-31', 510.80, 290.40, 52100.00, 9400.00, 5.0, 0)
        """)
        
        # Insert test ledger entries
        cursor.execute("""
            INSERT INTO ledger_entries (date, party_name, vehicle_wheel_no, amount, type, remarks)
            VALUES ('2026-05-30', 'Rahul Transport', 'HR-38-F-1234', 8500.00, 'udhaar', 'HSD credit sale')
        """)
        cursor.execute("""
            INSERT INTO ledger_entries (date, party_name, vehicle_wheel_no, amount, type, remarks)
            VALUES ('2026-05-30', 'Local Cash Office', 'N/A', 150.00, 'expense', 'Office Tea & Cleaning')
        """)
        cursor.execute("""
            INSERT INTO ledger_entries (date, party_name, vehicle_wheel_no, amount, type, remarks)
            VALUES ('2026-05-31', 'SBI Bank Drop', 'N/A', 40000.00, 'bank_drop', 'Daily cash collection bank deposit')
        """)
        
        conn.commit()
        conn.close()
        
        # 3. Trigger export
        export_db_to_excel("ledger_test.xlsx")
        logger.info("Self-test completed successfully. Visualised outputs at ledger_test.xlsx")
        
    except Exception as e:
        logger.error(f"Self-test failed: {str(e)}")
