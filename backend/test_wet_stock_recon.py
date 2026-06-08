#!/usr/bin/env python3
"""
Automated tests for backend/wet_stock_recon.py.

Covers:
- Opening stock lookup resolves to previous date's closing stock
- Inbound receipts and meter sales are fetched correctly
- Variance threshold flags: Normal Handling Shrinkage vs Abnormal Product Leakage Alert
- generate_reconciliation_report_data returns a correct Pandas DataFrame
- export_db_to_excel writes the "Inventory Reconciliation" tab
"""

import os
import sys
import sqlite3
import unittest
import tempfile

BACKEND_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, BACKEND_DIR)

import wet_stock_recon


def _make_test_db(path: str):
    """Creates a minimal test database with all required tables."""
    conn = sqlite3.connect(path)
    cursor = conn.cursor()

    # stock_recon
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS stock_recon (
        date TEXT PRIMARY KEY,
        hsd_opening_dip_liters REAL DEFAULT 0.0,
        hsd_receipt_liters REAL DEFAULT 0.0,
        hsd_closing_dip_liters REAL DEFAULT 0.0,
        ms_opening_dip_liters REAL DEFAULT 0.0,
        ms_receipt_liters REAL DEFAULT 0.0,
        ms_closing_dip_liters REAL DEFAULT 0.0,
        actual_cash_deposited REAL DEFAULT 0.0,
        digital_wallet_settlements REAL DEFAULT 0.0,
        logged_udhaar_entries REAL DEFAULT 0.0
    )
    """)

    # daily_summary (sales volumes)
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS daily_summary (
        date TEXT PRIMARY KEY,
        total_hsd_liters REAL DEFAULT 0.0,
        total_ms_liters REAL DEFAULT 0.0,
        total_cash_calculated REAL DEFAULT 0.0,
        total_credit_sales REAL DEFAULT 0.0,
        total_testing_deductions REAL DEFAULT 0.0,
        is_verified INTEGER DEFAULT 0,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
    """)

    # tanker_receipts
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS tanker_receipts (
        invoice_no TEXT PRIMARY KEY,
        date TEXT,
        tank_lorry_no TEXT,
        product_type TEXT,
        invoice_volume_liters REAL,
        invoice_density_at_15c REAL,
        observed_compartment_dips_mm TEXT,
        observed_density_raw REAL,
        observed_temperature_celsius REAL,
        actual_received_volume_liters REAL,
        transit_shortage_liters REAL
    )
    """)

    # tank_calibration_charts (empty → convert_dip_to_liters fallback)
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS tank_calibration_charts (
        tank_id TEXT,
        dip_level_mm INTEGER,
        volume_liters REAL,
        PRIMARY KEY (tank_id, dip_level_mm)
    )
    """)

    # tank_tilt_profiles (required by dip_profiler via tank_calibration)
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS tank_tilt_profiles (
        tank_id TEXT PRIMARY KEY,
        tilt_coefficient REAL DEFAULT 0.0,
        ols_intercept REAL DEFAULT 0.0,
        r_squared REAL DEFAULT 0.0,
        n_points INTEGER DEFAULT 0,
        analysis_start_date TEXT,
        analysis_end_date TEXT,
        anomaly_type TEXT DEFAULT 'none',
        anomaly_confidence REAL DEFAULT 0.0,
        correction_active INTEGER DEFAULT 0,
        last_updated TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
    """)

    # tank_dip_log (required by dip_profiler)
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS tank_dip_log (
        log_id INTEGER PRIMARY KEY AUTOINCREMENT,
        tank_id TEXT NOT NULL,
        reading_date TEXT NOT NULL,
        dip_mm REAL NOT NULL,
        chart_volume_L REAL,
        actual_variance_L REAL,
        meter_check_ok INTEGER DEFAULT 1,
        source TEXT DEFAULT 'manual',
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(tank_id, reading_date)
    )
    """)

    conn.commit()
    conn.close()


class TestGetOpeningStock(unittest.TestCase):
    """Test resolution of opening stock from the previous date's closing entry."""

    def setUp(self):
        self.tmp = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
        self.tmp.close()
        self.db = self.tmp.name
        _make_test_db(self.db)

    def tearDown(self):
        os.unlink(self.db)

    def test_previous_date_closing_used_as_opening(self):
        """The previous day's closing volume is returned as the opening stock."""
        conn = sqlite3.connect(self.db)
        conn.execute("""
            INSERT INTO stock_recon (date, hsd_closing_dip_liters, ms_closing_dip_liters)
            VALUES ('2026-05-29', 10000.0, 8000.0)
        """)
        conn.commit()
        conn.close()

        hsd_opening = wet_stock_recon.get_opening_stock("Tank_1_HSD", "2026-05-30", self.db)
        ms_opening = wet_stock_recon.get_opening_stock("Tank_2_MS", "2026-05-30", self.db)

        self.assertEqual(hsd_opening, 10000.0)
        self.assertEqual(ms_opening, 8000.0)

    def test_fallback_to_current_day_opening(self):
        """If no previous date exists, returns the current day's opening stock."""
        conn = sqlite3.connect(self.db)
        conn.execute("""
            INSERT INTO stock_recon (date, hsd_opening_dip_liters, ms_opening_dip_liters)
            VALUES ('2026-05-30', 12500.0, 9500.0)
        """)
        conn.commit()
        conn.close()

        hsd_opening = wet_stock_recon.get_opening_stock("Tank_1_HSD", "2026-05-30", self.db)
        self.assertEqual(hsd_opening, 12500.0)

    def test_returns_zero_if_no_data(self):
        """Returns 0.0 when no data exists at all."""
        result = wet_stock_recon.get_opening_stock("Tank_1_HSD", "2026-05-30", self.db)
        self.assertEqual(result, 0.0)


class TestGetInboundReceipts(unittest.TestCase):
    """Test inbound receipts summing from tanker_receipts."""

    def setUp(self):
        self.tmp = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
        self.tmp.close()
        self.db = self.tmp.name
        _make_test_db(self.db)

    def tearDown(self):
        os.unlink(self.db)

    def test_sums_multiple_tanker_receipts(self):
        """Multiple HSD receipts on the same date are summed correctly."""
        conn = sqlite3.connect(self.db)
        conn.execute("""
            INSERT INTO tanker_receipts (invoice_no, date, product_type, actual_received_volume_liters)
            VALUES ('INV001', '2026-05-30', 'HSD', 5000.0)
        """)
        conn.execute("""
            INSERT INTO tanker_receipts (invoice_no, date, product_type, actual_received_volume_liters)
            VALUES ('INV002', '2026-05-30', 'HSD', 3000.0)
        """)
        conn.commit()
        conn.close()

        receipts = wet_stock_recon.get_inbound_receipts("Tank_1_HSD", "2026-05-30", self.db)
        self.assertEqual(receipts, 8000.0)

    def test_no_receipts_returns_zero(self):
        """Returns 0.0 when no receipts recorded for that date."""
        receipts = wet_stock_recon.get_inbound_receipts("Tank_1_HSD", "2026-05-30", self.db)
        self.assertEqual(receipts, 0.0)

    def test_product_type_isolation(self):
        """MS receipts are not included in HSD totals."""
        conn = sqlite3.connect(self.db)
        conn.execute("""
            INSERT INTO tanker_receipts (invoice_no, date, product_type, actual_received_volume_liters)
            VALUES ('INVMS1', '2026-05-30', 'MS', 4000.0)
        """)
        conn.commit()
        conn.close()

        hsd_receipts = wet_stock_recon.get_inbound_receipts("Tank_1_HSD", "2026-05-30", self.db)
        ms_receipts = wet_stock_recon.get_inbound_receipts("Tank_2_MS", "2026-05-30", self.db)
        self.assertEqual(hsd_receipts, 0.0)
        self.assertEqual(ms_receipts, 4000.0)


class TestGetMeterSalesVolume(unittest.TestCase):
    """Test meter sales volume retrieval from daily_summary."""

    def setUp(self):
        self.tmp = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
        self.tmp.close()
        self.db = self.tmp.name
        _make_test_db(self.db)

    def tearDown(self):
        os.unlink(self.db)

    def test_hsd_and_ms_sales_returned_correctly(self):
        conn = sqlite3.connect(self.db)
        conn.execute("""
            INSERT INTO daily_summary (date, total_hsd_liters, total_ms_liters)
            VALUES ('2026-05-30', 1200.0, 700.0)
        """)
        conn.commit()
        conn.close()

        hsd = wet_stock_recon.get_meter_sales_volume("Tank_1_HSD", "2026-05-30", self.db)
        ms = wet_stock_recon.get_meter_sales_volume("Tank_2_MS", "2026-05-30", self.db)
        self.assertEqual(hsd, 1200.0)
        self.assertEqual(ms, 700.0)


class TestReconcileTankWetStock(unittest.TestCase):
    """Test the full reconcile_tank_wet_stock function."""

    def setUp(self):
        self.tmp = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
        self.tmp.close()
        self.db = self.tmp.name
        _make_test_db(self.db)

        conn = sqlite3.connect(self.db)
        # Insert previous day's closing
        conn.execute("""
            INSERT INTO stock_recon (date, hsd_closing_dip_liters, ms_closing_dip_liters)
            VALUES ('2026-05-29', 12000.0, 8000.0)
        """)
        # Insert current day's stock_recon entry (closing volumes stored in liters)
        conn.execute("""
            INSERT INTO stock_recon (date, hsd_opening_dip_liters, hsd_receipt_liters, hsd_closing_dip_liters,
                                     ms_opening_dip_liters, ms_receipt_liters, ms_closing_dip_liters)
            VALUES ('2026-05-30', 12000.0, 5000.0, 15800.0, 8000.0, 0.0, 7300.0)
        """)
        # Insert daily_summary meter sales
        conn.execute("""
            INSERT INTO daily_summary (date, total_hsd_liters, total_ms_liters)
            VALUES ('2026-05-30', 1200.0, 700.0)
        """)
        # Insert a tanker receipt row so get_inbound_receipts returns the 5000L correctly
        conn.execute("""
            INSERT INTO tanker_receipts (invoice_no, date, tank_lorry_no, product_type,
                invoice_volume_liters, invoice_density_at_15c, observed_compartment_dips_mm,
                observed_density_raw, observed_temperature_celsius, actual_received_volume_liters,
                transit_shortage_liters)
            VALUES ('TEST001', '2026-05-30', 'TN01AB1234', 'HSD',
                5000.0, 820.0, '500', 820.0, 30.0, 5000.0, 0.0)
        """)
        conn.commit()
        conn.close()

    def tearDown(self):
        os.unlink(self.db)

    def test_balanced_reconciliation(self):
        """Expected closing matches actual closing → variance should be 0.0."""
        conn = sqlite3.connect(self.db)
        # Adjust closing to match expected: 12000 + 5000 - 1200 = 15800
        conn.execute("UPDATE stock_recon SET hsd_closing_dip_liters = 15800.0 WHERE date = '2026-05-30'")
        conn.commit()
        conn.close()

        results = wet_stock_recon.reconcile_tank_wet_stock("2026-05-30", db_path=self.db)
        hsd = next(r for r in results if r["tank_id"] == "Tank_1_HSD")

        self.assertEqual(hsd["opening_volume"], 12000.0)
        self.assertEqual(hsd["inbound_receipts"], 5000.0)
        self.assertEqual(hsd["meter_sales_volume"], 1200.0)
        self.assertEqual(hsd["expected_closing_volume"], 15800.0)
        self.assertEqual(hsd["actual_closing_volume"], 15800.0)
        self.assertEqual(hsd["variance"], 0.0)

    def test_normal_shrinkage_classification(self):
        """
        A loss just under the threshold should be tagged 'Normal Handling Shrinkage'.
        HSD threshold = 0.20% × 1200L = 2.40L
        Closing: 15800 - 2 = 15798.0 → variance = -2.0 → within threshold
        """
        conn = sqlite3.connect(self.db)
        conn.execute("UPDATE stock_recon SET hsd_closing_dip_liters = 15798.0 WHERE date = '2026-05-30'")
        conn.commit()
        conn.close()

        results = wet_stock_recon.reconcile_tank_wet_stock("2026-05-30", db_path=self.db)
        hsd = next(r for r in results if r["tank_id"] == "Tank_1_HSD")
        self.assertEqual(hsd["status"], "Normal Handling Shrinkage")

    def test_abnormal_leakage_classification(self):
        """
        A loss exceeding the threshold should be tagged 'Abnormal Product Leakage Alert'.
        MS threshold = 0.60% × 700L = 4.20L
        Closing: 8000 - 700 - 10 = 7290 → variance = 7290 - 7300 = -10.0 → exceeds threshold
        """
        conn = sqlite3.connect(self.db)
        conn.execute("UPDATE stock_recon SET ms_closing_dip_liters = 7290.0 WHERE date = '2026-05-30'")
        conn.commit()
        conn.close()

        results = wet_stock_recon.reconcile_tank_wet_stock("2026-05-30", db_path=self.db)
        ms = next(r for r in results if r["tank_id"] == "Tank_2_MS")
        self.assertEqual(ms["status"], "Abnormal Product Leakage Alert")

    def test_result_has_all_required_keys(self):
        """Ensures the result dict contains all required fields."""
        results = wet_stock_recon.reconcile_tank_wet_stock("2026-05-30", db_path=self.db)
        required_keys = [
            "date", "tank_id", "product_type", "opening_volume", "inbound_receipts",
            "meter_sales_volume", "expected_closing_volume", "evening_dip_mm",
            "actual_closing_volume", "variance", "shrinkage_limit_pct",
            "shrinkage_limit_liters", "status"
        ]
        for r in results:
            for key in required_keys:
                self.assertIn(key, r, msg=f"Key '{key}' missing from result for {r.get('tank_id')}")


class TestGenerateReconciliationReportData(unittest.TestCase):
    """Test generate_reconciliation_report_data returns a correct DataFrame."""

    def setUp(self):
        self.tmp = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
        self.tmp.close()
        self.db = self.tmp.name
        _make_test_db(self.db)

    def tearDown(self):
        os.unlink(self.db)

    def test_empty_when_no_data(self):
        df = wet_stock_recon.generate_reconciliation_report_data(db_path=self.db)
        self.assertTrue(df.empty)

    def test_returns_dataframe_with_correct_columns(self):
        """DataFrame must have the key columns after loading data."""
        conn = sqlite3.connect(self.db)
        conn.execute("""
            INSERT INTO stock_recon (date, hsd_closing_dip_liters, ms_closing_dip_liters)
            VALUES ('2026-05-30', 15000.0, 7000.0)
        """)
        conn.execute("""
            INSERT INTO daily_summary (date, total_hsd_liters, total_ms_liters)
            VALUES ('2026-05-30', 1000.0, 500.0)
        """)
        conn.commit()
        conn.close()

        df = wet_stock_recon.generate_reconciliation_report_data(db_path=self.db)
        self.assertFalse(df.empty)
        self.assertIn("tank_id", df.columns)
        self.assertIn("status", df.columns)
        self.assertIn("variance", df.columns)
        # Verify exactly 2 rows (one per tank)
        self.assertEqual(len(df), 2)

    def test_both_tanks_appear_in_report(self):
        """Both Tank_1_HSD and Tank_2_MS must appear in the output."""
        conn = sqlite3.connect(self.db)
        conn.execute("""
            INSERT INTO stock_recon (date, hsd_closing_dip_liters, ms_closing_dip_liters)
            VALUES ('2026-05-30', 15000.0, 7000.0)
        """)
        conn.execute("""
            INSERT INTO daily_summary (date, total_hsd_liters, total_ms_liters)
            VALUES ('2026-05-30', 1000.0, 500.0)
        """)
        conn.commit()
        conn.close()

        df = wet_stock_recon.generate_reconciliation_report_data(db_path=self.db)
        tank_ids = set(df["tank_id"].tolist())
        self.assertIn("Tank_1_HSD", tank_ids)
        self.assertIn("Tank_2_MS", tank_ids)


class TestExcelExportInventoryReconciliationTab(unittest.TestCase):
    """Test that export_db_to_excel generates the Inventory Reconciliation sheet."""

    def setUp(self):
        self.tmp_db = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
        self.tmp_db.close()
        self.db = self.tmp_db.name
        _make_test_db(self.db)

        self.tmp_excel = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        self.tmp_excel.close()
        self.excel = self.tmp_excel.name

        # Populate test data
        conn = sqlite3.connect(self.db)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS ledger_entries (
                entry_id INTEGER PRIMARY KEY AUTOINCREMENT,
                date TEXT,
                party_name TEXT,
                vehicle_wheel_no TEXT,
                amount REAL DEFAULT 0.0,
                type TEXT,
                remarks TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        conn.execute("""
            INSERT INTO stock_recon (date, hsd_closing_dip_liters, ms_closing_dip_liters)
            VALUES ('2026-05-30', 15000.0, 7000.0)
        """)
        conn.execute("""
            INSERT INTO daily_summary (date, total_hsd_liters, total_ms_liters)
            VALUES ('2026-05-30', 1000.0, 500.0)
        """)
        conn.commit()
        conn.close()

    def tearDown(self):
        os.unlink(self.db)
        if os.path.exists(self.excel):
            os.unlink(self.excel)

    def test_inventory_reconciliation_tab_present(self):
        """export_db_to_excel must produce an 'Inventory Reconciliation' sheet."""
        import exporter
        import importlib

        # Temporarily patch DB_PATH in exporter module and wet_stock_recon
        original_exporter_db = exporter.DB_PATH
        original_recon_db = wet_stock_recon.DB_PATH
        exporter.DB_PATH = self.db
        wet_stock_recon.DB_PATH = self.db
        os.makedirs(os.path.dirname(self.excel), exist_ok=True)

        try:
            exporter.export_db_to_excel(excel_path=self.excel)
        finally:
            exporter.DB_PATH = original_exporter_db
            wet_stock_recon.DB_PATH = original_recon_db

        from openpyxl import load_workbook
        wb = load_workbook(self.excel)
        self.assertIn("Inventory Reconciliation", wb.sheetnames,
                      "Inventory Reconciliation sheet not found in exported workbook")

        ws = wb["Inventory Reconciliation"]
        # Verify header row
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        self.assertIn("Tank ID", headers)
        self.assertIn("Status", headers)
        self.assertIn("Variance (Liters)", headers)


if __name__ == "__main__":
    unittest.main()
