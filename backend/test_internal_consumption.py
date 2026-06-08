#!/usr/bin/env python3
"""
Unit and Integration Test Suite for the Internal Fuel Consumption Tracker.
"""

import os
import sys
import sqlite3
import pandas as pd
import unittest
from openpyxl import load_workbook

# Add backend directory to path
BACKEND_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.append(BACKEND_DIR)

import internal_consumption
import reconciliation
import exporter
import price_registry
from crypto_vault import decrypt_field

class TestInternalConsumption(unittest.TestCase):
    def setUp(self):
        # Override DB paths to a sandbox test ledger database
        self.original_ic_db = internal_consumption.DB_PATH
        self.original_recon_db = reconciliation.DB_PATH
        self.original_exporter_db = exporter.DB_PATH
        self.original_pr_db = price_registry.DB_PATH
        
        self.test_db = os.path.join(BACKEND_DIR, "test_internal_consumption.db")
        internal_consumption.DB_PATH = self.test_db
        reconciliation.DB_PATH = self.test_db
        exporter.DB_PATH = self.test_db
        price_registry.DB_PATH = self.test_db
        
        if os.path.exists(self.test_db):
            os.remove(self.test_db)
            
        # 1. Initialize Tables in the sandbox database
        internal_consumption.init_internal_consumption_db(self.test_db)
        reconciliation.init_recon_db(self.test_db)
        price_registry.init_rates_db()
        
        # Initialize ledger_entries, daily_summary, and daily_ledger tables
        conn = sqlite3.connect(self.test_db)
        cursor = conn.cursor()
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS ledger_entries (
            entry_id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT,
            party_name TEXT,
            vehicle_wheel_no TEXT,
            amount REAL DEFAULT 0.0,
            type TEXT,
            remarks TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """)
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS daily_summary (
            date TEXT PRIMARY KEY,
            total_hsd_liters REAL DEFAULT 0.0,
            total_ms_liters REAL DEFAULT 0.0,
            total_cash_calculated REAL DEFAULT 0.0,
            total_credit_sales REAL DEFAULT 0.0,
            total_testing_deductions REAL DEFAULT 0.0,
            is_verified INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """)
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS daily_ledger (
            date TEXT PRIMARY KEY,
            cash_tender REAL DEFAULT 0.0,
            upi_tender REAL DEFAULT 0.0,
            paytm_transfers REAL DEFAULT 0.0,
            card_tender REAL DEFAULT 0.0,
            udhaar_sales REAL DEFAULT 0.0,
            raw_data TEXT
        )
        """)
        conn.commit()
        conn.close()

    def tearDown(self):
        # Restore DB paths
        internal_consumption.DB_PATH = self.original_ic_db
        reconciliation.DB_PATH = self.original_recon_db
        exporter.DB_PATH = self.original_exporter_db
        price_registry.DB_PATH = self.original_pr_db
        
        if os.path.exists(self.test_db):
            os.remove(self.test_db)

    def test_database_init(self):
        """Verifies that internal_consumption table is correctly created."""
        conn = sqlite3.connect(self.test_db)
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='internal_consumption'")
        self.assertIsNotNone(cursor.fetchone())
        conn.close()

    def test_record_internal_consumption_validation(self):
        """Verifies validation logic in record_internal_consumption."""
        # 1. Invalid product type
        with self.assertRaises(ValueError):
            internal_consumption.record_internal_consumption("2026-06-01", "INVALID", 10.0, "Testing", "Auth", self.test_db)
            
        # 2. Negative/zero liters
        with self.assertRaises(ValueError):
            internal_consumption.record_internal_consumption("2026-06-01", "HSD", -5.0, "Testing", "Auth", self.test_db)
        with self.assertRaises(ValueError):
            internal_consumption.record_internal_consumption("2026-06-01", "HSD", 0, "Testing", "Auth", self.test_db)

    def test_record_internal_consumption_success_with_fallback_rates(self):
        """Verifies recording fuel consumption and checking fallback rate & journal entries."""
        # Record HSD draw (should use fallback rate 94.27)
        res_hsd = internal_consumption.record_internal_consumption(
            "2026-06-01", "HSD", 50.0, "Station Generator", "Manager", self.test_db
        )
        self.assertEqual(res_hsd["status"], "success")
        self.assertEqual(res_hsd["applied_rate"], 94.27)
        self.assertEqual(res_hsd["total_financial_cost"], 50.0 * 94.27)
        
        # Record MS draw (should use fallback rate 106.31)
        res_ms = internal_consumption.record_internal_consumption(
            "2026-06-01", "MS", 20.0, "Pump Testing", "Operator", self.test_db
        )
        self.assertEqual(res_ms["status"], "success")
        self.assertEqual(res_ms["applied_rate"], 106.31)
        self.assertEqual(res_ms["total_financial_cost"], 20.0 * 106.31)
        
        # Verify SQLite raw records in internal_consumption table
        conn = sqlite3.connect(self.test_db)
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM internal_consumption ORDER BY entry_id ASC")
        rows = cursor.fetchall()
        self.assertEqual(len(rows), 2)
        
        # HSD row
        self.assertEqual(rows[0][1], "2026-06-01")
        self.assertEqual(rows[0][2], "HSD")
        self.assertEqual(rows[0][3], 50.0)
        self.assertEqual(rows[0][4], "Station Generator")
        self.assertEqual(rows[0][5], "Manager")
        
        # MS row
        self.assertEqual(rows[1][1], "2026-06-01")
        self.assertEqual(rows[1][2], "MS")
        self.assertEqual(rows[1][3], 20.0)
        self.assertEqual(rows[1][4], "Pump Testing")
        self.assertEqual(rows[1][5], "Operator")
        
        # Verify entries in ledger_entries with base64 decryption
        cursor.execute("SELECT party_name, amount, type, remarks FROM ledger_entries ORDER BY entry_id ASC")
        l_rows = cursor.fetchall()
        self.assertEqual(len(l_rows), 2)
        
        # HSD journal expense
        dec_party_hsd = decrypt_field(l_rows[0][0], str)
        dec_amount_hsd = decrypt_field(l_rows[0][1], float)
        self.assertEqual(dec_party_hsd, "Internal Fuel Consumption - Station Generator")
        self.assertEqual(dec_amount_hsd, 50.0 * 94.27)
        self.assertEqual(l_rows[0][2], "expense")
        self.assertIn("Station Generator", l_rows[0][3])
        
        # MS journal expense
        dec_party_ms = decrypt_field(l_rows[1][0], str)
        dec_amount_ms = decrypt_field(l_rows[1][1], float)
        self.assertEqual(dec_party_ms, "Internal Fuel Consumption - Pump Testing")
        self.assertEqual(dec_amount_ms, 20.0 * 106.31)
        self.assertEqual(l_rows[1][2], "expense")
        self.assertIn("Pump Testing", l_rows[1][3])
        
        conn.close()

    def test_record_internal_consumption_with_price_registry(self):
        """Verifies recording fuel consumption using prices loaded from price registry."""
        # Save custom rates to fuel_rates table
        conn = sqlite3.connect(self.test_db)
        cursor = conn.cursor()
        cursor.execute("INSERT INTO fuel_rates (date, hsd_rate, ms_rate) VALUES (?, ?, ?)", ("2026-06-02", 90.00, 105.00))
        conn.commit()
        conn.close()
        
        res = internal_consumption.record_internal_consumption(
            "2026-06-02", "HSD", 10.0, "Own Vehicle", "Supervisor", self.test_db
        )
        self.assertEqual(res["status"], "success")
        self.assertEqual(res["applied_rate"], 90.00)
        self.assertEqual(res["total_financial_cost"], 900.00)
        
        # Verify ledger entry cost
        conn = sqlite3.connect(self.test_db)
        cursor = conn.cursor()
        cursor.execute("SELECT amount FROM ledger_entries WHERE date = '2026-06-02'")
        amt_enc = cursor.fetchone()[0]
        self.assertEqual(decrypt_field(amt_enc, float), 900.00)
        conn.close()

    def test_reconciliation_integration_totalizer_filter(self):
        """Verifies that reconciliation expected book stock calculations subtract internal draws."""
        # 1. Setup mock daily summary nozzle sales
        conn = sqlite3.connect(self.test_db)
        cursor = conn.cursor()
        # Daily sales: 1000L HSD, 800L MS
        cursor.execute("""
            INSERT INTO daily_summary (date, total_hsd_liters, total_ms_liters, total_cash_calculated)
            VALUES ('2026-06-03', 1000.0, 800.0, 150000.0)
        """)
        # 2. Setup mock dip stocks: Opening dip 5000L, Receipts 2000L, Closing dip 5900L (for HSD)
        # Expected stock = Opening + Receipt - Sales - Internal = 5000 + 2000 - 1000 - Internal
        cursor.execute("""
            INSERT INTO stock_recon (
                date, hsd_opening_dip_liters, hsd_receipt_liters, hsd_closing_dip_liters,
                ms_opening_dip_liters, ms_receipt_liters, ms_closing_dip_liters
            ) VALUES ('2026-06-03', 5000.0, 2000.0, 5900.0, 4000.0, 1000.0, 4150.0)
        """)
        conn.commit()
        conn.close()
        
        # 3. Record internal draws for 2026-06-03: 50L HSD, 30L MS
        internal_consumption.record_internal_consumption("2026-06-03", "HSD", 50.0, "Generator", "M", self.test_db)
        internal_consumption.record_internal_consumption("2026-06-03", "MS", 30.0, "Test", "O", self.test_db)
        
        # 4. Run calculation
        recon_result = reconciliation.calculate_daily_variance("2026-06-03", db_path=self.test_db)
        
        # Verify stock calculations:
        # Expected HSD = 5000 + 2000 - 1000 - 50 = 5950
        # Expected MS = 4000 + 1000 - 800 - 30 = 4170
        self.assertEqual(recon_result["expected_hsd_book_stock"], 5950.0)
        self.assertEqual(recon_result["expected_ms_book_stock"], 4170.0)
        
        # HSD variance = Closing (5900) - Expected (5950) = -50
        # MS variance = Closing (4150) - Expected (4170) = -20
        self.assertEqual(recon_result["hsd_variance_liters"], -50.0)
        self.assertEqual(recon_result["ms_variance_liters"], -20.0)
        
        # Confirm internal consumption values returned in response
        self.assertEqual(recon_result["hsd_internal_consumption"], 50.0)
        self.assertEqual(recon_result["ms_internal_consumption"], 30.0)

    def test_excel_export_adjustments(self):
        """Verifies that both export engines generate the new premium styled Internal Consumption sheet."""
        # 1. Add sample data
        internal_consumption.record_internal_consumption("2026-06-01", "HSD", 50.0, "Generator", "Auth1", self.test_db)
        internal_consumption.record_internal_consumption("2026-06-02", "MS", 25.0, "Testing", "Auth2", self.test_db)
        
        excel_path_master = os.path.join(BACKEND_DIR, "test_master_export.xlsx")
        
        try:
            # 2. Test master export
            exporter.export_db_to_excel(excel_path_master)
            self.assertTrue(os.path.exists(excel_path_master))
            
            # Check sheet names using openpyxl
            wb = load_workbook(excel_path_master)
            self.assertIn("Internal Consumption", wb.sheetnames)
            ws = wb["Internal Consumption"]
            
            # Verify cumulative headers on row 1
            self.assertEqual(ws.cell(row=1, column=1).value, "Month")
            self.assertEqual(ws.cell(row=1, column=2).value, "Product Type")
            self.assertEqual(ws.cell(row=1, column=3).value, "Cumulative Liters Drawn (Liters)")
            self.assertEqual(ws.cell(row=1, column=4).value, "Total Transactions")
            
            # Verify data
            self.assertEqual(ws.cell(row=2, column=1).value, "2026-06")
            self.assertEqual(ws.cell(row=2, column=2).value, "HSD")
            self.assertEqual(float(ws.cell(row=2, column=3).value), 50.0)
            self.assertEqual(int(ws.cell(row=2, column=4).value), 1)
            
            # Find the Entry ID header (second table)
            second_header_row = None
            for r in range(2, ws.max_row + 1):
                if ws.cell(row=r, column=1).value == "Entry ID":
                    second_header_row = r
                    break
            self.assertIsNotNone(second_header_row)
            
            # Check banner row
            banner_row = second_header_row - 1
            self.assertEqual(ws.cell(row=banner_row, column=1).value, "DETAILED INTERNAL FUEL DRAW LOGS")
            # The cell style check
            self.assertEqual(ws.cell(row=banner_row, column=1).font.size, 12)
            self.assertTrue(ws.cell(row=banner_row, column=1).font.bold)
            
            # Check second header row style (Navy Blue fill, white text)
            self.assertIn("1F497D", ws.cell(row=second_header_row, column=1).fill.start_color.rgb)
            self.assertTrue(ws.cell(row=second_header_row, column=1).font.bold)
            self.assertIn("FFFFFF", ws.cell(row=second_header_row, column=1).font.color.rgb)
            
            # Verify detailed table data
            self.assertEqual(ws.cell(row=second_header_row + 1, column=2).value, "2026-06-02")
            self.assertEqual(ws.cell(row=second_header_row + 1, column=3).value, "MS")
            self.assertEqual(float(ws.cell(row=second_header_row + 1, column=4).value), 25.0)
            self.assertEqual(ws.cell(row=second_header_row + 1, column=5).value, "Testing")
            self.assertEqual(ws.cell(row=second_header_row + 1, column=6).value, "Auth2")
            
            wb.close()
            
            # 3. Test continuous accounting export
            excel_path_acct, csv_path_acct = exporter.generate_accounting_export("all")
            self.assertTrue(os.path.exists(excel_path_acct))
            
            wb_acct = load_workbook(excel_path_acct)
            self.assertIn("Internal Consumption", wb_acct.sheetnames)
            ws_acct = wb_acct["Internal Consumption"]
            self.assertEqual(ws_acct.cell(row=1, column=1).value, "Month")
            wb_acct.close()
            
            # Cleanup accounting exports
            if os.path.exists(excel_path_acct):
                os.remove(excel_path_acct)
            if os.path.exists(csv_path_acct):
                os.remove(csv_path_acct)
                
        finally:
            if os.path.exists(excel_path_master):
                os.remove(excel_path_master)

if __name__ == "__main__":
    unittest.main()
